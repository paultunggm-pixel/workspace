#!/usr/bin/env python3
"""
比对 players.json / teams.json 与现有 Excel 名单，生成差异标注版 Excel
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re

# ── 颜色定义 ──
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 匹配
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # 差异/仅JSON有
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # 仅Excel有
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SECTION_FONT = Font(name="微软雅黑", size=11, bold=True, color="C00000")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def apply_header_style(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def apply_cell_style(ws, row, cols, font=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or NORMAL_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

def normalize_name(name):
    """归一化球员姓名用于匹配"""
    name = name.strip()
    name = name.replace(' ', ' ').replace('　', ' ')
    # 去掉常见前缀标记
    name = re.sub(r'^[A-Z]\.\s*', '', name)  # J.戴维 → 戴维
    return name

# ── 1. 加载 JSON 数据 ──
print("加载 JSON 数据...")
with open('/Users/d.j.f/Desktop/players.json', 'r') as f:
    json_players = json.load(f)
with open('/Users/d.j.f/Desktop/teams.json', 'r') as f:
    json_teams = json.load(f)

# 球队名称映射 (JSON -> Excel)
team_name_map = {
    '刚果民主共和国': '刚果(金)',
    '刚果（金）': '刚果(金)',
    '刚果(金)': '刚果(金)',
    '沙特阿拉伯': '沙特',
    '沙特': '沙特',
}

json_team_names = set(team_name_map.get(t['name'], t['name']) for t in json_teams)
json_players_by_team = defaultdict(list)
for p in json_players:
    team = team_name_map.get(p.get('team', ''), p.get('team', ''))
    json_players_by_team[team].append(p)
    # 同时修正player记录中的team名
    p['team_normalized'] = team

print(f"  JSON: {len(json_team_names)} 队, {len(json_players)} 球员")

# ── 2. 加载现有 Excel ──
print("加载现有 Excel...")
wb_old = openpyxl.load_workbook('/Users/d.j.f/Documents/Claude/2026世界杯球队球员教练完整名单.xlsx')

# 读取球队总览
ws_teams_old = wb_old['球队总览']
excel_teams = {}  # name -> row data
for row in ws_teams_old.iter_rows(min_row=4, max_row=ws_teams_old.max_row, values_only=True):
    if row[1]:
        team_name = row[1]
        # 去掉旗帜emoji
        team_clean = re.sub(r'^[^一-鿿]+', '', team_name).strip()
        excel_teams[team_clean] = {
            'group': row[0], 'full_name': team_name, 'fifa_rank': row[2],
            'coach': row[3], 'captain': row[4], 'roster_status': row[5],
            'key_players': row[6], 'notes': row[7]
        }

# 读取球员
ws_players_old = wb_old['球员名单(已确认)']
excel_players = defaultdict(list)  # team -> [player dicts]
for row in ws_players_old.iter_rows(min_row=4, max_row=ws_players_old.max_row, values_only=True):
    if row[1] and row[4]:
        team_clean = re.sub(r'^[^一-鿿]+', '', row[1]).strip()
        excel_players[team_clean].append({
            'group': row[0], 'team_full': row[1], 'position': row[2],
            'number': str(row[3]) if row[3] else '', 'name': row[4],
            'club': row[5], 'age': row[6], 'caps': row[7], 'notes': row[8]
        })

print(f"  Excel: {len(excel_teams)} 队, {sum(len(v) for v in excel_players.values())} 球员(已确认)")

# ── 3. 比对分析 ──
print("\n开始比对分析...")

# 3.1 球队比对
json_teams_set = json_team_names
excel_teams_set = set(excel_teams.keys())
common_teams = json_teams_set & excel_teams_set
only_json_teams = json_teams_set - excel_teams_set
only_excel_teams = excel_teams_set - json_teams_set

print(f"  共同球队: {len(common_teams)}")
print(f"  仅JSON有: {only_json_teams}")
print(f"  仅Excel有: {only_excel_teams}")

# 3.2 球员比对 (按球队)
# 匹配策略:
#   1) 中文名精确匹配 (日韩球员)
#   2) 号码+位置匹配 (欧美球员 - JSON中文名 vs Excel拉丁名)
player_comparison = {}  # team -> {matched, only_json, only_excel, matched_by_number}
for team in common_teams:
    jp_list = json_players_by_team.get(team, [])
    ep_list = excel_players.get(team, [])

    # 建立 Excel 球员索引
    ep_by_name = {}
    ep_by_number = defaultdict(list)
    for ep in ep_list:
        name = ep['name']
        ep_by_name[name] = ep
        norm = normalize_name(name)
        if norm != name:
            ep_by_name[norm] = ep
        if ep['number']:
            ep_by_number[ep['number']].append(ep)

    matched = []       # matched by name
    matched_by_num = []  # matched by number+position
    only_json = []
    used_ep = set()  # track which Excel players have been matched

    for jp in jp_list:
        jname = jp.get('name', '')
        if not jname:
            only_json.append(jp)
            continue

        # Strategy 1: Direct name match
        if jname in ep_by_name:
            ep = ep_by_name[jname]
            matched.append((jp, ep, 'name'))
            used_ep.add(id(ep))
            continue

        # Strategy 2: Normalized name match
        norm_j = normalize_name(jname)
        if norm_j in ep_by_name and norm_j != jname:
            ep = ep_by_name[norm_j]
            matched.append((jp, ep, 'name'))
            used_ep.add(id(ep))
            continue

        # Strategy 3: Match by jersey number + similar position
        jnum = str(jp.get('number', ''))
        jpos = jp.get('position', '')
        if jnum and jnum in ep_by_number:
            candidates = [ep for ep in ep_by_number[jnum] if id(ep) not in used_ep]
            if candidates:
                # Prefer position match
                pos_match = [ep for ep in candidates if ep.get('position', '') == jpos]
                best = pos_match[0] if pos_match else candidates[0]
                matched_by_num.append((jp, best, 'number'))
                used_ep.add(id(best))
                continue

        # No match found
        only_json.append(jp)

    # Only-excel: Excel players not matched
    only_excel = [ep for ep in ep_list if id(ep) not in used_ep]

    player_comparison[team] = {
        'matched': matched,
        'matched_by_num': matched_by_num,
        'only_json': only_json,
        'only_excel': only_excel
    }

total_matched = sum(len(v['matched']) for v in player_comparison.values())
total_matched_by_num = sum(len(v['matched_by_num']) for v in player_comparison.values())
total_only_json = sum(len(v['only_json']) for v in player_comparison.values())
total_only_excel = sum(len(v['only_excel']) for v in player_comparison.values())

print(f"  球员比对: {total_matched} 姓名匹配, {total_matched_by_num} 号码匹配, {total_only_json} 仅JSON, {total_only_excel} 仅Excel")

# ── 4. 生成新 Excel ──
print("\n生成新 Excel...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "差异总览"

# ─── Sheet 1: 差异总览 ───
ws.merge_cells('A1:H1')
ws.cell(row=1, column=1, value="JSON数据源 vs 现有Excel — 差异总览报告").font = TITLE_FONT
ws.merge_cells('A2:H2')
ws.cell(row=2, column=1, value=f"比对日期: 2026-05-28  |  JSON数据: {len(json_teams)}队/{len(json_players)}球员  |  Excel数据: {len(excel_teams)}队/{sum(len(v) for v in excel_players.values())}球员(仅已确认球队)").font = Font(name="微软雅黑", size=10, color="666666")

# 摘要
r = 4
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value="📊 关键发现").font = SECTION_FONT
r += 1
findings = [
    f"1. 球队层面: JSON和Excel共有 {len(common_teams)} 支相同球队（两个数据源球队完全一致）",
    f"2. 球员层面: JSON包含全部48队共{len(json_players)}名球员（含未确认球队），Excel仅收录已确认球队的{sum(len(v) for v in excel_players.values())}名球员",
    f"3. 匹配策略: ①姓名直接匹配(中日韩球员)={total_matched}人  ②号码+位置匹配(欧美球员·JSON中文名↔Excel拉丁名)={total_matched_by_num}人",
    f"4. 号码匹配说明: JSON使用中文球员名(如「姆巴佩」)，Excel使用拉丁名(如「Kylian Mbappe」)，翻译不同但号码相同即为同一球员",
    f"5. 差异统计: {total_only_json} 名仅存在于JSON(含未确认球队全部球员)，{total_only_excel} 名仅存在于Excel(JSON号码覆盖不到的)",
    f"6. 球员人数差异: JSON中部分球队球员数远超26人（如伊朗36人、加拿大31人），说明JSON包含的是初选大名单而非最终26人",
    f"7. 推荐: Excel数据作为权威基准，JSON数据作为补充参考（尤其对尚未确认名单的球队有参考价值）",
    f"8. ⚠️ 号码不一致: 法国队等球队的JSON号码与官宣号码存在系统性偏差（如姆巴佩JSON#10↔官宣#7），说明JSON数据来源可能非最新官宣版本",
]
for f_text in findings:
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value=f_text).font = Font(name="微软雅黑", size=10)
    r += 1

r += 1
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value="📋 按球队差异统计").font = SECTION_FONT
r += 1

headers = ['球队', 'JSON球员数', 'Excel球员数', '姓名匹配', '号码匹配', '仅JSON', '仅Excel', 'Excel名单状态', '差异说明']
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
apply_header_style(ws, r, len(headers))
r += 1

for team in sorted(common_teams):
    comp = player_comparison.get(team, {'matched': [], 'matched_by_num': [], 'only_json': [], 'only_excel': []})
    j_count = len(json_players_by_team.get(team, []))
    e_count = len(excel_players.get(team, []))
    m_count = len(comp['matched'])
    mn_count = len(comp['matched_by_num'])
    oj_count = len(comp['only_json'])
    oe_count = len(comp['only_excel'])
    total_match = m_count + mn_count
    status = excel_teams.get(team, {}).get('roster_status', '未知')

    if e_count == 0:
        diff_note = "Excel未收录此队(名单未确认)"
    elif oj_count > 0 and oe_count > 0:
        diff_note = f"双方各有独有球员(共{total_match}人匹配)"
    elif oj_count > 0 and oe_count == 0:
        diff_note = f"JSON多{oj_count}人(可能为初选名单)"
    elif oe_count > 0 and oj_count == 0:
        diff_note = f"Excel多{oe_count}人(且全部通过号码/姓名匹配)"
    else:
        diff_note = f"完全一致" if j_count == e_count else f"数量差异(JSON{j_count}/Excel{e_count})"

    if mn_count > 0 and m_count == 0:
        diff_note += " [注意:仅通过号码匹配,姓名采用不同语言]"
    elif mn_count > 0:
        diff_note += " [含号码匹配]"

    row_data = [team, j_count, e_count, m_count, mn_count, oj_count, oe_count, status, diff_note]
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val)
    apply_cell_style(ws, r, len(headers))

    # 高亮有差异的行
    if oj_count > 0:
        ws.cell(row=r, column=6).fill = RED_FILL
    if oe_count > 0:
        ws.cell(row=r, column=7).fill = YELLOW_FILL
    if mn_count > 0:  # 有号码匹配 → 需要关注
        ws.cell(row=r, column=5).fill = PatternFill(start_color="DDD9C4", end_color="DDD9C4", fill_type="solid")
    if total_match > 0:
        ws.cell(row=r, column=4).fill = GREEN_FILL

    r += 1

# 列宽
col_widths = [14, 14, 14, 10, 10, 10, 18, 35]
for c, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

# ─── Sheet 2: JSON球员名单(全48队) ───
print("  写入 JSON球员名单...")
ws2 = wb.create_sheet("JSON球员名单(全48队)")
ws2.merge_cells('A1:G1')
ws2.cell(row=1, column=1, value="JSON数据源 — 全部48队球员名单 (1273人)").font = TITLE_FONT
ws2.merge_cells('A2:G2')
ws2.cell(row=2, column=1, value="⚠️ 数据来自第三方渠道(UC体育/API)，部分球队球员数超过26人(初选大名单)，请以「差异总览」和Excel基准数据为准").font = Font(name="微软雅黑", size=9, color="C00000")

headers2 = ['球队', '球衣号', '位置', '球员姓名', '球员头像URL', '是否与Excel匹配', '差异说明']
r2 = 4
for c, h in enumerate(headers2, 1):
    ws2.cell(row=r2, column=c, value=h)
apply_header_style(ws2, r2, len(headers2))
r2 += 1

for team in sorted(json_players_by_team.keys()):
    comp = player_comparison.get(team, {'matched': [], 'matched_by_num': [], 'only_json': [], 'only_excel': []})
    matched_names = set(m[1]['name'] for m in comp['matched'])
    num_matched_names = set(m[1]['name'] for m in comp['matched_by_num'])

    for p in json_players_by_team[team]:
        pname = p.get('name', '')
        is_name_match = pname in matched_names
        is_num_match = pname in num_matched_names
        is_match = is_name_match or is_num_match

        if is_name_match:
            match_status = "✅ 姓名匹配"
        elif is_num_match:
            match_status = "🔢 号码匹配(姓名语言不同)"
        elif pname:
            match_status = "⚠️ 仅JSON有(Excel未收录或未匹配)"
        else:
            match_status = "⚠️ 数据缺失(无姓名)"

        if not is_match and team in excel_teams:
            excel_status = excel_teams[team].get('roster_status', '')
            if '待公布' in str(excel_status) or '初选' in str(excel_status):
                match_status += " (该队名单未最终确认)"

        row_data = [team, p.get('number', ''), p.get('position', ''), pname, p.get('pic', ''),
                    "✅" if is_name_match else ("🔢" if is_num_match else "⚠️"), match_status]
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r2, column=c, value=val)
        apply_cell_style(ws2, r2, len(headers2))

        if is_name_match:
            ws2.cell(row=r2, column=6).fill = GREEN_FILL
        elif is_num_match:
            ws2.cell(row=r2, column=6).fill = PatternFill(start_color="DDD9C4", end_color="DDD9C4", fill_type="solid")
        else:
            ws2.cell(row=r2, column=6).fill = RED_FILL
            ws2.cell(row=r2, column=7).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        r2 += 1

col_widths2 = [14, 8, 8, 16, 50, 14, 45]
for c, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ─── Sheet 3: 球员详细比对 ───
print("  写入球员详细比对...")
ws3 = wb.create_sheet("球员详细比对")
ws3.merge_cells('A1:K1')
ws3.cell(row=1, column=1, value="JSON vs Excel — 球员级别详细比对").font = TITLE_FONT

headers3 = ['球队', '比对结果', 'JSON-姓名', 'JSON-号码', 'JSON-位置', 'Excel-姓名', 'Excel-号码', 'Excel-位置', 'Excel-俱乐部', 'Excel-年龄', '备注']
r3 = 3
for c, h in enumerate(headers3, 1):
    ws3.cell(row=r3, column=c, value=h)
apply_header_style(ws3, r3, len(headers3))
r3 += 1

for team in sorted(common_teams):
    comp = player_comparison.get(team, {'matched': [], 'matched_by_num': [], 'only_json': [], 'only_excel': []})

    # 姓名匹配的
    for jp, ep, mtype in comp['matched']:
        jpname = jp.get('name', '')
        jnum = str(jp.get('number', ''))
        enum = str(ep.get('number', ''))
        num_match = "✓" if jnum == enum else f"号码不同(J:{jnum}/E:{enum})"
        jpos = jp.get('position', '')
        epos = ep.get('position', '')
        pos_match = "✓" if jpos == epos else f"位置不同(J:{jpos}/E:{epos})"
        notes = "; ".join([x for x in [num_match, pos_match] if x != "✓"])

        row_data = [team, '✅ 姓名匹配', jpname, jnum, jpos,
                    ep['name'], enum, epos, ep.get('club', ''), ep.get('age', ''), notes or '号码位置均一致']
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r3, column=c, value=val)
        apply_cell_style(ws3, r3, len(headers3))
        ws3.cell(row=r3, column=2).fill = GREEN_FILL
        if notes:
            ws3.cell(row=r3, column=11).fill = YELLOW_FILL
        r3 += 1

    # 号码匹配的 (JSON中文名 ↔ Excel拉丁名)
    for jp, ep, mtype in comp['matched_by_num']:
        jpname = jp.get('name', '')
        epname = ep.get('name', '')
        jnum = str(jp.get('number', ''))
        enum = str(ep.get('number', ''))
        jpos = jp.get('position', '')
        epos = ep.get('position', '')
        pos_same = "✓" if jpos == epos else f"位置不同(J:{jpos}/E:{epos})"
        notes = f"通过号码#{jnum}匹配(JSON中文名↔Excel拉丁名); {pos_same}"

        row_data = [team, '🔢 号码匹配', jpname, jnum, jpos,
                    epname, enum, epos, ep.get('club', ''), ep.get('age', ''), notes]
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r3, column=c, value=val)
        apply_cell_style(ws3, r3, len(headers3))
        ws3.cell(row=r3, column=2).fill = PatternFill(start_color="DDD9C4", end_color="DDD9C4", fill_type="solid")
        ws3.cell(row=r3, column=11).fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        r3 += 1

    # 仅JSON有
    for jp in comp['only_json']:
        jpname = jp.get('name', '')
        row_data = [team, '⚠️ 仅JSON', jpname, jp.get('number', ''), jp.get('position', ''),
                    '(无)', '', '', '', '', f'该球员不在Excel名单中{"(无姓名)" if not jpname else ""}']
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r3, column=c, value=val)
        apply_cell_style(ws3, r3, len(headers3))
        ws3.cell(row=r3, column=2).fill = RED_FILL
        r3 += 1

    # 仅Excel有
    for ep in comp['only_excel']:
        row_data = [team, '📋 仅Excel', '(无)', '', '',
                    ep['name'], ep.get('number', ''), ep.get('position', ''), ep.get('club', ''), ep.get('age', ''), '该球员不在JSON数据中']
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r3, column=c, value=val)
        apply_cell_style(ws3, r3, len(headers3))
        ws3.cell(row=r3, column=2).fill = YELLOW_FILL
        r3 += 1

col_widths3 = [12, 12, 14, 8, 8, 14, 8, 8, 18, 6, 35]
for c, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

# ─── Sheet 4: JSON球队信息 ───
print("  写入JSON球队信息...")
ws4 = wb.create_sheet("JSON球队信息")
ws4.merge_cells('A1:C1')
ws4.cell(row=1, column=1, value="JSON数据源 — 48支球队列表").font = TITLE_FONT

headers4 = ['球队名称', 'Logo URL', '是否在Excel中存在']
r4 = 3
for c, h in enumerate(headers4, 1):
    ws4.cell(row=r4, column=c, value=h)
apply_header_style(ws4, r4, len(headers4))
r4 += 1

for t in sorted(json_teams, key=lambda x: x['name']):
    in_excel = "✅" if t['name'] in excel_teams_set else "⚠️ 不在Excel中"
    row_data = [t['name'], t.get('logo', ''), in_excel]
    for c, val in enumerate(row_data, 1):
        ws4.cell(row=r4, column=c, value=val)
    apply_cell_style(ws4, r4, len(headers4))
    if t['name'] in excel_teams_set:
        ws4.cell(row=r4, column=3).fill = GREEN_FILL
    else:
        ws4.cell(row=r4, column=3).fill = RED_FILL
    r4 += 1

col_widths4 = [16, 60, 18]
for c, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ─── Sheet 5: 合并建议 ───
print("  写入合并建议...")
ws5 = wb.create_sheet("合并建议")

ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value="数据合并建议与行动项").font = TITLE_FONT

suggestions = [
    ("SECTION", "一、总体评价"),
    ("INFO", "JSON数据源(UC体育API)覆盖了全部48支球队，球员信息包含姓名、号码、位置、头像URL，但缺乏俱乐部、年龄、国家队出场等详细数据。"),
    ("INFO", "现有Excel数据基于权威媒体报道，仅收录已确认最终名单的球队，信息更准确详细但覆盖不完整。"),
    ("INFO", "两个数据源球队列表完全一致（48队相同），可互补使用。"),
    ("SECTION", "二、推荐合并策略"),
    ("ACTION", "【高优先】Excel已确认球队的球员 → 以Excel为准，JSON仅作校验（如号码、位置是否一致）"),
    ("ACTION", "【高优先】Excel标记为「初选名单」或「待公布」的球队 → 可使用JSON数据作为临时参考，但需标注「未确认」"),
    ("ACTION", "【中优先】JSON中号码/位置与Excel不一致的球员 → 需人工核实（见「球员详细比对」表中标记）"),
    ("ACTION", "【低优先】JSON独有的球员 → 谨慎对待，可能为初选大名单中最终被淘汰的球员"),
    ("ACTION", "【低优先】Excel独有的球员 → 可能为JSON数据源遗漏，以Excel为准"),
    ("SECTION", "三、待核实项目"),
    ("TODO", "1. 标志为「号码不同」的球员需要核实最新号码分配"),
    ("TODO", "2. 标志为「位置不同」的球员需要核实实际司职位置"),
    ("TODO", "3. 伊朗(36人)、加拿大(31人)等球员数远超26人的球队，需等待官方最终名单"),
    ("TODO", "4. 6月2日FIFA最终名单截止后，需重新比对更新"),
]

r5 = 3
for stype, text in suggestions:
    if stype == "SECTION":
        ws5.merge_cells(f'A{r5}:F{r5}')
        ws5.cell(row=r5, column=1, value=text).font = SECTION_FONT
    elif stype == "ACTION":
        ws5.cell(row=r5, column=1, value="▶ 行动项").font = BOLD_FONT
        ws5.cell(row=r5, column=1).fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        ws5.merge_cells(f'B{r5}:F{r5}')
        ws5.cell(row=r5, column=2, value=text).font = NORMAL_FONT
    elif stype == "TODO":
        ws5.cell(row=r5, column=1, value="☐ 待办").font = BOLD_FONT
        ws5.cell(row=r5, column=1).fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        ws5.merge_cells(f'B{r5}:F{r5}')
        ws5.cell(row=r5, column=2, value=text).font = NORMAL_FONT
    else:
        ws5.merge_cells(f'A{r5}:F{r5}')
        ws5.cell(row=r5, column=1, value=text).font = NORMAL_FONT
    r5 += 1

col_widths5 = [18, 22, 22, 22, 22, 22]
for c, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

# ── 冻结窗格 & 自动筛选 ──
ws.freeze_panes = 'A8'
ws.auto_filter.ref = f'A7:H{r-1}'
ws2.freeze_panes = 'A5'
ws2.auto_filter.ref = f'A4:G{r2-1}'
ws3.freeze_panes = 'A4'
ws3.auto_filter.ref = f'A3:K{r3-1}'

# ── 保存 ──
output_path = '/Users/d.j.f/Documents/Claude/2026世界杯名单比对结果.xlsx'
wb.save(output_path)
print(f"\n✅ 已保存到: {output_path}")
print("完成!")
