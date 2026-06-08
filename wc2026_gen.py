import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# Team strength tiers
# ============================================================
tier1 = {"France", "Argentina", "Brazil", "England", "Spain", "Germany", "Portugal", "Netherlands"}
tier2 = {"Belgium", "Uruguay", "Croatia", "Colombia", "Mexico", "USA", "Japan", "South Korea",
         "Morocco", "Senegal", "Switzerland", "Austria", "Sweden", "Norway", "Ecuador"}
tier3 = {"Canada", "Turkey", "Scotland", "Egypt", "Ivory Coast", "Iran", "Tunisia", "Australia",
         "Paraguay", "Ghana", "Czechia", "Qatar", "Saudi Arabia", "Algeria", "Panama",
         "South Africa", "Bosnia & Herzegovina", "Iraq", "Jordan"}
tier4 = {"Cape Verde", "Curaçao", "Haiti", "DR Congo", "Uzbekistan", "New Zealand"}

order = ["France","Argentina","Brazil","England","Spain","Germany","Portugal","Netherlands",
         "Belgium","Uruguay","Croatia","Colombia","Mexico","USA","Japan","South Korea",
         "Morocco","Senegal","Switzerland","Austria","Sweden","Norway","Ecuador",
         "Canada","Turkey","Scotland","Egypt","Ivory Coast","Iran","Tunisia","Australia",
         "Paraguay","Ghana","Czechia","Qatar","Saudi Arabia","Algeria","Panama",
         "South Africa","Bosnia & Herzegovina","Iraq","Jordan",
         "Cape Verde","Curaçao","Haiti","DR Congo","Uzbekistan","New Zealand"]

def get_tier(team):
    if team in tier1: return 1
    if team in tier2: return 2
    if team in tier3: return 3
    if team in tier4: return 4
    return 3

def predict(team_a, team_b):
    ta = get_tier(team_a)
    tb = get_tier(team_b)
    if ta < tb:
        return team_a
    elif tb < ta:
        return team_b
    else:
        ra = order.index(team_a) if team_a in order else 99
        rb = order.index(team_b) if team_b in order else 99
        return team_a if ra < rb else team_b

def match_type(team_a, team_b):
    ta = get_tier(team_a)
    tb = get_tier(team_b)
    high = min(ta, tb)
    low = max(ta, tb)
    diff = low - high
    if diff == 0:
        if high <= 2:
            return "强强五五开"
        else:
            return "弱弱五五开"
    elif diff == 1:
        if high <= 2:
            return "碾压局"
        else:
            return "弱弱五五开"
    else:
        return "碾压局"

# ============================================================
# Full match data
# ============================================================
matches = [
    # Group Stage Matchday 1
    ("6月11日","15:00","Mexico","South Africa","A组第1轮","Estadio Azteca, 墨西哥城"),
    ("6月11日","22:00","South Korea","Czechia","A组第1轮","Estadio Akron, 萨波潘"),
    ("6月12日","15:00","Canada","Bosnia & Herzegovina","B组第1轮","BMO Field, 多伦多"),
    ("6月12日","21:00","USA","Paraguay","D组第1轮","SoFi Stadium, 洛杉矶"),
    ("6月13日","15:00","Qatar","Switzerland","B组第1轮","Levi's Stadium, 旧金山"),
    ("6月13日","18:00","Brazil","Morocco","C组第1轮","MetLife Stadium, 纽约/新泽西"),
    ("6月13日","21:00","Haiti","Scotland","C组第1轮","Gillette Stadium, 波士顿"),
    ("6月14日","12:00","Australia","Turkey","D组第1轮","BC Place, 温哥华"),
    ("6月14日","13:00","Germany","Curaçao","E组第1轮","NRG Stadium, 休斯顿"),
    ("6月14日","16:00","Netherlands","Japan","F组第1轮","AT&T Stadium, 达拉斯"),
    ("6月14日","19:00","Ivory Coast","Ecuador","E组第1轮","Lincoln Financial Field, 费城"),
    ("6月14日","22:00","Sweden","Tunisia","F组第1轮","Estadio BBVA, 蒙特雷"),
    ("6月15日","12:00","Spain","Cape Verde","H组第1轮","Mercedes-Benz Stadium, 亚特兰大"),
    ("6月15日","15:00","Belgium","Egypt","G组第1轮","Lumen Field, 西雅图"),
    ("6月15日","18:00","Saudi Arabia","Uruguay","H组第1轮","Hard Rock Stadium, 迈阿密"),
    ("6月15日","21:00","Iran","New Zealand","G组第1轮","SoFi Stadium, 洛杉矶"),
    ("6月16日","15:00","France","Senegal","I组第1轮","MetLife Stadium, 纽约/新泽西"),
    ("6月16日","18:00","Iraq","Norway","I组第1轮","Gillette Stadium, 波士顿"),
    ("6月16日","21:00","Argentina","Algeria","J组第1轮","Arrowhead Stadium, 堪萨斯城"),
    ("6月17日","12:00","Austria","Jordan","J组第1轮","Levi's Stadium, 旧金山"),
    ("6月17日","13:00","Portugal","DR Congo","K组第1轮","NRG Stadium, 休斯顿"),
    ("6月17日","16:00","England","Croatia","L组第1轮","AT&T Stadium, 达拉斯"),
    ("6月17日","19:00","Ghana","Panama","L组第1轮","BMO Field, 多伦多"),
    ("6月17日","22:00","Uzbekistan","Colombia","K组第1轮","Estadio Azteca, 墨西哥城"),

    # Group Stage Matchday 2
    ("6月18日","12:00","Czechia","South Africa","A组第2轮","Mercedes-Benz Stadium, 亚特兰大"),
    ("6月18日","15:00","Switzerland","Bosnia & Herzegovina","B组第2轮","SoFi Stadium, 洛杉矶"),
    ("6月18日","18:00","Canada","Qatar","B组第2轮","BC Place, 温哥华"),
    ("6月18日","21:00","Mexico","South Korea","A组第2轮","Estadio Akron, 萨波潘"),
    ("6月19日","15:00","USA","Australia","D组第2轮","Lumen Field, 西雅图"),
    ("6月19日","18:00","Scotland","Morocco","C组第2轮","Gillette Stadium, 波士顿"),
    ("6月19日","21:00","Brazil","Haiti","C组第2轮","Lincoln Financial Field, 费城"),
    ("6月20日","12:00","Turkey","Paraguay","D组第2轮","Levi's Stadium, 旧金山"),
    ("6月20日","13:00","Netherlands","Sweden","F组第2轮","NRG Stadium, 休斯顿"),
    ("6月20日","16:00","Germany","Ivory Coast","E组第2轮","BMO Field, 多伦多"),
    ("6月20日","20:00","Ecuador","Curaçao","E组第2轮","Arrowhead Stadium, 堪萨斯城"),
    ("6月21日","12:00","Tunisia","Japan","F组第2轮","Estadio BBVA, 蒙特雷"),
    ("6月21日","12:00","Spain","Saudi Arabia","H组第2轮","Mercedes-Benz Stadium, 亚特兰大"),
    ("6月21日","15:00","Belgium","Iran","G组第2轮","SoFi Stadium, 洛杉矶"),
    ("6月21日","18:00","Uruguay","Cape Verde","H组第2轮","Hard Rock Stadium, 迈阿密"),
    ("6月21日","21:00","New Zealand","Egypt","G组第2轮","BC Place, 温哥华"),
    ("6月22日","13:00","Argentina","Austria","J组第2轮","AT&T Stadium, 达拉斯"),
    ("6月22日","17:00","France","Iraq","I组第2轮","Lincoln Financial Field, 费城"),
    ("6月22日","20:00","Norway","Senegal","I组第2轮","MetLife Stadium, 纽约/新泽西"),
    ("6月22日","23:00","Jordan","Algeria","J组第2轮","Levi's Stadium, 旧金山"),
    ("6月23日","13:00","Portugal","Uzbekistan","K组第2轮","NRG Stadium, 休斯顿"),
    ("6月23日","16:00","England","Ghana","L组第2轮","Gillette Stadium, 波士顿"),
    ("6月23日","19:00","Panama","Croatia","L组第2轮","BMO Field, 多伦多"),
    ("6月23日","22:00","Colombia","DR Congo","K组第2轮","Estadio Akron, 萨波潘"),

    # Group Stage Matchday 3
    ("6月24日","15:00","Switzerland","Canada","B组第3轮","BC Place, 温哥华"),
    ("6月24日","15:00","Bosnia & Herzegovina","Qatar","B组第3轮","Lumen Field, 西雅图"),
    ("6月24日","18:00","Scotland","Brazil","C组第3轮","Hard Rock Stadium, 迈阿密"),
    ("6月24日","18:00","Morocco","Haiti","C组第3轮","Mercedes-Benz Stadium, 亚特兰大"),
    ("6月24日","21:00","Czechia","Mexico","A组第3轮","Estadio Azteca, 墨西哥城"),
    ("6月24日","21:00","South Africa","South Korea","A组第3轮","Estadio BBVA, 蒙特雷"),
    ("6月25日","16:00","Curaçao","Ivory Coast","E组第3轮","Lincoln Financial Field, 费城"),
    ("6月25日","16:00","Ecuador","Germany","E组第3轮","MetLife Stadium, 纽约/新泽西"),
    ("6月25日","19:00","Japan","Sweden","F组第3轮","AT&T Stadium, 达拉斯"),
    ("6月25日","19:00","Tunisia","Netherlands","F组第3轮","Arrowhead Stadium, 堪萨斯城"),
    ("6月25日","22:00","Turkey","USA","D组第3轮","SoFi Stadium, 洛杉矶"),
    ("6月25日","22:00","Paraguay","Australia","D组第3轮","Levi's Stadium, 旧金山"),
    ("6月26日","15:00","Norway","France","I组第3轮","Gillette Stadium, 波士顿"),
    ("6月26日","15:00","Senegal","Iraq","I组第3轮","BMO Field, 多伦多"),
    ("6月26日","20:00","Cape Verde","Saudi Arabia","H组第3轮","NRG Stadium, 休斯顿"),
    ("6月26日","20:00","Uruguay","Spain","H组第3轮","Estadio Akron, 萨波潘"),
    ("6月26日","23:00","Egypt","Iran","G组第3轮","Lumen Field, 西雅图"),
    ("6月26日","23:00","New Zealand","Belgium","G组第3轮","BC Place, 温哥华"),
    ("6月27日","17:00","Panama","England","L组第3轮","MetLife Stadium, 纽约/新泽西"),
    ("6月27日","17:00","Croatia","Ghana","L组第3轮","Lincoln Financial Field, 费城"),
    ("6月27日","19:30","Colombia","Portugal","K组第3轮","Hard Rock Stadium, 迈阿密"),
    ("6月27日","19:30","DR Congo","Uzbekistan","K组第3轮","Mercedes-Benz Stadium, 亚特兰大"),
    ("6月27日","22:00","Algeria","Austria","J组第3轮","Arrowhead Stadium, 堪萨斯城"),
    ("6月27日","22:00","Jordan","Argentina","J组第3轮","AT&T Stadium, 达拉斯"),

    # Round of 32
    ("6月28日","15:00","A组第2","B组第2","32强赛","SoFi Stadium, 洛杉矶"),
    ("6月29日","13:00","C组第1","F组第2","32强赛","NRG Stadium, 休斯顿"),
    ("6月29日","16:30","E组第1","最佳第3","32强赛","Gillette Stadium, 波士顿"),
    ("6月29日","21:00","F组第1","C组第2","32强赛","Estadio BBVA, 蒙特雷"),
    ("6月30日","13:00","E组第2","I组第2","32强赛","AT&T Stadium, 达拉斯"),
    ("6月30日","17:00","I组第1","最佳第3","32强赛","MetLife Stadium, 纽约/新泽西"),
    ("6月30日","21:00","A组第1","最佳第3","32强赛","Estadio Azteca, 墨西哥城"),
    ("7月1日","12:00","L组第1","最佳第3","32强赛","Mercedes-Benz Stadium, 亚特兰大"),
    ("7月1日","16:00","G组第1","最佳第3","32强赛","Lumen Field, 西雅图"),
    ("7月1日","20:00","D组第1","最佳第3","32强赛","Levi's Stadium, 旧金山"),
    ("7月2日","15:00","H组第1","J组第2","32强赛","SoFi Stadium, 洛杉矶"),
    ("7月2日","19:00","K组第2","L组第2","32强赛","BMO Field, 多伦多"),
    ("7月2日","23:00","B组第1","最佳第3","32强赛","BC Place, 温哥华"),
    ("7月3日","14:00","D组第2","G组第2","32强赛","AT&T Stadium, 达拉斯"),
    ("7月3日","18:00","J组第1","H组第2","32强赛","Hard Rock Stadium, 迈阿密"),
    ("7月3日","21:30","K组第1","最佳第3","32强赛","Arrowhead Stadium, 堪萨斯城"),

    # Round of 16
    ("7月4日","13:00","M73胜者","M75胜者","16强赛","NRG Stadium, 休斯顿"),
    ("7月4日","17:00","M74胜者","M77胜者","16强赛","Lincoln Financial Field, 费城"),
    ("7月5日","16:00","M76胜者","M78胜者","16强赛","MetLife Stadium, 纽约/新泽西"),
    ("7月5日","20:00","M79胜者","M80胜者","16强赛","Estadio Azteca, 墨西哥城"),
    ("7月6日","15:00","M83胜者","M84胜者","16强赛","AT&T Stadium, 达拉斯"),
    ("7月6日","20:00","M81胜者","M82胜者","16强赛","Lumen Field, 西雅图"),
    ("7月7日","12:00","M86胜者","M88胜者","16强赛","Mercedes-Benz Stadium, 亚特兰大"),
    ("7月7日","16:00","M85胜者","M87胜者","16强赛","BC Place, 温哥华"),

    # Quarter-finals
    ("7月9日","16:00","M89胜者","M90胜者","1/4决赛","Gillette Stadium, 波士顿"),
    ("7月10日","15:00","M93胜者","M94胜者","1/4决赛","SoFi Stadium, 洛杉矶"),
    ("7月11日","17:00","M91胜者","M92胜者","1/4决赛","Hard Rock Stadium, 迈阿密"),
    ("7月11日","21:00","M95胜者","M96胜者","1/4决赛","Arrowhead Stadium, 堪萨斯城"),

    # Semi-finals
    ("7月14日","15:00","M97胜者","M98胜者","半决赛","AT&T Stadium, 达拉斯"),
    ("7月15日","15:00","M99胜者","M100胜者","半决赛","Mercedes-Benz Stadium, 亚特兰大"),

    # Third place
    ("7月18日","17:00","M101负者","M102负者","季军赛","Hard Rock Stadium, 迈阿密"),

    # Final
    ("7月19日","15:00","M101胜者","M102胜者","决赛","MetLife Stadium, 纽约/新泽西"),
]

# ============================================================
# Create workbook
# ============================================================
wb = openpyxl.Workbook()

# --- Sheet 1: Full Schedule ---
ws1 = wb.active
ws1.title = "完整赛程"

header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ko_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
final_fill = PatternFill(start_color="F4B4C2", end_color="F4B4C2", fill_type="solid")
cell_font = Font(name="Microsoft YaHei", size=10)
bold_font = Font(name="Microsoft YaHei", size=10, bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

headers = ["序号","日期","开球时间(ET)","主队","客队","阶段","预测胜方","比赛类型","比赛场地"]

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

widths = [6, 12, 16, 22, 22, 18, 22, 14, 42]
for col, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = w

for i, m in enumerate(matches):
    row = i + 2
    date, time_, team_a, team_b, stage, venue = m

    is_knockout = any(kw in stage for kw in ["32强","16强","1/4","半","季军","决赛"])
    is_final = stage == "决赛"

    if is_knockout:
        predicted = "待定"
        mtype = "待定"
    else:
        predicted = predict(team_a, team_b)
        mtype = match_type(team_a, team_b)

    values = [i+1, date, time_, team_a, team_b, stage, predicted, mtype, venue]

    for col, val in enumerate(values, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = cell_font
        cell.alignment = center if col != 9 else left_align
        cell.border = thin_border

        if is_final:
            cell.fill = final_fill
            cell.font = bold_font
        elif is_knockout:
            cell.fill = ko_fill
        elif i < 72:
            cell.fill = group_fill

ws1.freeze_panes = "A2"

# --- Sheet 2: Group Predictions ---
ws2 = wb.create_sheet("分组预判汇总")

group_headers = ["组别","球队","档次","预测小组排名","备注"]
for col, h in enumerate(group_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

groups = [
    ("A", [("Mexico",1),("South Africa",3),("South Korea",2),("Czechia",3)]),
    ("B", [("Canada",3),("Bosnia & Herzegovina",3),("Qatar",3),("Switzerland",2)]),
    ("C", [("Brazil",1),("Morocco",2),("Haiti",4),("Scotland",3)]),
    ("D", [("USA",2),("Paraguay",3),("Australia",3),("Turkey",3)]),
    ("E", [("Germany",1),("Curaçao",4),("Ivory Coast",3),("Ecuador",2)]),
    ("F", [("Netherlands",1),("Japan",2),("Sweden",2),("Tunisia",3)]),
    ("G", [("Belgium",2),("Egypt",3),("Iran",3),("New Zealand",4)]),
    ("H", [("Spain",1),("Cape Verde",4),("Saudi Arabia",3),("Uruguay",2)]),
    ("I", [("France",1),("Senegal",2),("Iraq",3),("Norway",2)]),
    ("J", [("Argentina",1),("Algeria",3),("Austria",2),("Jordan",3)]),
    ("K", [("Portugal",1),("DR Congo",4),("Uzbekistan",4),("Colombia",2)]),
    ("L", [("England",1),("Croatia",2),("Ghana",3),("Panama",3)]),
]

tier_names = {1:"顶级强队",2:"二档强队",3:"中游球队",4:"弱旅"}

row = 2
for group_name, teams in groups:
    ranked = sorted(teams, key=lambda x: (x[1], order.index(x[0]) if x[0] in order else 99))
    for rank, (team, tier) in enumerate(ranked, 1):
        note = ""
        if rank == 1: note = "预测小组第1出线"
        elif rank == 2: note = "预测小组第2出线"
        elif rank == 3: note = "有望竞争最佳第3"
        values = [group_name, team, tier_names[tier], f"第{rank}名", note]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.alignment = center
            cell.border = thin_border
            if tier == 1:
                cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="C00000")
        row += 1

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 24
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 24
ws2.freeze_panes = "A2"

# --- Sheet 3: Stats ---
ws3 = wb.create_sheet("统计汇总")

stats = [
    ("预测统计",""),
    ("",""),
    ("小组赛总场次","72"),
    ("强强五五开","待计算"),
    ("弱弱五五开","待计算"),
    ("碾压局","待计算"),
    ("淘汰赛场次","32"),
    ("总赛场次","104"),
]

# Count match types
count_strong = 0
count_weak = 0
count_stomp = 0
for m in matches[:72]:
    _, _, team_a, team_b, _, _ = m
    mt = match_type(team_a, team_b)
    if mt == "强强五五开": count_strong += 1
    elif mt == "弱弱五五开": count_weak += 1
    else: count_stomp += 1

stats[3] = ("强强五五开", f"{count_strong} 场")
stats[4] = ("弱弱五五开", f"{count_weak} 场")
stats[5] = ("碾压局", f"{count_stomp} 场")

for i, (k, v) in enumerate(stats, 1):
    c1 = ws3.cell(row=i, column=1, value=k)
    c2 = ws3.cell(row=i, column=2, value=v)
    if i == 1:
        c1.font = Font(name="Microsoft YaHei", size=14, bold=True)
    elif k:
        c1.font = Font(name="Microsoft YaHei", size=10, bold=True)
        c2.font = Font(name="Microsoft YaHei", size=10)
    c1.alignment = left_align
    c2.alignment = center

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 16

# --- Sheet 4: Legend ---
ws4 = wb.create_sheet("图例说明")
legends = [
    ("图例说明",""),
    ("",""),
    ("比赛类型说明：",""),
    ("强强五五开","两队均为顶级或二档强队，实力接近，胜负难料"),
    ("弱弱五五开","两队均为中游或弱旅，实力接近"),
    ("碾压局","两队实力差距明显（差两个档次或以上），强队大概率获胜"),
    ("",""),
    ("球队档次说明：",""),
    ("顶级强队 (Tier 1)","法国、阿根廷、巴西、英格兰、西班牙、德国、葡萄牙、荷兰"),
    ("二档强队 (Tier 2)","比利时、乌拉圭、克罗地亚、哥伦比亚、墨西哥、美国、日本、韩国、摩洛哥、塞内加尔、瑞士、奥地利、瑞典、挪威、厄瓜多尔"),
    ("中游球队 (Tier 3)","加拿大、土耳其、苏格兰、埃及、科特迪瓦、伊朗、突尼斯、澳大利亚、巴拉圭、加纳、捷克、卡塔尔、沙特、阿尔及利亚、巴拿马、南非、波黑、伊拉克、约旦"),
    ("弱旅 (Tier 4)","佛得角、库拉索、海地、刚果(金)、乌兹别克斯坦、新西兰"),
    ("",""),
    ("免责声明：","本预判仅基于纸面实力和FIFA排名做出的主观推测，仅供娱乐参考。足球是圆的，一切皆有可能！"),
]

for i, (k, v) in enumerate(legends, 1):
    c1 = ws4.cell(row=i, column=1, value=k)
    c2 = ws4.cell(row=i, column=2, value=v)
    if i == 1:
        c1.font = Font(name="Microsoft YaHei", size=14, bold=True)
    elif k and not v:
        c1.font = Font(name="Microsoft YaHei", size=11, bold=True)
    else:
        c1.font = Font(name="Microsoft YaHei", size=10, bold=True)
        c2.font = Font(name="Microsoft YaHei", size=10)
    c1.alignment = left_align
    c2.alignment = left_align

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 90

# Save
output_path = "/Users/d.j.f/Documents/Claude/2026世界杯完整赛程预判.xlsx"
wb.save(output_path)

print(f"✅ 文件已保存至: {output_path}")
print(f"总比赛场次: {len(matches)}")
print(f"  - 小组赛: 72 场")
print(f"  - 淘汰赛: 32 场")
print(f"小组赛预判分布:")
print(f"  - 强强五五开: {count_strong} 场")
print(f"  - 弱弱五五开: {count_weak} 场")
print(f"  - 碾压局: {count_stomp} 场")
