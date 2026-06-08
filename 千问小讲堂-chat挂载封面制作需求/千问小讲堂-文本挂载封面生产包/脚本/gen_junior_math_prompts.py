#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""初中数学底表 → 4-case 生图指令（A/B/C/D 各一列，无钩子/学名/标签）。

用法：
  python3 gen_junior_math_prompts.py \\
    --csv "/path/初中数学-生产运营底表.csv" \\
    --out outputs/初中数学-生图指令-4cases.xlsx
"""
import argparse
import csv
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
except ImportError:
    Workbook = None

from junior_math_lib import NUM_CASES, gen_four_image_prompts, row_from_csv

CASE_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
CASE_COLS = [f"{c}生图指令" for c in ("A", "B", "C", "D")]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--also-csv", action="store_true", help="同时输出同路径 .csv")
    args = ap.parse_args()

    if Workbook is None:
        print("需要 openpyxl: pip3 install openpyxl", file=sys.stderr)
        sys.exit(1)

    with open(args.csv, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        print("CSV 无数据", file=sys.stderr)
        sys.exit(1)

    orig_keys = list(rows[0].keys())
    header = orig_keys + CASE_COLS

    wb = Workbook()
    ws = wb.active
    ws.title = "初中数学生图"
    ws.append(header)
    case_start = len(orig_keys)

    out_rows = []
    n = 0
    for d in rows:
        if not (d.get("课文标题") or "").strip():
            continue
        if args.limit and n >= args.limit:
            break
        r = row_from_csv(d)
        cases = gen_four_image_prompts(r)
        line = [d.get(k, "") for k in orig_keys]
        line += [cases[i]["prompt"] for i in range(NUM_CASES)]
        ws.append(line)
        out_rows.append(dict(zip(header, line)))
        n += 1

    for col_idx in range(case_start + 1, case_start + NUM_CASES + 1):
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).fill = CASE_FILL

    os.makedirs(os.path.dirname(os.path.abspath(args.out)) or ".", exist_ok=True)
    wb.save(args.out)
    print(f"完成 {n} 条 → {args.out}")

    if args.also_csv or args.out.lower().endswith(".csv"):
        csv_path = args.out if args.out.lower().endswith(".csv") else (
            os.path.splitext(args.out)[0] + ".csv"
        )
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=header)
            w.writeheader()
            w.writerows(out_rows)
        print(f"CSV: {csv_path}")


if __name__ == "__main__":
    main()
