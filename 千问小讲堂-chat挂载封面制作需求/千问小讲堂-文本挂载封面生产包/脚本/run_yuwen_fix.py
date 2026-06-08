#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""问题语文封面重压：style3 + 底部自适应压暗带。

读取 语文类封面有问题.xlsx，逐行：
  背景 = 『url·修改版』(有则优先) 否则 『url确认版』
  学名 = 『学名确认版』
  渲染 style3 → 压 JPG(≤50K) → 上传 OSS → 回写『修复压缩后url』
"""
import hashlib
import os
import sys

import openpyxl

import render_covers as rc

XLSX = "/Users/xiyunmeng/Desktop/语文类封面有问题.xlsx"
OUT = "output/yuwen_fix/jpg"
CACHE = "output/yuwen_fix/_bg_cache"

# 1-based 列号
C_NAME = 13   # 学名确认版
C_SRC = 14    # url确认版
C_FIX = 18    # 修复压缩后url
C_MODBG = 19  # url·修改版


def md5(s):
    return hashlib.md5(s.encode()).hexdigest()


def main():
    os.makedirs(OUT, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    cache_urls = {}   # (bg_url, name) -> jpg_url
    ok = err = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, C_NAME).value
        name = (name or "").strip()
        if not name:
            continue
        modbg = (ws.cell(r, C_MODBG).value or "").strip()
        src = (ws.cell(r, C_SRC).value or "").strip()
        bg_url = modbg or src
        if not bg_url:
            print(f"  [行{r}] 跳过：无背景 {name}", file=sys.stderr)
            continue
        key = (bg_url, name)
        try:
            if key in cache_urls:
                jpg_url = cache_urls[key]
            else:
                bg = rc.fetch_bg(bg_url, CACHE)
                img = rc.render(3, bg, name)
                jpg_path = os.path.join(OUT, md5(bg_url + name) + ".jpg")
                rc.save(img, jpg_path)
                jpg_url = rc.upload_to_oss(jpg_path)
                cache_urls[key] = jpg_url
            ws.cell(r, C_FIX).value = jpg_url
            ok += 1
            print(f"  [行{r}] OK {name} -> {jpg_url}")
        except Exception as e:  # noqa: BLE001
            err += 1
            print(f"  [行{r}] 失败 {name}: {e}", file=sys.stderr)

    wb.save(XLSX)
    print(f"完成 {ok} 行（失败 {err}）；已回写 {XLSX}")


if __name__ == "__main__":
    main()
