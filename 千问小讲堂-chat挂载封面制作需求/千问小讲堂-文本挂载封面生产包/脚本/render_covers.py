#!/usr/bin/env python3
"""Qianwen lecture-cover renderer (3 styles).

Renders one of three cover layouts adapted from the Figma file
`小讲堂挂载封面优化`:

    style=1  彩虹里        — top-left title/subtitle, rounded pill tag, drape badge (#0293D2)
    style=2  数字1·百倍超能力 — top-left bold title/subtitle, rectangular tag, square-bottom badge (#09692A)
    style=3  揠苗助长        — centered song-style title + subtitle, plain badge (#457447)

The user supplies:
    --bg        path to background image (auto cover-fit to 144x192 frame)
    --title     primary title text
    --subtitle  secondary line (style 1/2/3)
    --tag       tag/category text (style 1 & 2 only)

Frame is 144x192 design-px, output at 4x = 576x768 PNG.

Usage
-----
Single render:
    python render_covers.py --style 1 --bg ./bg.jpg \\
        --title "彩虹里" --subtitle "藏着几种颜色" --tag "学龄前★生活发现" \\
        --output ./out.png

Batch from xlsx (columns: style, bg, title, subtitle, tag, output):
    python render_covers.py --xlsx ./covers.xlsx --out-dir ./output
"""
import argparse
import colorsys
import csv
import math
import os
import re
import subprocess
import sys
import uuid

from PIL import Image, ImageDraw, ImageFilter, ImageFont

HERE = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(HERE, "assets")

SCALE = 4
DESIGN_W, DESIGN_H = 144, 192
W, H = DESIGN_W * SCALE, DESIGN_H * SCALE

# Brand fonts (bundled under assets/). 单 ttf 字体，index 恒为 0。
FONT_YUANHEI = os.path.join(ASSETS_DIR, "ZaoziGongfangYuanhei-Regular.ttf")   # 小低钩子
FONT_PUHUI_BOLD = os.path.join(ASSETS_DIR, "Alibaba-PuHuiTi-Bold.ttf")        # 小高钩子 + 徽标
FONT_PUHUI_MEDIUM = os.path.join(ASSETS_DIR, "Alibaba-PuHuiTi-Medium.ttf")    # 标签
FONT_XINRENWENSONG = os.path.join(ASSETS_DIR, "HYXinRenWenSong65W.ttf")       # style3 学名（标题 65W）
FONT_XINRENWENSONG_W = os.path.join(ASSETS_DIR, "HYXinRenWenSongW.ttf")        # 成语钩子（常规 W）

WHITE = "#FFFFFF"

# ---- 徽标文字 ----
BADGE_TEXT = "千问小讲堂"
BADGE_TEXT_PX = 11.52
BADGE_TEXT_LEFT = 12.0        # 设计稿文本框 X
# 设计稿文本框 Y=5.76、H=9 → 垂直中心 10.26dp，墨迹按此居中
BADGE_TEXT_CY = 5.76 + 9.0 / 2
THEME_DARKEN = 0.70          # 徽标底色 = 底图顶部主色 × 0.70

# ---- 钩子（标题）排版 ----
# 单一文本，按宽度贪心换行（最多 2 行），固定字号，仅 2 行仍溢出时轻微缩字。
TITLE_X = 8.64
TITLE_Y = 28.8               # 首行顶
TITLE_MAX_W = 128.0          # 可用宽度（6 字钩子可一行）
TITLE_LOW_PX = 20.0          # 小低固定字号
TITLE_LOW_TRACK = 1.44       # 小低字间距
TITLE_LOW_LH = 24.1 / 22.99  # 小低行高比 ≈1.048
TITLE_HIGH_PX = 20.0         # 小高字号（沿用小低；待 Figma 规格确认）
TITLE_HIGH_TRACK = 0.0
TITLE_HIGH_LH = 1.16
TITLE_MIN_PX = 12.0          # 极端长钩子的缩字下限

# ---- 标签 ----
TAG_PX = 9.12
TAG_H = 16.0                 # 胶囊高（对齐设计稿）
TAG_PAD_X = 6.75            # 左右内边距（对齐设计稿）
TAG_GAP = 7.0                # 标签距钩子块底部

# ---- style3 学名（语文成语类）----
TITLE3_PX = 22.99            # HYXinRenWenSong 65W
TITLE3_TRACK = 1.44          # 字间距
TITLE3_LH = 24.1 / 22.99     # 行高比
TITLE3_CY = 150.0            # 学名块垂直中心（较设计稿下移，落在下三分之一）
TITLE3_MAX_W = 128.0         # 学名换行可用宽（适当放宽，尽量一行）
TITLE3_MIN_PX = 16.0         # 超长学名（需换行时）缩字下限
TITLE3_ONELINE_MIN_PX = 18.0  # 一行可接受的最小字号（小幅缩字仍尽量一行）
SCRIM_START = 0.46           # 底部压暗带起点（占画面高比例）
SCRIM_ALPHA = 0.55           # 压暗带最大不透明度
SCRIM_DARKEN = 0.40          # 压暗带颜色 = 底部主色 × 该系数


def sp(x):
    """Scale a design-px value to output-px (float-safe)."""
    return int(round(x * SCALE))


def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i : i + 2], 16) for i in (0, 2, 4))


def load_font(path, design_px, index=0):
    return ImageFont.truetype(path, sp(design_px), index=index)


def theme_color(img):
    """取底图顶部留白区主色并加深，作为徽标底色（随底图自适应）。"""
    strip = img.convert("RGB").crop((0, 0, W, sp(26)))
    small = strip.resize((1, 1), Image.LANCZOS)
    r, g, b = small.getpixel((0, 0))
    return (int(r * THEME_DARKEN), int(g * THEME_DARKEN), int(b * THEME_DARKEN))


def cover_resize(img, target_w, target_h):
    """Scale + center-crop `img` to exactly fill (target_w, target_h)."""
    sw, sh = img.size
    scale = max(target_w / sw, target_h / sh)
    nw, nh = max(1, int(math.ceil(sw * scale))), max(1, int(math.ceil(sh * scale)))
    resized = img.resize((nw, nh), Image.LANCZOS)
    left = (nw - target_w) // 2
    top = (nh - target_h) // 2
    return resized.crop((left, top, left + target_w, top + target_h))


def load_bg(path):
    bg = Image.open(path).convert("RGB")
    return cover_resize(bg, W, H)


def card_base(bg_path):
    base = load_bg(bg_path)
    return base.convert("RGBA")


def cubic_bezier_points(p0, p1, p2, p3, steps):
    out = []
    for i in range(1, steps + 1):
        t = i / steps
        u = 1 - t
        bx = u**3 * p0[0] + 3 * u**2 * t * p1[0] + 3 * u * t**2 * p2[0] + t**3 * p3[0]
        by = u**3 * p0[1] + 3 * u**2 * t * p1[1] + 3 * u * t**2 * p2[1] + t**3 * p3[1]
        out.append((bx, by))
    return out


# Supersample factor for badge/pill shapes. PIL's polygon/rect drawing has no
# native anti-aliasing, so we draw at SS× resolution and downsample with
# LANCZOS to smooth the curved edges (drape bottom, pill outline, etc.).
SS = 8


def _ss_buffer(w_px, h_px):
    return Image.new("RGBA", (w_px * SS, h_px * SS), (0, 0, 0, 0))


def _ss_paste(img, layer, x_px, y_px):
    w = layer.size[0] // SS
    h = layer.size[1] // SS
    smooth = layer.resize((w, h), Image.LANCZOS)
    img.paste(smooth, (x_px, y_px), smooth)


def draw_drape_badge(img, x_dp, y_dp, w_dp, h_dp, color):
    """Draw the style-1 drape/banner — flat top, downward-arc bottom (anti-aliased)."""
    bw, bh = sp(w_dp), sp(h_dp)
    layer = _ss_buffer(bw, bh)

    VBW, VBH = 63.36, 22.9955
    sx = bw * SS / VBW
    sy = bh * SS / VBH

    def P(px, py):
        return (px * sx, py * sy)

    pts = [P(0, 0), P(VBW, 0), P(VBW, 14.5478)]
    pts.extend(P(*pt) for pt in cubic_bezier_points(
        (63.36, 14.5478), (63.36, 16.9405), (61.5978, 18.9678), (59.2284, 19.301), 16))
    pts.append(P(33.2888, 22.9488))
    pts.extend(P(*pt) for pt in cubic_bezier_points(
        (33.2888, 22.9488), (32.8582, 23.0093), (32.4215, 23.0111), (31.9905, 22.954), 8))
    pts.append(P(4.1702, 19.2719))
    pts.extend(P(*pt) for pt in cubic_bezier_points(
        (4.1702, 19.2719), (1.78348, 18.956), (0, 16.921), (0, 14.5134), 16))
    ImageDraw.Draw(layer).polygon(pts, fill=color)
    _ss_paste(img, layer, sp(x_dp), sp(y_dp))


def draw_bottom_rounded_badge(img, x_dp, y_dp, w_dp, h_dp, r_dp, color):
    """Rectangle with only the bottom corners rounded (style 2 badge, anti-aliased)."""
    bw, bh = sp(w_dp), sp(h_dp)
    layer = _ss_buffer(bw, bh)
    d = ImageDraw.Draw(layer)
    rr = sp(r_dp) * SS
    d.rounded_rectangle((0, 0, bw * SS - 1, bh * SS - 1), radius=rr, fill=color)
    d.rectangle((0, 0, bw * SS - 1, rr), fill=color)
    _ss_paste(img, layer, sp(x_dp), sp(y_dp))


def draw_plain_badge(img, x_dp, y_dp, w_dp, h_dp, color):
    ImageDraw.Draw(img).rectangle(
        (sp(x_dp), sp(y_dp), sp(x_dp + w_dp) - 1, sp(y_dp + h_dp) - 1),
        fill=color,
    )


def draw_pill_outline(img, x_dp, y_dp, w_dp, h_dp, r_dp, color, stroke_dp):
    """Rounded-rect outline anti-aliased via supersampling."""
    bw, bh = sp(w_dp), sp(h_dp)
    layer = _ss_buffer(bw, bh)
    ImageDraw.Draw(layer).rounded_rectangle(
        (0, 0, bw * SS - 1, bh * SS - 1),
        radius=sp(r_dp) * SS,
        outline=color,
        width=max(1, sp(stroke_dp) * SS),
    )
    _ss_paste(img, layer, sp(x_dp), sp(y_dp))


def draw_badge_text(img):
    """绘制徽标文字「千问小讲堂」：普惠 Bold，墨迹按设计框垂直居中。"""
    font = load_font(FONT_PUHUI_BOLD, BADGE_TEXT_PX)
    bbox = font.getbbox(BADGE_TEXT, anchor="la")
    ink_center = (bbox[1] + bbox[3]) / 2
    tx = sp(BADGE_TEXT_LEFT)
    ty = sp(BADGE_TEXT_CY) - ink_center
    ImageDraw.Draw(img).text((tx, ty), BADGE_TEXT, font=font, fill=WHITE, anchor="la")


def dominant_color(img, region=None, palette_size=8):
    """量化取主色，跳过过亮/过暗，保证徽标可读。"""
    work = img.crop(region) if region else img
    work = work.resize((96, 96)).convert("RGB")
    q = work.quantize(colors=palette_size)
    palette = q.getpalette()
    counts = sorted(q.getcolors() or [], reverse=True)
    for _cnt, idx in counts:
        r, g, b = palette[idx * 3], palette[idx * 3 + 1], palette[idx * 3 + 2]
        lum = 0.299 * r + 0.587 * g + 0.114 * b
        if 28 < lum < 228:
            return (r, g, b)
    if counts:
        _cnt, idx = counts[0]
        return (palette[idx * 3], palette[idx * 3 + 1], palette[idx * 3 + 2])
    return (80, 80, 80)


def region_luma(img, region):
    """区域平均亮度（0~1），用于判断底色深浅。"""
    work = img.crop(region).resize((24, 24)).convert("RGB")
    px = list(work.getdata())
    s = sum(0.299 * r + 0.587 * g + 0.114 * b for r, g, b in px)
    return s / (len(px) * 255)


def adapt_tint(rgb, luma, darken=0.50, lighten=0.30,
               light_cap=0.52, sat_boost=1.10):
    """自适应着色：浅底加深、深底提亮（克制+封顶到中灰，保住白字对比）。
    luma 为徽标区域底色平均亮度(0~1)。"""
    r, g, b = rgb
    h, l, s = colorsys.rgb_to_hls(r / 255, g / 255, b / 255)
    if luma >= 0.5:                       # 浅底 → 加深
        l = max(0.06, l * darken)
    else:                                  # 深底 → 提亮（封顶到 light_cap）
        l = min(light_cap, l + (1 - l) * lighten)
    s = max(0.0, min(1.0, s * sat_boost))
    nr, ng, nb = colorsys.hls_to_rgb(h, l, s)
    return (int(nr * 255), int(ng * 255), int(nb * 255))


def draw_frosted_badge(img, bg_img, x_dp, y_dp, w_dp, h_dp, r_dp,
                       alpha=0.60, blur_radius_px=3):
    """半透明毛玻璃徽标（自适应）：
    取徽标所在区域的主色 + 平均亮度，浅底加深 / 深底提亮；
    徽标处做轻度高斯模糊（保留背景形状）后叠 ~alpha 半透明色块。
    """
    x0, y0 = sp(x_dp), sp(y_dp)
    bw, bh = sp(w_dp), sp(h_dp)
    region = (x0, y0, x0 + bw, y0 + bh)
    luma = region_luma(bg_img, region)
    fill_rgb = adapt_tint(dominant_color(bg_img, region=region), luma)
    if blur_radius_px > 0:
        blurred = img.crop(region).filter(ImageFilter.GaussianBlur(blur_radius_px))
        img.paste(blurred, (x0, y0))
    fill = fill_rgb + (int(round(255 * alpha)),)
    layer = _ss_buffer(bw, bh)
    d = ImageDraw.Draw(layer)
    rr = sp(r_dp) * SS
    d.rounded_rectangle((0, 0, bw * SS - 1, bh * SS - 1), radius=rr, fill=fill)
    d.rectangle((0, 0, bw * SS - 1, rr), fill=fill)
    _ss_paste(img, layer, x0, y0)


def draw_badge(img, style, color):
    """style1/2 徽标：实心色块（theme_color）+ 文字。"""
    x_dp, y_dp = 8.64, 0
    if style == 1:
        draw_drape_badge(img, x_dp, y_dp, 63.36, 23.04, color)
    elif style == 2:
        draw_bottom_rounded_badge(img, x_dp, y_dp, 63.36, 21.6, 4.8, color)
    else:
        draw_plain_badge(img, x_dp, y_dp, 63.36, 21.6, color)
    draw_badge_text(img)


# ---------------------------------------------------------------------------
# 钩子（标题）排版：单一文本，按宽度贪心换行（≤2 行），固定字号 + 字间距
# ---------------------------------------------------------------------------

def tracked_width(text, font, track_px):
    """字符串施加字间距后的输出像素宽度。track_px 为输出像素。"""
    if not text:
        return 0
    return sum(font.getlength(ch) for ch in text) + track_px * (len(text) - 1)


def wrap_greedy(text, font, track_px, max_w_px):
    """从左往右贪心断行：当前行放不下就换行。最多不限，由调用方判定行数。"""
    lines, cur = [], ""
    for ch in text:
        trial = cur + ch
        if cur and tracked_width(trial, font, track_px) > max_w_px:
            lines.append(cur)
            cur = ch
        else:
            cur = trial
    if cur:
        lines.append(cur)
    return lines


def avoid_orphan(lines):
    """轻量防孤字：2 行且末行只剩 1 字时，从上一行末尾拉一个字下来。"""
    if len(lines) == 2 and len(lines[1]) == 1 and len(lines[0]) >= 2:
        return [lines[0][:-1], lines[0][-1] + lines[1]]
    return lines


def layout_hook(text, font_path, base_px, max_w_dp, track_dp, min_px=TITLE_MIN_PX):
    """把钩子排成 ≤2 行：固定 base_px 贪心换行；若超过 2 行才逐步缩字。

    返回 (font, lines, track_px)。
    """
    max_w_px = sp(max_w_dp)
    size = base_px
    while size >= min_px:
        font = load_font(font_path, size)
        track_px = sp(track_dp) if track_dp else 0
        lines = wrap_greedy(text, font, track_px, max_w_px)
        if len(lines) <= 2:
            return font, avoid_orphan(lines), track_px
        size -= 0.5
    font = load_font(font_path, min_px)
    track_px = sp(track_dp) if track_dp else 0
    return font, avoid_orphan(wrap_greedy(text, font, track_px, max_w_px)), track_px


def draw_tracked(img, line, font, track_px, x_px, y_px, color):
    """逐字渲染并施加字间距，anchor 'la'（左-ascender）。"""
    d = ImageDraw.Draw(img)
    x = x_px
    for ch in line:
        d.text((x, y_px), ch, font=font, fill=color, anchor="la")
        x += font.getlength(ch) + track_px


def draw_hook(img, text, font_path, base_px, track_dp, lh_ratio, color,
              x_dp=TITLE_X, y_dp=TITLE_Y, max_w_dp=TITLE_MAX_W):
    """绘制钩子，返回钩子块底部的设计 px（供标签定位）。"""
    if not text:
        return y_dp
    font, lines, track_px = layout_hook(text, font_path, base_px, max_w_dp, track_dp)
    line_h_px = sp(base_px * lh_ratio)
    x_px = sp(x_dp)
    y_px = sp(y_dp)
    for i, line in enumerate(lines):
        draw_tracked(img, line, font, track_px, x_px, y_px + i * line_h_px, color)
    bottom_px = y_px + (len(lines) - 1) * line_h_px + line_h_px
    return bottom_px / SCALE


# ---------------------------------------------------------------------------
# 标签
# ---------------------------------------------------------------------------

def parse_tag(raw):
    """把『【左】★【右】』解析为『左★右』；其它格式原样去掉【】。"""
    if not raw:
        return ""
    return raw.replace("【", "").replace("】", "").strip()


def draw_tag(img, tag, x_dp, y_dp, radius_dp):
    if not tag:
        return
    tag_font = load_font(FONT_PUHUI_MEDIUM, TAG_PX)
    text_w_px = tag_font.getlength(tag)
    pill_w_dp = (text_w_px / SCALE) + TAG_PAD_X * 2
    draw_pill_outline(img, x_dp, y_dp, pill_w_dp, TAG_H, radius_dp, WHITE, 0.96)
    tx = sp(x_dp + pill_w_dp / 2) - int(text_w_px) // 2
    ty = sp(y_dp + TAG_H / 2)
    ImageDraw.Draw(img).text((tx, ty), tag, font=tag_font, fill=WHITE, anchor="lm")


# ---------------------------------------------------------------------------
# Layouts
# ---------------------------------------------------------------------------

def render_style1(bg_path, hook, tag):
    """小低数学 — 造字工房元黑钩子（≤2 行）+ 描边胶囊标签 + 垂帘徽标。"""
    img = card_base(bg_path)
    badge_color = theme_color(img)

    hook = hook or ""
    tag = parse_tag(tag)

    bottom = draw_hook(img, hook, FONT_YUANHEI, TITLE_LOW_PX, TITLE_LOW_TRACK,
                       TITLE_LOW_LH, WHITE)
    if tag:
        draw_tag(img, tag, TITLE_X, bottom + TAG_GAP, 10.56)

    draw_badge(img, 1, badge_color)
    return img


def render_style2(bg_path, hook, tag):
    """小高数学 — 普惠体 Bold 钩子（≤2 行）+ 矩形标签 + 下圆角徽标（过渡版）。"""
    img = card_base(bg_path)
    badge_color = theme_color(img)

    hook = hook or ""
    tag = parse_tag(tag)

    bottom = draw_hook(img, hook, FONT_PUHUI_BOLD, TITLE_HIGH_PX, TITLE_HIGH_TRACK,
                       TITLE_HIGH_LH, WHITE)
    if tag:
        draw_tag(img, tag, TITLE_X, bottom + TAG_GAP, 1.92)

    draw_badge(img, 2, badge_color)
    return img


def bottom_color(img):
    """取底图下半部主色并加深，作为底部压暗带颜色。"""
    strip = img.convert("RGB").crop((0, int(H * 0.6), W, H))
    r, g, b = strip.resize((1, 1), Image.LANCZOS).getpixel((0, 0))
    return (int(r * SCRIM_DARKEN), int(g * SCRIM_DARKEN), int(b * SCRIM_DARKEN))


def add_bottom_scrim(img, color, start_frac=SCRIM_START, max_alpha=SCRIM_ALPHA):
    """底部竖向渐变压暗带：start_frac 处全透明 → 底部 max_alpha。"""
    start = int(H * start_frac)
    grad = Image.new("L", (1, H), 0)
    px = grad.load()
    span = max(1, H - start)
    for y in range(start, H):
        px[0, y] = int(max_alpha * 255 * (y - start) / span)
    alpha = grad.resize((W, H))
    overlay = Image.new("RGBA", (W, H), color + (0,))
    overlay.putalpha(alpha)
    img.alpha_composite(overlay)


# ---------------------------------------------------------------------------
# 学名（style3）标点/语义感知换行
# ---------------------------------------------------------------------------

# 行首禁则：闭合/后置标点不可出现在行首
NO_LINE_START = set("）)】》」』〕〉、，。．！？：；·…—’”%》")
# 行尾禁则：开启/前置标点不可出现在行尾
NO_LINE_END = set("（(【《「『〔〈‘“")
CLOSERS = set("）)】》」』〕〉”’")    # 闭合括号/引号：其后断行最自然
OPENERS = set("（(【《「『〔〈“‘")     # 开启括号/引号：其前断行最自然
MIDSEP = set("·・、，；")            # 间隔/分隔符：其后断行次优


# 字库缺字的标点归一化：日文/全角中点 → 间隔号 ·
PUNCT_NORMALIZE = {"\u30fb": "·", "\uff65": "·", "\u2027": "·", "\u00b7": "·"}


def normalize_title_punct(s):
    for k, v in PUNCT_NORMALIZE.items():
        s = s.replace(k, v)
    return s


def split_two_lines_cjk(text, font, track_px, max_w_px):
    """把 text 切成 2 行，遵守标点禁则并优先在标点边界断开。

    返回 (priority, balance, [l1, l2]) 或 None。
    优先级：闭合括号后 / 开启括号前 (0) > 间隔符后 (1) > 普通位置 (2)。
    """
    n = len(text)
    cands = []
    for i in range(1, n):
        l1, l2 = text[:i], text[i:]
        if l1[-1] in NO_LINE_END or l2[0] in NO_LINE_START:
            continue
        w1 = tracked_width(l1, font, track_px)
        w2 = tracked_width(l2, font, track_px)
        if w1 > max_w_px or w2 > max_w_px:
            continue
        if l1[-1] in CLOSERS or l2[0] in OPENERS:
            pri = 0
        elif l1[-1] in MIDSEP:
            pri = 1
        else:
            pri = 2
        cands.append((pri, abs(w1 - w2), [l1, l2]))
    if not cands:
        return None
    cands.sort(key=lambda c: (c[0], c[1]))
    return cands[0]


def layout_title_cjk(text, font_path, base_px, max_w_dp, track_dp,
                     min_px=TITLE3_MIN_PX):
    """学名排版：尽量一行；放不下按标点/语义切 2 行。

    2 行时全字号搜索：标点边界(括号/间隔号)优先级最高，必要时略微缩字以
    在标点处断开；其次取缩字最少、两行最平衡的方案。返回 (font, lines, track_px)。
    """
    max_w_px = sp(max_w_dp)
    track_px = sp(track_dp) if track_dp else 0

    # 尽量一行：在 [ONELINE_MIN, base] 内取最大可容纳一行的字号
    size = base_px
    while size >= TITLE3_ONELINE_MIN_PX:
        font = load_font(font_path, size)
        if tracked_width(text, font, track_px) <= max_w_px:
            return font, [text], track_px
        size -= 0.5

    best = None  # (key, font, [l1,l2])
    size = base_px
    while size >= min_px:
        f = load_font(font_path, size)
        cand = split_two_lines_cjk(text, f, track_px, max_w_px)
        if cand:
            pri, bal, two = cand
            key = (pri, round(base_px - size, 3), bal)
            if best is None or key < best[0]:
                best = (key, f, two)
        size -= 0.5
    if best:
        return best[1], best[2], track_px

    f = load_font(font_path, min_px)
    return f, wrap_greedy(text, f, track_px, max_w_px), track_px


def draw_name_centered(img, name, color=WHITE, cy_dp=TITLE3_CY, max_w_dp=TITLE3_MAX_W):
    """style3 学名：HYXinRenWenSong 居中，标点/语义换行（尽量一行，最多 2 行）。"""
    if not name:
        return
    name = normalize_title_punct(name)
    font, lines, track_px = layout_title_cjk(
        name, FONT_XINRENWENSONG, TITLE3_PX, max_w_dp, TITLE3_TRACK)
    line_h = int(round(font.size * TITLE3_LH))
    total_h = line_h * len(lines)
    top = sp(cy_dp) - total_h // 2
    for i, line in enumerate(lines):
        line_w = tracked_width(line, font, track_px)
        x = (W - line_w) / 2
        draw_tracked(img, line, font, track_px, int(x), top + i * line_h, color)


# 学名区底部自适应压暗带：仅在底色偏亮时触发，确保白字可读（非文字投影）
TITLE3_SCRIM_START = 0.50    # 渐变起点（占画面高比例）
TITLE3_SCRIM_FULL = 0.74     # 到此达到满强度并保持到底
TITLE3_SCRIM_RGB = (12, 12, 16)


def title_area_luma(img):
    """学名所在区域的平均亮度（0–255）。"""
    reg = img.convert("RGB").crop(
        (int(W * 0.12), int(H * 0.70), int(W * 0.88), int(H * 0.92))
    ).resize((32, 16))
    px = list(reg.getdata())
    return sum(0.299 * r + 0.587 * g + 0.114 * b for r, g, b in px) / len(px)


def add_title_scrim(img):
    """按学名区底色亮度自适应叠加底部渐变压暗带；深底几乎不触发。"""
    luma = title_area_luma(img)
    alpha = max(0.0, min(0.78, (luma - 90) / 140.0))
    if alpha <= 0.02:
        return
    start = int(H * TITLE3_SCRIM_START)
    full = int(H * TITLE3_SCRIM_FULL)
    grad = Image.new("L", (1, H), 0)
    pp = grad.load()
    for y in range(start, H):
        if y < full:
            v = alpha * 255 * (y - start) / (full - start)
        else:
            v = alpha * 255
        pp[0, y] = int(v)
    overlay = Image.new("RGBA", (W, H), TITLE3_SCRIM_RGB + (0,))
    overlay.putalpha(grad.resize((W, H)))
    img.alpha_composite(overlay)


def display_name_from_xueming(name):
    """封面主文案：有《》则只取书名号内文字，否则用整段学名。"""
    if not name:
        return ""
    m = re.search(r"《([^》]+)》", name)
    return m.group(1).strip() if m else name.strip()


def render_style3(bg_path, name, tag=""):
    """语文成语类 — HYXinRenWenSong 学名居中 + 底部自适应压暗带 + 半透明毛玻璃徽标。

    `tag` 参数仅为统一签名，本布局不使用。
    """
    img = card_base(bg_path)
    bg_rgb = img.convert("RGB")
    draw_name_centered(img, display_name_from_xueming(name))
    draw_frosted_badge(img, bg_rgb, 8.64, 0, 63.36, 21.6, r_dp=4.8)
    draw_badge_text(img)
    return img


# ---------------------------------------------------------------------------
# 成语故事：学名（大标题）+ 钩子（小副标题），二者居中、整体下三分之一
# ---------------------------------------------------------------------------
# 标题（学名）：HYXinRenWenSong 65W 22.99 / 行高24.1 / 间距1.44，框 Y128 H25 → 中心140.5
TH_TITLE_MAXW = 120.0
TH_TITLE_CY = 128.0 + 25.0 / 2          # 140.5
# 钩子：HYXinRenWenSong **W 常规** 11.52 / 行高13.4 / 间距1.09 / 白70%，框 Y156.48 H27 → 中心≈170
TH_HOOK_PX = 11.52
TH_HOOK_TRACK = 1.09
TH_HOOK_LH = 13.4 / 11.52
TH_HOOK_MAXW = 100.0
TH_HOOK_CY = 156.48 + 27.0 / 2          # ≈169.98
TH_HOOK_ALPHA = 0.70


def wrap_hook_lines(text, font, track_px, max_w_px):
    """钩子断行：优先在中文逗号/顿号/分号处断为两行（去掉该标点）；
    无标点则按宽度贪心，并尽量均衡为两行。"""
    text = (text or "").strip()
    if not text:
        return []
    seps = "，、；,"
    idxs = [i for i, ch in enumerate(text) if ch in seps]
    if idxs:
        mid = len(text) / 2
        i = min(idxs, key=lambda k: abs(k - mid))
        return [text[:i].strip(), text[i + 1:].strip()]
    lines = wrap_greedy(text, font, track_px, max_w_px)
    if len(lines) <= 1:
        return lines
    cut = (len(text) + 1) // 2          # 无标点时按字数对半，二行更均衡
    return [text[:cut], text[cut:]]


def _draw_centered_lines(img, lines, font, track_px, line_h, top, color):
    """居中绘制多行，返回该块底部 y(px)。"""
    for i, line in enumerate(lines):
        w = tracked_width(line, font, track_px)
        x = (W - w) / 2
        draw_tracked(img, line, font, track_px, int(x), int(top + i * line_h), color)
    return top + len(lines) * line_h


def render_chengyu(bg_path, name, hook=""):
    """成语故事 — 学名(HYXinRenWenSong 65W 大字) + 钩子(HYXinRenWenSong W 常规小字, 白70%)。
    二者居中，按 Figma 固定位：标题中心 Y140.5、钩子中心 Y170。
    """
    img = card_base(bg_path)
    bg_rgb = img.convert("RGB")

    # 学名标题（固定中心）
    tf, tlines, ttrack = layout_hook(
        name or "", FONT_XINRENWENSONG, TITLE3_PX, TH_TITLE_MAXW, TITLE3_TRACK)
    tlh = sp(TITLE3_PX * TITLE3_LH)
    t_top = sp(TH_TITLE_CY) - tlh * len(tlines) / 2
    _draw_centered_lines(img, tlines, tf, ttrack, tlh, t_top, WHITE)

    # 钩子副标题（固定中心，70% 白，单独图层合成以保证半透明正确）
    hook = (hook or "").strip()
    if hook:
        hf = load_font(FONT_XINRENWENSONG_W, TH_HOOK_PX)
        htrack = sp(TH_HOOK_TRACK) if TH_HOOK_TRACK else 0
        hlines = wrap_hook_lines(hook, hf, htrack, sp(TH_HOOK_MAXW))
        hlh = sp(TH_HOOK_PX * TH_HOOK_LH)
        h_top = sp(TH_HOOK_CY) - hlh * len(hlines) / 2
        layer = Image.new("RGBA", img.size, (0, 0, 0, 0))
        a = int(round(255 * TH_HOOK_ALPHA))
        _draw_centered_lines(layer, hlines, hf, htrack, hlh, h_top, (255, 255, 255, a))
        img.alpha_composite(layer)

    draw_frosted_badge(img, bg_rgb, 8.64, 0, 63.36, 21.6, r_dp=4.8)
    draw_badge_text(img)
    return img


RENDERERS = {
    1: render_style1,
    2: render_style2,
    3: render_style3,
}


def render(style, bg_path, hook, tag=""):
    if style not in RENDERERS:
        raise ValueError(f"unknown style {style}; expected 1/2/3")
    return RENDERERS[style](bg_path, hook, tag)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

import io

TARGET_MAX_KB = 50   # JPEG 体积上限
TARGET_MIN_KB = 30   # 期望下限（仅用于提示）


def save(img, path):
    """png → 无损；jpg/jpeg → 二分搜索质量，卡在 ≤TARGET_MAX_KB 的最高画质。"""
    flat = Image.new("RGB", img.size, "#FFFFFF")
    flat.paste(img, mask=img.split()[-1])
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".jpg", ".jpeg"):
        flat.save(path, "PNG", optimize=True, compress_level=9)
        return

    cap = TARGET_MAX_KB * 1024
    lo, hi, best = 30, 92, None
    while lo <= hi:
        mid = (lo + hi) // 2
        buf = io.BytesIO()
        flat.save(buf, "JPEG", quality=mid, optimize=True)
        if buf.getbuffer().nbytes <= cap:
            best = buf.getvalue()
            lo = mid + 1
        else:
            hi = mid - 1
    if best is None:  # 连 q=30 都超标，用 q=30 兜底
        buf = io.BytesIO()
        flat.save(buf, "JPEG", quality=30, optimize=True)
        best = buf.getvalue()
    with open(path, "wb") as f:
        f.write(best)


LOW_GRADES = {"一年级", "二年级", "三年级"}


def grade_to_style(grade):
    """一/二/三年级 → style1（小低数学）；其余 → style2（小高数学）。"""
    return 1 if str(grade).strip() in LOW_GRADES else 2


def fetch_bg(url, cache_dir):
    """下载 URL 背景图到本地缓存，返回本地路径（已存在则直接复用）。"""
    import hashlib
    import urllib.request

    os.makedirs(cache_dir, exist_ok=True)
    ext = os.path.splitext(url.split("?")[0])[1] or ".jpg"
    name = hashlib.md5(url.encode()).hexdigest() + ext
    path = os.path.join(cache_dir, name)
    if not os.path.exists(path):
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp, open(path, "wb") as f:
            f.write(resp.read())
    return path


def run_single(args):
    img = render(args.style, args.bg, args.hook, args.tag or "")
    save(img, args.output)
    print(f"wrote {args.output} (style={args.style})")


# 确认版 xlsx 列名
COL_GRADE = "年级"
COL_HOOK = "钩子确认版"
COL_TAG = "标签确认版"
COL_URL = "url确认版"

# ---- OSS 上传（复用 mf4_xise.py 的 ossutil + Souti OSS 配置）----
OSSUTIL = os.path.expanduser("~/Desktop/ossutilmac64")
OSS_ENDPOINT = "oss-cn-hangzhou.aliyuncs.com"
OSS_AK = os.environ["SOUTI_OSS_ACCESS_KEY_ID"]
OSS_SK = os.environ["SOUTI_OSS_ACCESS_KEY_SECRET"]
OSS_DEST = "oss://sm-frontend-private-img/souti-imgs-lasting/"
OSS_CDN = "https://cdn-private.sm.cn/souti-imgs-lasting/"


def upload_to_oss(path):
    """纯上传本地文件到 Souti OSS（不吸色），返回 CDN URL。"""
    cmd = [OSSUTIL, "-e", OSS_ENDPOINT, "-i", OSS_AK, "-k", OSS_SK,
           "cp", "-f", path, OSS_DEST]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(r.stderr.strip() or r.stdout.strip())
    return OSS_CDN + os.path.basename(path)


def run_batch(args):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("error: openpyxl is required for batch mode (pip install openpyxl)", file=sys.stderr)
        sys.exit(1)

    os.makedirs(args.out_dir, exist_ok=True)
    cache_dir = os.path.join(args.out_dir, "_bg_cache")
    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(headers)}
    for r in (COL_GRADE, COL_HOOK, COL_URL):
        if r not in col:
            print(f"error: xlsx 缺少列 '{r}' (现有: {headers})", file=sys.stderr)
            sys.exit(1)

    count, errors = 0, 0
    manifest = []
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[col[COL_HOOK]]:
            continue
        hook = str(row[col[COL_HOOK]]).strip()
        grade = str(row[col[COL_GRADE]] or "").strip()
        tag = str(row[col[COL_TAG]] or "") if COL_TAG in col else ""
        url = row[col[COL_URL]]
        style = grade_to_style(grade)
        if not url:
            print(f"  [行{n}] 跳过：无 url（{hook}）", file=sys.stderr)
            errors += 1
            continue
        src_url = str(url).strip()
        try:
            bg = fetch_bg(src_url, cache_dir)
            img = render(style, bg, hook, tag)
            fname = uuid.uuid4().hex + "." + args.format
            out_path = os.path.join(args.out_dir, fname)
            save(img, out_path)
            cover_url = ""
            if args.upload:
                cover_url = upload_to_oss(out_path)
            manifest.append([grade, style, hook, parse_tag(tag), src_url, fname, cover_url])
            count += 1
            if args.upload and count % 20 == 0:
                print(f"  ...已上传 {count} 张")
        except Exception as e:  # noqa: BLE001
            print(f"  [行{n}] 失败：{hook} -> {e}", file=sys.stderr)
            errors += 1

    manifest_path = os.path.join(args.out_dir, "manifest.csv")
    with open(manifest_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["年级", "style", "钩子", "标签", "原图url", "文件名", "url"])
        w.writerows(manifest)
    print(f"wrote {count} images to {args.out_dir} (失败 {errors})")
    print(f"清单: {manifest_path}")


def run_merge(args):
    """把 manifest.csv 的封面 URL 按『原图url』匹配写回 xlsx，新增一列『封面url』。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("error: openpyxl is required (pip install openpyxl)", file=sys.stderr)
        sys.exit(1)

    url_map = {}
    with open(args.manifest, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if row.get("原图url") and row.get("url"):
                url_map[row["原图url"].strip()] = row["url"].strip()

    wb = load_workbook(args.xlsx)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    src_idx = headers.index(COL_URL)
    # 若同名列已存在则复用，否则新增到最右
    if args.col in headers:
        new_col = headers.index(args.col) + 1
    else:
        new_col = len(headers) + 1
    ws.cell(row=1, column=new_col, value=args.col)
    matched = 0
    for n in range(2, ws.max_row + 1):
        src = ws.cell(row=n, column=src_idx + 1).value
        if src and str(src).strip() in url_map:
            ws.cell(row=n, column=new_col, value=url_map[str(src).strip()])
            matched += 1
    wb.save(args.out)
    print(f"matched {matched} 行，写入 {args.out}")


def build_parser():
    p = argparse.ArgumentParser(
        description="Render 千问小讲堂 数学挂载封面 (style1=小低, style2=小高).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = p.add_subparsers(dest="mode")

    one = sub.add_parser("one", help="render one cover")
    one.add_argument("--style", type=int, required=True, choices=[1, 2])
    one.add_argument("--bg", required=True, help="background image path")
    one.add_argument("--hook", "--title", dest="hook", required=True,
                     help="钩子文案（完整一段，自动换行）")
    one.add_argument("--tag", default="", help="标签，如 【一年级】★【超好玩】")
    one.add_argument("--output", "-o", required=True)
    one.set_defaults(func=run_single)

    batch = sub.add_parser("batch", help="按确认版 xlsx 批量生成")
    batch.add_argument("--xlsx", required=True,
                       help="列：年级 / 钩子确认版 / 标签确认版 / url确认版")
    batch.add_argument("--out-dir", required=True)
    batch.add_argument("--format", choices=["jpg", "png"], default="jpg",
                       help="jpg=压到≤50K；png=无损（未压缩）")
    batch.add_argument("--upload", action="store_true",
                       help="生成后纯上传 OSS 并在 manifest.csv 写回 URL")
    batch.set_defaults(func=run_batch)

    merge = sub.add_parser("merge", help="把 manifest.csv 的封面 URL 写回 xlsx")
    merge.add_argument("--xlsx", required=True, help="原始 xlsx")
    merge.add_argument("--manifest", required=True, help="manifest.csv")
    merge.add_argument("--out", required=True, help="输出新 xlsx")
    merge.add_argument("--col", default="封面url", help="写入的列名（默认 封面url）")
    merge.set_defaults(func=run_merge)

    return p


def main(argv=None):
    parser = build_parser()
    args = parser.parse_args(argv)
    if not getattr(args, "func", None):
        parser.print_help()
        sys.exit(1)
    args.func(args)


if __name__ == "__main__":
    main()
