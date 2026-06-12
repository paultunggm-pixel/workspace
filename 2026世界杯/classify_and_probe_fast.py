#!/usr/bin/env python3
"""
世界杯时效性站点 — 分类 + 并发拨测 + 生成HTML (快速版)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import requests
import re
import time
from urllib.parse import urlparse
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import json

INPUT_PATH = '/Users/d.j.f/Documents/Claude/2026世界杯/世界杯时效性站点列表页/世界杯时效性站点列表页.xlsx'
OUTPUT_EXCEL = '/Users/d.j.f/Documents/Claude/2026世界杯/世界杯时效性站点分类排查结果.xlsx'
OUTPUT_HTML = '/Users/d.j.f/Documents/Claude/2026世界杯/世界杯时效性站点排查报告.html'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
}

# 国内站点域名关键词
CN_DOMAINS = [
    '.cn', 'sina.com.cn', '163.com', 'qq.com', 'sohu.com',
    'zhibo8.com', 'dongqiudi.com', 'hupu.com', 'ppsport.com',
    'cctv.com', 'people.com.cn', 'xinhuanet.com', 'ifeng.com',
    'toutiao.com', 'lesports.com', 'leisu.com'
]

SITE_COUNTRY = {
    'espn.com': '美国','bbc.com': '英国','bbc.co.uk': '英国',
    'theguardian.com': '英国','goal.com': '英国','cbssports.com': '美国',
    'foxsports.com': '美国','sportsmole.co.uk': '英国','marca.com': '西班牙',
    'lequipe.fr': '法国','bild.de': '德国','sportingnews.com': '美国',
    'gazzetta.it': '意大利','si.com': '美国','actionnetwork.com': '美国',
    'fourfourtwo.com': '英国','theanalyst.com': '英国',
    'sofascore.com': '克罗地亚','reuters.com': '英国',
    'skysports.com': '英国','nytimes.com': '美国',
    'whoscored.com': '英国','beinsports.com': '法国',
    'expertpicks.com': '美国','leaguelane.com': '英国',
    'uefa.com': '瑞士','kickoff.guide': '英国','fifa.com': '瑞士',
    'apnews.com': '美国','bleacherreport.com': '美国',
    'newsnow.co.uk': '英国','90min.com': '美国',
    'premierleague.com': '英国','football365.com': '英国',
    'hudl.com': '美国','planetfootball.com': '英国',
    'spielverlagerung.de': '德国','onefootball.com': '德国',
    'manchestereveningnews.co.uk': '英国','abola.pt': '葡萄牙',
    'record.pt': '葡萄牙','bundesliga.com': '德国',
    'sbs.com.au': '澳大利亚','tribalfootball.com': '英国',
    'worldsoccertalk.com': '美国','insideworldfootball.com': '英国',
    'mirror.co.uk': '英国','fotmob.com': '挪威',
    'as.com': '西班牙','independent.co.uk': '英国',
    'dailymail.com': '美国','mundodeportivo.com': '西班牙',
    'football-italia.net': '意大利','flashscore.com': '捷克',
}

def classify_site(name, url):
    if not url:
        return 'unknown', ''
    parsed = urlparse(url)
    domain = parsed.netloc.lower()
    for cn in CN_DOMAINS:
        if cn in domain or domain.endswith('.cn'):
            return '国内', ''
    for key, country in SITE_COUNTRY.items():
        if key in domain:
            return '境外', country
    return '境外', '未知(国际)'

def probe_one(r):
    """并发拨测一个URL"""
    url, site_name = r['url'], r['site']
    result = {'http_status': 0, 'redirect': '', 'is_list': False,
              'probe_result': '', 'corrected_url': '', 'corrected_note': ''}

    try:
        resp = requests.get(url, headers=HEADERS, timeout=8, allow_redirects=True, verify=False)
        http_code = resp.status_code
        final_url = resp.url
        content = resp.text
        content_lower = content.lower()

        # WC关键词
        wc_keywords = ['world cup', 'world-cup', 'worldcup', 'fifa world cup',
                       'fifa-world-cup', '世界杯', '国际足球', 'w杯',
                       'wm 2026', 'coupe du monde', 'copa mundial', 'coppa del mondo']
        has_wc = any(kw in content_lower for kw in wc_keywords)

        # 列表特征
        list_markers = ['<article ', 'class="article', 'class="post', 'class="news',
                        'class="list', 'class="feed', 'class="timeline', 'class="latest',
                        'article_list', 'news-list', 'news_list', 'articleList',
                        'class="news-item', 'class="hot-news', 'class="content-list',
                        'class="story', 'class="headline', 'class="card--',
                        'class="entry', 'class="result-',
                        'data-testid="article', '<li><a href=', 'class="title',
                        'class="c-news', '新闻列表']

        list_score = sum(1 for m in list_markers if m in content_lower)

        # 单篇详情页特征
        single_markers = ['article-detail', 'article_detail', 'single-article',
                          'class="article-body', 'article-full', 'article-content',
                          '<article class=', 'data-post-type=']
        is_single = sum(1 for m in single_markers if m in content_lower) >= 2

        # 决策
        if is_single and list_score < 5:
            result['probe_result'] = '疑似单篇详情页'
        elif not has_wc:
            result['probe_result'] = '未找到世界杯相关内容'
        elif list_score < 3:
            result['probe_result'] = '列表特征不明显'
        else:
            result['is_list'] = True
            result['probe_result'] = '✅ 确认为资讯列表页'

        result['http_status'] = http_code
        if final_url != url:
            result['redirect'] = final_url

        # 非有效? 尝试找正确URL
        if not result['is_list'] and http_code == 200:
            parsed = urlparse(url)
            base = f"{parsed.scheme}://{parsed.netloc}"
            wc_paths = ['/football/world-cup', '/football/world-cup-2026',
                        '/soccer/world-cup', '/football/worldcup',
                        '/sport/football/world-cup', '/en/football/world-cup',
                        '/football/coupe-du-monde', '/futbol/mundial',
                        '/football/wm-2026']
            for path in wc_paths:
                try:
                    r2 = requests.get(base + path, headers=HEADERS, timeout=6, verify=False)
                    if r2.status_code == 200 and ('world cup' in r2.text.lower() or 'world-cup' in r2.text.lower()):
                        result['corrected_url'] = base + path
                        result['corrected_note'] = f'自动发现(含WC内容)'
                        break
                except:
                    continue
            if not result['corrected_url']:
                result['corrected_note'] = '未找到有效替代'

    except requests.Timeout:
        result['probe_result'] = '超时(>8s)'
    except requests.ConnectionError:
        result['probe_result'] = '连接失败'
    except Exception as e:
        result['probe_result'] = f'异常:{str(e)[:60]}'

    r.update(result)
    return r

# ── 主流程 ──
import urllib3
urllib3.disable_warnings()

print("1. 加载Excel + 分类...")
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb.active
rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name, url, date = row
    if name and url:
        r = {'site': name, 'url': str(url).strip(), 'date': str(date) if date else ''}
        cat, country = classify_site(name, r['url'])
        r['category'] = cat
        r['country'] = country
        rows.append(r)

cn_count = sum(1 for r in rows if r['category'] == '国内')
intl_count = sum(1 for r in rows if r['category'] == '境外')
print(f"   国内:{cn_count} | 境外:{intl_count} | 总计:{len(rows)}")

# ── 2. 并发拨测 ──
print(f"\n2. 并发拨测 {len(rows)} 个URL (8s超时, 10并发)...")
start_t = time.time()
with ThreadPoolExecutor(max_workers=10) as pool:
    futures = {pool.submit(probe_one, r): r for r in rows}
    for i, f in enumerate(as_completed(futures), 1):
        r = f.result()
        status = '✅' if r['is_list'] else ('🔍' if r['corrected_url'] else '❌')
        print(f"   [{i:2d}/{len(rows)}] {status} {r['site']:20s} HTTP{r['http_status']} {r['probe_result'][:50]}")

elapsed = time.time() - start_t
valid = sum(1 for r in rows if r['is_list'])
invalid = sum(1 for r in rows if not r['is_list'])
corrected = sum(1 for r in rows if r['corrected_url'])
print(f"\n   耗时:{elapsed:.1f}s | 有效:{valid} | 非有效:{invalid} | 已补充:{corrected}")

# ── 3. 生成 Excel ──
print("\n3. 生成 Excel...")
wb_out = openpyxl.Workbook()
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def write_sheet(ws, headers, data_rows, widths):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for ri, vals in enumerate(data_rows, 2):
        for c, v in enumerate(vals, 1):
            ws.cell(row=ri, column=c, value=v)
        if not data_rows[ri-2][5] if len(data_rows[ri-2]) > 5 else False:
            pass  # handled below
    # highlight invalid rows
    for ri, item in enumerate(data_rows, 2):
        if len(item) > 5 and item[5] != '✅ 是':
            if item[5] == '⚠️ 否':
                for c in range(1, len(headers)+1):
                    ws.cell(row=ri, column=c).fill = red_fill
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

# Sheet 1: 国内站点
ws_cn = wb_out.active
ws_cn.title = "国内站点"
cn_headers = ['序号','站点','URL','HTTP','跳转','是否资讯列表','排查说明','补充URL','补充说明']
cn_data = []
for i, r in enumerate([r for r in rows if r['category']=='国内'], 1):
    cn_data.append([i, r['site'], r['url'], r['http_status'], r['redirect'],
                    '✅ 是' if r['is_list'] else '⚠️ 否',
                    r['probe_result'], r['corrected_url'], r['corrected_note']])
write_sheet(ws_cn, cn_headers, cn_data, [6,18,55,8,45,12,28,55,22])

# Sheet 2: 境外站点
ws_intl = wb_out.create_sheet("境外站点")
intl_headers = ['序号','站点','URL','国家','HTTP','跳转','是否资讯列表','排查说明','补充URL','补充说明']
intl_data = []
for i, r in enumerate([r for r in rows if r['category']=='境外'], 1):
    intl_data.append([i, r['site'], r['url'], r['country'], r['http_status'], r['redirect'],
                      '✅ 是' if r['is_list'] else '⚠️ 否',
                      r['probe_result'], r['corrected_url'], r['corrected_note']])
write_sheet(ws_intl, intl_headers, intl_data, [6,20,55,12,8,45,12,28,55,22])

wb_out.save(OUTPUT_EXCEL)
print(f"   Excel: {OUTPUT_EXCEL}")

# ── 4. 生成 HTML ──
print("4. 生成 HTML...")
cn_sites = [r for r in rows if r['category']=='国内']
intl_sites = [r for r in rows if r['category']=='境外']

html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>2026世界杯时效性站点排查报告</title>
<style>
:root{{--bg:#0f1117;--card:#1a1d27;--border:#2a2d3a;--text:#e1e4eb;--muted:#888b96;--accent:#5b7fff;--green:#00c853;--red:#ff5252;--orange:#ff9100;--yellow:#ffd600}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Microsoft YaHei",sans-serif;background:var(--bg);color:var(--text);line-height:1.7}}
.container{{max-width:1300px;margin:0 auto;padding:32px 24px 80px}}
.hero{{text-align:center;padding:48px 0 32px;border-bottom:1px solid var(--border);margin-bottom:40px}}
.hero h1{{font-size:28px;margin-bottom:8px}}
.hero p{{color:var(--muted);font-size:14px}}
.stats{{display:flex;gap:16px;margin-bottom:40px;flex-wrap:wrap}}
.sc{{flex:1;min-width:130px;background:var(--card);border:1px solid var(--border);border-radius:12px;padding:18px;text-align:center}}
.sc .n{{font-size:34px;font-weight:700}}.sc .l{{color:var(--muted);font-size:13px;margin-top:4px}}
.sc.g .n{{color:var(--green)}}.sc.r .n{{color:var(--red)}}.sc.o .n{{color:var(--orange)}}
h2{{font-size:19px;margin:36px 0 16px;padding-bottom:8px;border-bottom:2px solid var(--accent)}}
.badge{{font-size:12px;padding:2px 10px;border-radius:12px;background:var(--accent);color:#fff;margin-left:8px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#252836;padding:10px 10px;text-align:left;position:sticky;top:0}}
td{{padding:8px 10px;border-bottom:1px solid var(--border);vertical-align:top}}
tr:hover td{{background:rgba(91,127,255,0.06)}}
tr.bad td{{background:rgba(255,82,82,0.07)}}
tr.fix td{{border-left:3px solid var(--yellow);background:rgba(255,214,0,0.05)}}
.ok{{color:var(--green);font-weight:bold}}.no{{color:var(--red);font-weight:bold}}
.url{{font-family:"SF Mono",monospace;font-size:11px;word-break:break-all}}
.url a{{color:var(--accent);text-decoration:none}}.url a:hover{{text-decoration:underline}}
.fix-url a{{color:var(--orange)!important}}.note{{font-size:11px;color:var(--muted)}}
.footer{{text-align:center;color:var(--muted);font-size:12px;padding-top:32px;border-top:1px solid var(--border);margin-top:48px}}
</style>
</head>
<body>
<div class="container">
<div class="hero">
<h1>🏆 2026世界杯时效性站点排查报告</h1>
<p>排查时间: {datetime.now().strftime('%Y-%m-%d %H:%M')} ｜ 总URL: {len(rows)} ｜ 有效: {valid} ｜ 非有效: {invalid} ｜ 已补充: {corrected} ｜ 并发拨测耗时: {elapsed:.0f}s</p>
</div>
<div class="stats">
<div class="sc g"><div class="n">{valid}</div><div class="l">有效资讯列表页</div></div>
<div class="sc r"><div class="n">{invalid}</div><div class="l">非有效/待替换</div></div>
<div class="sc"><div class="n">{cn_count}</div><div class="l">国内站点</div></div>
<div class="sc o"><div class="n">{intl_count}</div><div class="l">境外站点</div></div>
</div>
<h2>🇨🇳 国内站点 <span class="badge">{cn_count}个</span></h2>
<table>
<tr><th>#</th><th>站点</th><th>URL</th><th>HTTP</th><th>列表?</th><th>排查说明</th><th>补充URL</th></tr>
'''

for i, r in enumerate(cn_sites, 1):
    cls = 'bad' if not r['is_list'] else ''
    fix = 'fix' if r['corrected_url'] else ''
    st = 'ok' if r['is_list'] else 'no'
    curl = f'<span class="fix-url"><a href="{r["corrected_url"]}" target="_blank">{r["corrected_url"][:90]}</a></span>' if r['corrected_url'] else '-'
    html += f'<tr class="{cls} {fix}"><td>{i}</td><td>{r["site"]}</td><td class="url"><a href="{r["url"]}" target="_blank">{r["url"][:100]}</a></td><td>{r["http_status"]}</td><td class="{st}">{"✅" if r["is_list"] else "⚠️"}</td><td class="note">{r["probe_result"]}</td><td class="url">{curl}<br><span class="note">{r.get("corrected_note","")}</span></td></tr>\n'

html += f'''</table>
<h2>🌍 境外站点 <span class="badge">{intl_count}个</span></h2>
<table>
<tr><th>#</th><th>站点</th><th>URL</th><th>国家</th><th>HTTP</th><th>列表?</th><th>排查说明</th><th>补充URL</th></tr>
'''

for i, r in enumerate(intl_sites, 1):
    cls = 'bad' if not r['is_list'] else ''
    fix = 'fix' if r['corrected_url'] else ''
    st = 'ok' if r['is_list'] else 'no'
    curl = f'<span class="fix-url"><a href="{r["corrected_url"]}" target="_blank">{r["corrected_url"][:90]}</a></span>' if r['corrected_url'] else '-'
    html += f'<tr class="{cls} {fix}"><td>{i}</td><td>{r["site"]}</td><td class="url"><a href="{r["url"]}" target="_blank">{r["url"][:100]}</a></td><td>{r["country"]}</td><td>{r["http_status"]}</td><td class="{st}">{"✅" if r["is_list"] else "⚠️"}</td><td class="note">{r["probe_result"]}</td><td class="url">{curl}<br><span class="note">{r.get("corrected_note","")}</span></td></tr>\n'

html += f'''</table>
<div class="footer">Generated by AI · {datetime.now().strftime('%Y-%m-%d %H:%M')} · 2026 World Cup Timeliness Site Audit</div>
</div></body></html>'''

with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"   HTML: {OUTPUT_HTML}")

# Save JSON
with open('/Users/d.j.f/Documents/Claude/2026世界杯/probe_results.json', 'w') as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print("\n✅ 全部完成！")
