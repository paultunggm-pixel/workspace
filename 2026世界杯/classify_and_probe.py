#!/usr/bin/env python3
"""
世界杯时效性站点列表 — 分类 + 拨测 + 生成HTML
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
import re
import time
from urllib.parse import urlparse
import tldextract
from datetime import datetime

# ── 配置 ──
INPUT_PATH = '/Users/d.j.f/Documents/Claude/2026世界杯/世界杯时效性站点列表页/世界杯时效性站点列表页.xlsx'
OUTPUT_EXCEL = '/Users/d.j.f/Documents/Claude/2026世界杯/世界杯时效性站点分类排查结果.xlsx'
OUTPUT_HTML = '/Users/d.j.f/Documents/Claude/2026世界杯/世界杯时效性站点排查报告.html'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
}

# ── 境外域名→国家映射 ──
DOMAIN_COUNTRY_MAP = {
    '.uk': '英国', '.co.uk': '英国', '.de': '德国', '.fr': '法国', '.es': '西班牙',
    '.it': '意大利', '.pt': '葡萄牙', '.br': '巴西', '.au': '澳大利亚',
    '.jp': '日本', '.kr': '韩国', '.nl': '荷兰', '.se': '瑞典',
    '.ca': '加拿大', '.mx': '墨西哥', '.ar': '阿根廷', '.uy': '乌拉圭',
    '.ru': '俄罗斯', '.sa': '沙特阿拉伯', '.qa': '卡塔尔', '.ae': '阿联酋',
    '.eg': '埃及', '.ma': '摩洛哥', '.sn': '塞内加尔', '.gh': '加纳',
    '.ci': '科特迪瓦', '.tn': '突尼斯',
}

SITE_COUNTRY_OVERRIDE = {
    'espn.com': '美国', 'bbc.com': '英国', 'bbc.co.uk': '英国',
    'theguardian.com': '英国', 'goal.com': '英国', 'cbssports.com': '美国',
    'foxsports.com': '美国', 'sportsmole.co.uk': '英国', 'marca.com': '西班牙',
    'lequipe.fr': '法国', 'bild.de': '德国', 'sportingnews.com': '美国',
    'gazzetta.it': '意大利', 'si.com': '美国', 'actionnetwork.com': '美国',
    'fourfourtwo.com': '英国', 'theanalyst.com': '英国',
    'sofascore.com': '克罗地亚', 'reuters.com': '英国',
    'skysports.com': '英国', 'nytimes.com': '美国',
    'whoscored.com': '英国', 'beinsports.com': '法国',
    'expertpicks.com': '美国', 'leaguelane.com': '英国',
    'uefa.com': '瑞士', 'kickoff.guide': '英国', 'fifa.com': '瑞士',
    'apnews.com': '美国', 'bleacherreport.com': '美国',
    'newsnow.co.uk': '英国', '90min.com': '美国',
    'premierleague.com': '英国', 'football365.com': '英国',
    'hudl.com': '美国', 'planetfootball.com': '英国',
    'spielverlagerung.de': '德国', 'onefootball.com': '德国',
    'manchestereveningnews.co.uk': '英国', 'abola.pt': '葡萄牙',
    'record.pt': '葡萄牙', 'bundesliga.com': '德国',
    'sbs.com.au': '澳大利亚', 'tribalfootball.com': '英国',
    'worldsoccertalk.com': '美国', 'insideworldfootball.com': '英国',
    'mirror.co.uk': '英国', 'fotmob.com': '挪威',
    'as.com': '西班牙', 'independent.co.uk': '英国',
    'dailymail.com': '美国', 'mundodeportivo.com': '西班牙',
    'football-italia.net': '意大利', 'flashscore.com': '捷克',
    'sofascore.com': '克罗地亚',
}

# ── 国内站点域名特征 ──
CN_DOMAINS = {'.cn', 'sina.com.cn', '163.com', 'qq.com', 'sohu.com', 'zhibo8.com',
              'dongqiudi.com', 'hupu.com', 'ppsport.com', 'cctv.com', 'people.com.cn',
              'xinhuanet.com', 'ifeng.com', 'toutiao.com', 'miguvideo.com',
              'lesports.com', 'zhanqi.tv', 'quanmin.tv', 'leisu.com'}

def classify_site(name, url):
    """分类：国内/境外，境外补充国家"""
    if not url:
        return 'unknown', ''

    parsed = urlparse(url)
    domain = parsed.netloc.lower()
    # 去掉www/m前缀
    domain_clean = re.sub(r'^(www\.|m\.)', '', domain)

    # 检查是否国内
    for cn in CN_DOMAINS:
        if cn in domain or domain.endswith('.cn'):
            return '国内', ''

    # 境外 — 确定国家
    # 先查覆盖表
    for key, country in SITE_COUNTRY_OVERRIDE.items():
        if key in domain:
            return '境外', country

    # 查TLD表
    extracted = tldextract.extract(url)
    suffix = '.' + extracted.suffix if extracted.suffix else ''
    for tld, country in DOMAIN_COUNTRY_MAP.items():
        if suffix == tld:
            return '境外', country

    # 默认
    return '境外', '未知(国际)'

def probe_url(url, site_name):
    """拨测URL：返回 (http_status, redirect_url, is_list, verdict, snippet)"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        http_code = resp.status_code
        final_url = resp.url
        redirect = url if final_url == url else final_url

        content = resp.text
        content_lower = content.lower()

        # 检查是否是列表页特征
        # WC keywords
        wc_keywords_cn = ['世界杯', '世预赛', '国际足球', '全球', '足球新闻']
        wc_keywords_en = ['world cup', 'world-cup', 'worldcup', 'fifa world cup',
                          'fifa-world-cup', 'international', 'w杯']

        has_wc = any(kw in content_lower for kw in wc_keywords_en) or \
                 any(kw in content for kw in wc_keywords_cn)

        # 列表页特征
        list_indicators = [
            ('<article', '</article>'),
            ('class="article', ''),
            ('class="post', ''),
            ('class="news', ''),
            ('class="entry', ''),
            ('class="list', ''),
            ('class="feed', ''),
            ('<li class=', ''),
            ('class="timeline', ''),
            ('class="latest', ''),
            ('article_list', ''),
            ('news-list', ''),
            ('news_list', ''),
            ('articleList', ''),
            ('ul class="bbs-list-ul', ''),   # 虎扑
            ('class="news-item', ''),  # 网易/新浪
            ('class="hot-news', ''),
            ('class="content-list', ''),
        ]

        list_score = sum(1 for ind, _ in list_indicators if ind in content_lower or ind in content)

        # 检测是否是单篇文章（内容详情而非列表）
        single_indicators = ['article-detail', 'article_detail', 'single-', 'post-detail',
                             'class="article-body', 'article-body']
        is_single = any(ind in content_lower for ind in single_indicators)

        # 检测是否首页（而非列表）
        home_indicators = [
            ('<title>', 'home</title>'),
            ('<title>', ' - wikipedia</title>'),
        ]

        # 判断逻辑
        if '<title>' in content_lower:
            title_start = content_lower.find('<title>')
            title_end = content_lower.find('</title>', title_start)
            title = content[title_start:title_end] if title_start >= 0 and title_end > 0 else ''
        else:
            title = ''

        if is_single and list_score < 3:
            is_list = False
            reason = '疑似单篇详情页，非资讯列表'
        elif not has_wc and list_score < 3:
            is_list = False
            reason = '无世界杯相关内容且非列表结构'
        elif not has_wc:
            is_list = False
            reason = '内容非世界杯相关'
        elif list_score < 2:
            is_list = False
            reason = '页面缺乏列表页特征（article/list/news结构）'
        else:
            is_list = True
            reason = ''

        snippet = (title[:200] if title else '') + ' | ' + str(list_score) + ' list indicators'

        return http_code, redirect, is_list, reason, snippet[:400]

    except requests.Timeout:
        return 0, '', False, '请求超时(>15s)', ''
    except requests.ConnectionError:
        return 0, '', False, '连接失败（域名可能不可达）', ''
    except Exception as e:
        return 0, '', False, f'请求异常: {str(e)[:100]}', ''

def find_correct_url(site_name, original_url):
    """尝试找到正确的世界杯资讯列表页"""
    parsed = urlparse(original_url)
    base = f"{parsed.scheme}://{parsed.netloc}"

    # 可能的WC列表路径
    candidates = []
    name_lower = site_name.lower()

    # 世界杯相关路径模式
    wc_paths = [
        '/football/world-cup',
        '/football/world-cup-2026',
        '/football/fifa-world-cup',
        '/soccer/world-cup',
        '/sport/football/world-cup',
        '/football/worldcup',
        '/news/football/world-cup',
        '/sports/soccer/world-cup',
        '/football/competitions/fifa-world-cup',
        '/en/football/world-cup',
    ]

    # 新闻列表路径
    news_paths = [
        '/football/news',
        '/soccer/news',
        '/news/football',
        '/sports/football',
        '/football',
    ]

    # 先从wc路径测试
    for path in wc_paths + news_paths:
        test_url = base + path
        try:
            r = requests.get(test_url, headers=HEADERS, timeout=10)
            if r.status_code == 200:
                content = r.text.lower()
                if 'world cup' in content or 'world-cup' in content or '世界杯' in content:
                    # 检查是否是列表
                    list_score = sum(1 for ind, _ in
                        [('article', ''), ('class="news', ''), ('class="post', '')]
                        if ind in content)
                    if list_score >= 1:
                        return test_url, f'自动发现(HTTP{r.status_code}，含WC+列表特征)'
        except:
            continue

    return '', '未能自动发现有效列表页URL'

# ── 主流程 ──
print("=" * 60)
print("1. 加载 Excel...")
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb.active

rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name, url, date = row
    if name and url:
        rows.append({'site': name, 'url': str(url).strip(), 'date': str(date) if date else ''})

print(f"   共加载 {len(rows)} 个URL")

# ── 2. 分类 ──
print("\n2. 分类站点...")
for r in rows:
    cat, country = classify_site(r['site'], r['url'])
    r['category'] = cat
    r['country'] = country

cn_count = sum(1 for r in rows if r['category'] == '国内')
intl_count = sum(1 for r in rows if r['category'] == '境外')
print(f"   国内: {cn_count} | 境外: {intl_count}")

# ── 3. 拨测 ──
print("\n3. 拨测URL（可能需要几分钟）...")
for i, r in enumerate(rows):
    print(f"   [{i+1}/{len(rows)}] {r['site']:20s} {r['url'][:60]}")
    http_code, redirect, is_list, reason, snippet = probe_url(r['url'], r['site'])
    r['http_status'] = http_code
    r['redirect'] = redirect if redirect != r['url'] else ''
    r['is_list'] = is_list
    r['probe_result'] = reason
    r['snippet'] = snippet
    r['corrected_url'] = ''
    r['corrected_note'] = ''

    # 如果不是有效列表页，尝试找正确URL
    if not is_list and http_code != 0:
        print(f"      ⚠️ 非有效: {reason}")
        corrected, note = find_correct_url(r['site'], r['url'])
        if corrected:
            r['corrected_url'] = corrected
            r['corrected_note'] = note
            print(f"      ✓ 找到候选: {corrected}")
        else:
            r['corrected_note'] = '未找到有效替代URL'
    elif http_code == 0:
        print(f"      ❌ 不可达: {reason}")

    time.sleep(0.3)  # 避免被限速

valid = sum(1 for r in rows if r['is_list'])
invalid = sum(1 for r in rows if not r['is_list'])
print(f"\n   有效列表页: {valid} | 非有效: {invalid}")

# ── 4. 生成 Excel ──
print("\n4. 生成 Excel...")
wb_out = openpyxl.Workbook()

# Sheet 1: 国内站点
ws_cn = wb_out.active
ws_cn.title = "国内站点"

cn_headers = ['序号', '站点名称', 'URL', 'HTTP状态', '重定向', '是否资讯列表', '排查说明', '补充URL', '补充说明']
for c, h in enumerate(cn_headers, 1):
    ws_cn.cell(row=1, column=c, value=h)
    ws_cn.cell(row=1, column=c).font = Font(bold=True)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for c in range(1, len(cn_headers)+1):
    ws_cn.cell(row=1, column=c).fill = header_fill
    ws_cn.cell(row=1, column=c).font = header_font

row_idx = 2
for i, r in enumerate([r for r in rows if r['category'] == '国内'], 1):
    vals = [i, r['site'], r['url'], r['http_status'], r['redirect'],
            '✅' if r['is_list'] else '⚠️ 否', r['probe_result'], r['corrected_url'], r['corrected_note']]
    for c, v in enumerate(vals, 1):
        ws_cn.cell(row=row_idx, column=c, value=v)
    if not r['is_list']:
        for c in range(1, len(cn_headers)+1):
            ws_cn.cell(row=row_idx, column=c).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    row_idx += 1

cn_widths = [6, 18, 55, 10, 50, 12, 30, 55, 25]
for c, w in enumerate(cn_widths, 1):
    ws_cn.column_dimensions[get_column_letter(c)].width = w

# Sheet 2: 境外站点
ws_intl = wb_out.create_sheet("境外站点")
intl_headers = ['序号', '站点名称', 'URL', '所属国家', 'HTTP状态', '重定向', '是否资讯列表', '排查说明', '补充URL', '补充说明']

for c, h in enumerate(intl_headers, 1):
    ws_intl.cell(row=1, column=c, value=h)
    ws_intl.cell(row=1, column=c).font = Font(bold=True)
for c in range(1, len(intl_headers)+1):
    ws_intl.cell(row=1, column=c).fill = header_fill
    ws_intl.cell(row=1, column=c).font = header_font

row_idx = 2
for i, r in enumerate([r for r in rows if r['category'] == '境外'], 1):
    vals = [i, r['site'], r['url'], r['country'], r['http_status'], r['redirect'],
            '✅' if r['is_list'] else '⚠️ 否', r['probe_result'], r['corrected_url'], r['corrected_note']]
    for c, v in enumerate(vals, 1):
        ws_intl.cell(row=row_idx, column=c, value=v)
    if not r['is_list']:
        for c in range(1, len(intl_headers)+1):
            ws_intl.cell(row=row_idx, column=c).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    row_idx += 1

intl_widths = [6, 18, 55, 12, 10, 50, 12, 30, 55, 25]
for c, w in enumerate(intl_widths, 1):
    ws_intl.column_dimensions[get_column_letter(c)].width = w

# Sheet 3: 总览统计
ws_sum = wb_out.create_sheet("总览统计")
ws_sum.cell(row=1, column=1, value="世界杯时效性站点排查统计").font = Font(bold=True, size=14)
ws_sum.cell(row=2, column=1, value=f"排查时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
ws_sum.cell(row=3, column=1, value=f"总URL数: {len(rows)}")
ws_sum.cell(row=4, column=1, value=f"国内站点: {cn_count} | 有效: {sum(1 for r in rows if r['category']=='国内' and r['is_list'])} | 非有效: {sum(1 for r in rows if r['category']=='国内' and not r['is_list'])}")
ws_sum.cell(row=5, column=1, value=f"境外站点: {intl_count} | 有效: {sum(1 for r in rows if r['category']=='境外' and r['is_list'])} | 非有效: {sum(1 for r in rows if r['category']=='境外' and not r['is_list'])}")
ws_sum.cell(row=6, column=1, value=f"已补充URL: {sum(1 for r in rows if r['corrected_url'])}")

wb_out.save(OUTPUT_EXCEL)
print(f"   Excel 已保存: {OUTPUT_EXCEL}")

# ── 5. 生成 HTML ──
print("\n5. 生成 HTML...")
cn_sites = [r for r in rows if r['category'] == '国内']
intl_sites = [r for r in rows if r['category'] == '境外']

html_content = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>2026世界杯时效性站点排查报告</title>
<style>
:root {{
  --bg: #0f1117; --card: #1a1d27; --border: #2a2d3a;
  --text: #e1e4eb; --muted: #888b96; --accent: #5b7fff;
  --green: #00c853; --red: #ff5252; --orange: #ff9100;
  --yellow: #ffd600;
}}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, "PingFang SC", "Microsoft YaHei", sans-serif; background:var(--bg); color:var(--text); line-height:1.7; }}
.container {{ max-width:1200px; margin:0 auto; padding:32px 24px 80px; }}
.hero {{ text-align:center; padding:48px 0 32px; border-bottom:1px solid var(--border); margin-bottom:40px; }}
.hero h1 {{ font-size:28px; margin-bottom:8px; }}
.hero p {{ color:var(--muted); font-size:14px; }}

.stats {{ display:flex; gap:16px; margin-bottom:40px; flex-wrap:wrap; }}
.stat-card {{ flex:1; min-width:140px; background:var(--card); border:1px solid var(--border); border-radius:12px; padding:20px; text-align:center; }}
.stat-card .num {{ font-size:36px; font-weight:700; }}
.stat-card .label {{ color:var(--muted); font-size:13px; margin-top:4px; }}
.stat-card.green .num {{ color:var(--green); }}
.stat-card.red .num {{ color:var(--red); }}
.stat-card.orange .num {{ color:var(--orange); }}

.section {{ margin-bottom:48px; }}
.section h2 {{ font-size:20px; margin-bottom:20px; padding-bottom:8px; border-bottom:2px solid var(--accent); display:flex; align-items:center; gap:10px; }}
.section h2 .badge {{ font-size:12px; padding:2px 10px; border-radius:12px; background:var(--accent); color:#fff; }}

table {{ width:100%; border-collapse:collapse; font-size:13px; }}
th {{ background:#252836; text-align:left; padding:10px 12px; font-weight:600; position:sticky; top:0; z-index:1; }}
td {{ padding:9px 12px; border-bottom:1px solid var(--border); vertical-align:top; }}
tr:hover td {{ background:rgba(91,127,255,0.05); }}
tr.invalid td {{ background:rgba(255,82,82,0.08); }}
.status-ok {{ color:var(--green); font-weight:bold; }}
.status-bad {{ color:var(--red); font-weight:bold; }}
.status-warn {{ color:var(--orange); }}

.url {{ font-family:"SF Mono",monospace; font-size:12px; word-break:break-all; color:var(--accent); }}
.corrected {{ background:rgba(255,214,0,0.08); }}
.corrected td {{ border-left:3px solid var(--yellow); }}
.note {{ font-size:12px; color:var(--muted); }}

.footer {{ text-align:center; color:var(--muted); font-size:12px; padding-top:40px; border-top:1px solid var(--border); margin-top:40px; }}
</style>
</head>
<body>
<div class="container">

<div class="hero">
  <h1>🏆 2026世界杯时效性站点排查报告</h1>
  <p>排查时间: {datetime.now().strftime('%Y-%m-%d %H:%M')} ｜ 总URL: {len(rows)} 个 ｜ 有效: {valid} ｜ 非有效: {invalid}</p>
</div>

<div class="stats">
  <div class="stat-card green"><div class="num">{valid}</div><div class="label">有效资讯列表页</div></div>
  <div class="stat-card red"><div class="num">{invalid}</div><div class="label">非有效/待替换</div></div>
  <div class="stat-card"><div class="num">{cn_count}</div><div class="label">国内站点</div></div>
  <div class="stat-card orange"><div class="num">{intl_count}</div><div class="label">境外站点</div></div>
</div>

<div class="section">
  <h2>🇨🇳 国内站点 <span class="badge">{cn_count}个</span></h2>
  <table>
  <tr><th>#</th><th>站点</th><th>URL</th><th>HTTP</th><th>列表页?</th><th>排查说明</th><th>补充URL</th></tr>
'''

for i, r in enumerate(cn_sites, 1):
    status_class = 'status-ok' if r['is_list'] else 'status-bad'
    row_class = '' if r['is_list'] else 'invalid'
    corrected_row = ' corrected' if r['corrected_url'] else ''

    html_content += f'''
  <tr class="{row_class}{corrected_row}">
    <td>{i}</td>
    <td>{r['site']}</td>
    <td class="url"><a href="{r['url']}" target="_blank" style="color:var(--accent)">{r['url'][:80]}{"..." if len(r['url'])>80 else ""}</a></td>
    <td>{r['http_status']}</td>
    <td class="{status_class}">{'✅ 是' if r['is_list'] else '⚠️ 否'}</td>
    <td class="note">{r['probe_result']}</td>
    <td class="url">{'<a href="'+r['corrected_url']+'" target="_blank" style="color:var(--orange)">'+r['corrected_url'][:80]+'</a>' if r['corrected_url'] else '-'}<br><span class="note">{r['corrected_note'] if r['corrected_url'] else ''}</span></td>
  </tr>'''

html_content += '''
  </table>
</div>

<div class="section">
  <h2>🌍 境外站点 <span class="badge">''' + str(intl_count) + '''个</span></h2>
  <table>
  <tr><th>#</th><th>站点</th><th>URL</th><th>国家</th><th>HTTP</th><th>列表页?</th><th>排查说明</th><th>补充URL</th></tr>
'''

for i, r in enumerate(intl_sites, 1):
    status_class = 'status-ok' if r['is_list'] else 'status-bad'
    row_class = '' if r['is_list'] else 'invalid'
    corrected_row = ' corrected' if r['corrected_url'] else ''

    html_content += f'''
  <tr class="{row_class}{corrected_row}">
    <td>{i}</td>
    <td>{r['site']}</td>
    <td class="url"><a href="{r['url']}" target="_blank" style="color:var(--accent)">{r['url'][:80]}{"..." if len(r['url'])>80 else ""}</a></td>
    <td>{r['country']}</td>
    <td>{r['http_status']}</td>
    <td class="{status_class}">{'✅ 是' if r['is_list'] else '⚠️ 否'}</td>
    <td class="note">{r['probe_result']}</td>
    <td class="url">{'<a href="'+r['corrected_url']+'" target="_blank" style="color:var(--orange)">'+r['corrected_url'][:80]+'</a>' if r['corrected_url'] else '-'}<br><span class="note">{r['corrected_note'] if r['corrected_url'] else ''}</span></td>
  </tr>'''

html_content += f'''
  </table>
</div>

<div class="footer">
  Generated by AI · {datetime.now().strftime('%Y-%m-%d %H:%M')} · 2026 World Cup Timeliness Site Audit
</div>

</div>
</body>
</html>
'''

with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"   HTML 已保存: {OUTPUT_HTML}")

# ── 6. 保存JSON中间数据 ──
import json
json_path = '/Users/d.j.f/Documents/Claude/2026世界杯/probe_results.json'
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)
print(f"   中间数据: {json_path}")

print("\n✅ 全部完成！")
