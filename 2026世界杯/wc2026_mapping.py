import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = "/Users/d.j.f/Documents/Claude/claude code 2026 worldcup.xlsx"
wb = openpyxl.load_workbook(path)

# Styles
header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ai_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
external_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
none_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
match_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
team_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
player_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
cell_font = Font(name="Microsoft YaHei", size=9)
bold_font = Font(name="Microsoft YaHei", size=9, bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# Complete mapping: row -> (快鱼模块, 数据字段, 字段注释)
# ============================================================
# Module abbreviations:
#   M1 = 赛程列表快鱼结构
#   M2 = 单个比赛快鱼结构
#   M3 = 榜单数据结构
#   M4 = 球员详细信息快鱼结构
#   M5 = 球队详细信息快鱼结构
#   M6 = 赛季列表快鱼结构
#   EXT = 外部数据源
#   AI = AI综合生成

data_rows = [
    # R1
    ("预测结论","P0","主队名/客队名","-","是","是","参考世界杯专题数据包-比赛包",
     "赛程列表快鱼结构","team_infos[].team_name","队伍名字，team_infos数组 home=1为主队/home=0为客队"),
    # R2
    ("预测结论","P0","比赛时间","精确到分钟","是","是","参考世界杯专题数据包-比赛包",
     "赛程列表快鱼结构","_product_info.valid_date.start_at","比赛开始时间，UTC+8格式精确到秒"),
    # R3
    ("预测结论","P0","比赛所属小组","如'E组'","是","是","参考世界杯专题数据包-比赛包",
     "单个比赛快鱼结构","_extra.group","比赛分组，如'小组赛A组''B组'等"),
    # R4
    ("预测结论","P0","预测胜负结果","主胜/平/客胜","否","是","综合生成字段",
     "AI综合生成","—","大模型综合推理生成，不直接取快鱼字段；纳米盘口数据作辅助输入"),
    # R5
    ("预测结论","P0","预测比分","如'4-0'","否","是","综合生成字段",
     "AI综合生成","—","大模型综合推理生成，结合球队攻防统计和盘口数据"),
    # R6
    ("预测结论","P0","信心指数","1-5星","否","是","综合生成字段",
     "AI综合生成","—","大模型根据实力差距、赔率隐含概率综合评估"),
    # R7
    ("预测结论","P0","预测锐评文案","一句话","否","否","综合生成字段",
     "AI综合生成","—","大模型根据比赛背景和预测结果生成有态度的锐评"),
    # R8
    ("胜负理由","P0","FIFA排名","双方","是","是","",
     "球队详细信息快鱼结构","_product_info.extra.fifa_ranking","FIFA世界排名，字符串类型如'6'"),
    # R9
    ("胜负理由","P0","Elo评分","双方","否","否","直接参照https://www.eloratings.net/",
     "外部数据源","eloratings.net","非快鱼数据，需从外部网站抓取，Elo评分为整数"),
    # R10
    ("胜负理由","P0","全队总身价","欧元","是","是","",
     "球队详细信息快鱼结构","_product_info.extra.values","球队球员总身价，单位欧元，字符串如'911250000'"),
    # R11
    ("胜负理由","P0","近N场战绩","胜/平/负","是","是","观察接口目前近2年信息",
     "单个比赛快鱼结构","_extra.team_latest_history_matches.{hTeam|vTeam}","victory(胜)/flat(平)/lost(负) + items[]近场明细"),
    # R12
    ("胜负理由","P0","近N场场均进球","-","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.statistics.points_items[].gfAverage","平均进球数，如'1.1'"),
    # R13
    ("胜负理由","P0","近N场场均失球","-","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.statistics.points_items[].gaAverage","平均失球数，如'0.55'"),
    # R14
    ("胜负理由","P0","最新伤病/停赛名单","哪些核心球员缺阵","否","否","纳米有接口但未接入",
     "球员详细信息快鱼结构","_product_info.extra.injuries[]","伤病列表: team_name/injury/start_time/end_time"),
    # R15
    ("胜负理由","P0","最新大名单/首发预测","26人名单+预计首发11人","否","否","有大名单无首发预测",
     "单个比赛快鱼结构","_extra.team_infos[].lineup.members[]","球员名单: number/playerName/position/player_content_id"),
    # R16
    ("胜负理由","P0","核心球员姓名","双方各1-2人","否","否","可人工标注",
     "球员详细信息快鱼结构","title 或 _product_info.extra.name","球员名称，title为快鱼标题，name为详细字段"),
    # R17
    ("胜负理由","P0","核心球员身价","欧元","是","是","有基础信息需要汇总",
     "球员详细信息快鱼结构","_product_info.extra.transfers[].fee","转会费(欧元)，反映球员市场身价，type=1为转会"),
    # R18
    ("胜负理由","P0","核心球员赛季进球/助攻","俱乐部+国家队","是","是","有基础信息需要汇总",
     "球员详细信息快鱼结构","_product_info.extra.statistics.{total|national_items}[]","进球goals/助攻assists/出场次数等，分联赛/杯赛/国家队/总计"),
    # R19
    ("胜负理由","P1","核心球员年龄","-","是","是","有基础信息需要汇总",
     "球员详细信息快鱼结构","_product_info.extra.age","球员年龄，字符串如'29'"),
    # R20
    ("胜负理由","P1","核心球员效力俱乐部","-","是","是","有基础信息需要汇总",
     "球员详细信息快鱼结构","_product_info.extra.team_name","现效力俱乐部名称"),
    # R21
    ("胜负理由","P1","核心球员位置","-","是","是","有基础信息需要汇总",
     "球员详细信息快鱼结构","_product_info.extra.position","场上位置，如'前锋''后卫''中场''守门员'"),
    # R22
    ("胜负理由","P1","核心球员国家队近N场进球/助攻","-","是","是","有基础信息需要汇总",
     "球员详细信息快鱼结构","_product_info.extra.statistics.national_items[]","国家队数据: goals/assists/出场(ptime)等"),
    # R23
    ("胜负理由","P1","核心球员身高","cm","是","是","有基础信息需要汇总",
     "球员详细信息快鱼结构","_product_info.extra.height","身高，单位cm，字符串如'175'"),
    # R24
    ("胜负理由","P1","核心球员最新伤病/体能状态","新闻来源","否","否","",
     "球员详细信息快鱼结构","_product_info.extra.injuries[]","当前伤病信息: injury描述/start_time/end_time"),
    # R25
    ("胜负理由","P1","核心球员关键传球/射门/过人","细分数据","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.stats[].info[]","球员统计: keyPass(关键传球)/shoot(射门)/dribbles(过人)/dribblesSuccess(过人成功)"),
    # R26
    ("胜负理由","P1","球队平均身高","cm","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.members[].height","需遍历members聚合计算平均身高(非直接字段)"),
    # R27
    ("胜负理由","P1","球队所属大洲/联赛级别","如'中北美''五大联赛'","否","否","是否可通用获取？",
     "球队详细信息快鱼结构","_product_info.extra.league","所在联赛，如'英超'；需结合country字段推导大洲"),
    # R28
    ("胜负理由","P1","世界杯参赛次数","历史","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.statistics.points_items[]","需筛选league='世界杯'的历史赛季记录，计次"),
    # R29
    ("胜负理由","P1","世界杯历史最佳成绩","如'八强''冠军'","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.honor[]","荣誉列表，需筛选世界杯相关title，取最高排名"),
    # R30
    ("胜负理由","P1","国家人口","-","否","否","是否可通用获取？",
     "外部数据源","世界银行/维基百科","非快鱼数据，需外部补充，用于生成趣味类比"),
    # R31
    ("胜负理由","P1","核心球员效力联赛分布","几人在五大联赛","是","是","有基础信息需要汇总",
     "球员详细信息快鱼结构","_product_info.extra.active_league[]","球员参与赛事列表，统计五大联赛(英超/西甲/德甲/意甲/法甲)人数"),
    # R32
    ("胜负理由","P1","近期与已知球队的交手结果","如'3月被国足2-0'","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.team_latest_history_matches.{hTeam|vTeam}.items[]","近场: hTeam/vTeam/matchDate/matchName/score，用于提取特定参照系"),
    # R33
    ("胜负理由","P1","对弱旅/同档次球队的场均进球","区分整体场均","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.statistics.points_items[].gf","总进球gf和场次session可计算场均，需按对手实力过滤"),
    # R34
    ("胜负理由","P1","与对手的历史交锋记录","H2H","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.battle_history.list[]","交锋历史: matchDate/hTeam/vTeam/score/matchName"),
    # R35
    ("胜负理由","P1","常用阵型","如'4-2-3-1'","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.team_infos[].lineup.formation","比赛阵型，如'4-2-3-1''4-3-3'"),
    # R36
    ("胜负理由","P1","控球率","近N场平均","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.stats[].poct","控球率( possession % )，如0.58=58%；可跨场平均"),
    # R37
    ("胜负理由","P1","赛前主教练发布会要点","摘要/关键语录","否","否","纳米有接口但未接入",
     "外部数据源","纳米数据接口","非快鱼数据，需从纳米新闻接口获取赛前发布会摘要"),
    # R38
    ("胜负理由","P1","赛前球队新闻热点","如'主教练辞职''球员内讧'","否","否","纳米有接口但未接入",
     "外部数据源","纳米数据接口","非快鱼数据，需从纳米新闻/舆情接口获取突发事件"),
    # R39
    ("胜负理由","P1","球队备战动态","如'提前3天到达适应海拔'","否","否","纳米有接口但未接入",
     "外部数据源","纳米数据接口","非快鱼数据，需从纳米新闻接口获取备战信息"),
    # R40
    ("胜负理由","P1","核心球员最新状态","如'孙兴慜训练中扭伤脚踝'","否","否","纳米有接口但未接入",
     "外部数据源","纳米数据接口","非快鱼数据，需从纳米新闻接口获取球员实时状态"),
    # R41
    ("胜负理由","P1","小组赛当前积分/排名","如果非首轮","是","是","",
     "球队详细信息快鱼结构","_product_info.extra.statistics.points_items[].{ranking|points}","当前排名ranking/积分points/胜victory/平flat/负lost，需指定赛季"),
    # R42
    ("胜负理由","P1","小组赛形势分析","出线/淘汰的条件","否","否","",
     "AI综合生成","—","大模型根据小组积分/净胜球/剩余赛程综合推理生成"),
    # R43
    ("胜负理由","P2","两队赛程密度","距上一场比赛间隔天数","是","是","",
     "单个比赛快鱼结构","_extra.future_matches.{hTeam|vTeam}.items[]","未来赛程matchDate，与当前比赛日期计算间隔天数"),
    # R44
    ("胜负理由","P2","知名球星身价/年收入","如C罗年收入3亿美元","否","否","有身价无收入",
     "外部数据源","福布斯/公开报道","非快鱼数据，身价可从转会费推算，年收入需外部数据"),
    # R45
    ("胜负理由","P2","球队风格标签","如'防反''控球''铁桶阵'","否","否","可人工标注",
     "AI综合生成","—","可基于stats统计(控球率poct/传球pass)推断风格，或人工标注"),
    # R46
    ("胜负理由","P2","主教练姓名/背景","-","否","是","快鱼仅基础信息",
     "单个比赛快鱼结构","_extra.team_infos[].lineup.coach","主教练姓名，如'Matthias Jaissle'，纳米侧有更多背景信息"),
    # R47
    ("胜负理由","P2","参赛国体育媒体赛前报道","当地媒体预测/态度","否","否","纳米有接口但未接入",
     "外部数据源","纳米数据接口","非快鱼数据，需从纳米新闻/媒体接口获取当地报道"),
    # R48
    ("胜负理由","P2","参赛国社交媒体热度","话题讨论量/情绪倾向","否","否","",
     "外部数据源","社交媒体API/自建爬虫","非快鱼数据，需外部获取社交热度/情绪分析"),
    # R49
    ("胜负理由","P2","赛前赔率/市场预期","不直接展示但转化为'市场信心'","否","是","赛前赔率有纳米数据",
     "外部数据源","纳米数据接口（赔率/盘口）","非快鱼数据，纳米接口提供欧赔/亚盘等，可转化为信心指数"),
    # R50
    ("比分模块","P0","双方场均进球","-","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.statistics.points_items[].gfAverage","平均进球数，需与场次session结合计算近N场均值"),
    # R51
    ("比分模块","P0","双方场均失球","-","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.statistics.points_items[].gaAverage","平均失球数，需与场次session结合计算近N场均值"),
    # R52
    ("比分模块","P1","首粒进球平均出现时间","如'第63分钟'","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.events[]","遍历events[]取第一个eventCode为进球的time字段，跨场平均"),
    # R53
    ("比分模块","P1","上半场/下半场进球占比","-","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.stats[].p1{} / p2{}","上半场p1和下半场p2中的goals字段，计算占比"),
    # R54
    ("比分模块","P1","先进球后拿分率","-","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.events[]","遍历events[]结合score判断首球归属+最终胜负，跨场统计"),
    # R55
    ("比分模块","P1","被先进球后扳平/逆转率","-","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.events[]","遍历events[]判断先失球场景下最终比分走向，跨场统计"),
    # R56
    ("比分模块","P1","进球方式分布","运动战/定位球/反击/点球占比","否","否","",
     "单个比赛快鱼结构","_extra.events[].eventCode / eventCodeName","通过事件代码区分: 点球(特定code)/定位球/运动战，需分类后统计占比"),
    # R57
    ("比分模块","P1","近N场总进球≤2球的比赛占比","-","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.team_latest_history_matches.{hTeam|vTeam}.items[].score","遍历近场score[0]+score[1]≤2的比例"),
    # R58
    ("比分模块","P1","近N场0-0比赛占比","-","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.team_latest_history_matches.{hTeam|vTeam}.items[].score","遍历近场score[0]=0且score[1]=0的比例"),
    # R59
    ("比分模块","P2","射正率","双方","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.stats[].shoot / shootR","射正率 = shootR(射正) / shoot(射门)，跨场平均"),
    # R60
    ("比分模块","P2","xG（预期进球）","近N场","否","否","",
     "外部数据源","Opta/StatsBomb等专业数据商","非快鱼数据，需从专业足球数据商获取xG数据"),
    # R61
    ("比分模块","P2","角球/任意球场均次数","-","否","否","",
     "单个比赛快鱼结构","_extra.stats[].corner / freekick","角球corner和任意球freekick次数，跨场取平均"),
    # R62
    ("比分模块","P2","点球获得率","近N场获得点球次数","否","否","",
     "单个比赛快鱼结构","_extra.stats[].penalty","点球penalty次数，遍历近场计算场均"),
    # R63
    ("比分模块","P2","近N场被零封次数","-","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.team_latest_history_matches.{hTeam|vTeam}.items[].score","己方score为0的场次数"),
    # R64
    ("场外信息","P0","比赛城市","-","是","是","",
     "单个比赛快鱼结构","_product_info.extra.venue 或从title解析","比赛地点，奥运会/全运会场馆字段为venue；足球可从title解析"),
    # R65
    ("场外信息","P0","比赛球场名称","-","是","是","",
     "球队详细信息快鱼结构","_product_info.extra.home_court","主队球场名称，需与比赛城市匹配确定比赛球场"),
    # R66
    ("场外信息","P1","球场类型","专业足球场/改建球场/有无顶棚","否","是","纳米接口有",
     "外部数据源","纳米数据接口（球场信息）","非快鱼数据，纳米接口提供球场详细信息含类型/顶棚"),
    # R67
    ("场外信息","P1","球场海拔","米","否","否","",
     "外部数据源","地理数据/维基百科","非快鱼数据，需外部获取球场所在城市海拔高度"),
    # R68
    ("场外信息","P1","比赛日气温","°C","是","是","",
     "单个比赛快鱼结构","_product_info.extra.environment.temperature","场地环境温度，如'12°C'；仅奥运类结构明确有此字段，足球需确认"),
    # R69
    ("场外信息","P1","比赛日湿度","%","是","是","",
     "单个比赛快鱼结构","_product_info.extra.environment.humidity","场地环境湿度，如'72%'；仅奥运类结构明确有此字段，足球需确认"),
    # R70
    ("场外信息","P1","比赛日降雨概率","%","是","是","",
     "外部数据源","天气API（OpenWeather/和风等）","快鱼environment字段仅含weather/temperature/humidity，降雨概率需外部天气API"),
    # R71
    ("场外信息","P2","草坪类型","天然草/人工草/临时铺设","否","是","纳米接口有",
     "外部数据源","纳米数据接口（球场信息）","非快鱼数据，纳米接口提供球场草坪类型信息"),
    # R72
    ("场外信息","P2","球场容量","-","否","否","可查",
     "球队详细信息快鱼结构","_product_info.extra.home_court_seats","主队球场座位数，如'74879'；世界杯球场可能不同于俱乐部主场"),
    # R73
    ("场外信息","P2","球队在当地的集训天数","提前几天到达适应","否","否","可查",
     "外部数据源","纳米数据接口（新闻/备战）","非快鱼数据，需从新闻接口获取球队到达/集训时间线"),
    # R74
    ("场外信息","P1","主裁判姓名/国籍","-","","","不确定",
     "单个比赛快鱼结构","_extra.events[]中涉及裁判的event，或需纳米补充","快鱼events中可抽取裁判相关事件；完整裁判信息需纳米数据"),
    # R75
    ("场外信息","P2","主裁场均黄牌/红牌","-","","","不确定",
     "外部数据源","纳米数据接口（裁判统计）","非快鱼数据，需纳米裁判统计接口获取历史执法数据"),
    # R76
    ("场外信息","P2","主裁场均点球判罚","-","","","不确定",
     "外部数据源","纳米数据接口（裁判统计）","非快鱼数据，需纳米裁判统计接口获取点球判罚频率"),
    # R77
    ("场外信息","P2","主裁历史执法该球队的记录","-","","","不确定",
     "外部数据源","纳米数据接口（裁判vs球队）","非快鱼数据，需交叉查询裁判执法记录与该球队的匹配"),
    # R78
    ("场外信息","P1","世界杯历史首轮战绩","-","是","是","有基础信息需要汇总",
     "球队详细信息快鱼结构","_product_info.extra.honor[] + statistics.points_items[]","从荣誉和积分中筛选世界杯首轮相关历史数据"),
    # R79
    ("场外信息","P1","历史世界杯爆冷案例库","如'2022沙特胜阿根廷'","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.battle_history.list[]","历史交锋中筛选排名/身价悬殊但弱者胜的案例"),
    # R80
    ("场外信息","P2","相似场景历史结果","如'2006德国首轮也碰中北美队'","是","是","有基础信息需要汇总",
     "单个比赛快鱼结构","_extra.team_latest_history_matches.{hTeam|vTeam}.items[]","从历史战绩中匹配对手大洲+轮次相似的场景"),
    # R81
    ("互动","P1","同组其他比赛信息","-","是","是","",
     "赛程列表快鱼结构","_extra.related + group匹配","同组筛选: 从赛程列表中按group字段过滤同组其他场次"),
    # R82
    ("互动","P1","球队下一场对手/时间","-","是","是","",
     "单个比赛快鱼结构","_extra.future_matches.{hTeam|vTeam}.items[]","未来赛程matchDate/hTeam/vTeam，取当前比赛后最近一场"),
    # R83
    ("互动","P2","当前热门话题/争议","如'梅西是否最后一届'","否","否","",
     "AI综合生成","—","大模型根据球员年龄/赛事阶段/社交媒体热度综合分析生成"),
]

# ============================================================
# Create sheet
# ============================================================
ws = wb.create_sheet("数据字段映射表")

headers = ["模块","优先级","字段","说明","UC接口","纳米接口","备注",
           "所在快鱼数据模块","具体使用数据字段","字段含义注释"]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# Column widths
widths = [10, 6, 28, 24, 8, 8, 28, 24, 42, 48]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

for i, d in enumerate(data_rows):
    row = i + 2
    # d = (模块, 优先级, 字段, 说明, UC, 纳米, 备注, 快鱼模块, 数据字段, 字段注释)
    for col, val in enumerate(d, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = cell_font
        cell.alignment = center if col <= 7 else left_align
        cell.border = thin_border

    # Color coding for last 3 columns
    module = d[7]  # 所在快鱼数据模块
    if "AI综合" in module:
        for c in [8,9,10]:
            ws.cell(row=row, column=c).fill = ai_fill
    elif "外部数据" in module:
        for c in [8,9,10]:
            ws.cell(row=row, column=c).fill = external_fill
    elif "赛程列表" in module:
        for c in [8,9,10]:
            ws.cell(row=row, column=c).fill = match_fill
    elif "球队详细" in module:
        for c in [8,9,10]:
            ws.cell(row=row, column=c).fill = team_fill
    elif "球员详细" in module:
        for c in [8,9,10]:
            ws.cell(row=row, column=c).fill = player_fill

ws.freeze_panes = "H2"
ws.auto_filter.ref = f"A1:J{len(data_rows)+1}"

# Add legend rows at bottom
lr = len(data_rows) + 3
ws.merge_cells(f'A{lr}:J{lr}')
c = ws.cell(row=lr, column=1, value="颜色图例：🟢 绿底=AI综合生成 | 🟠 橙底=外部数据源(纳米/天气API/公开数据) | 🔵 蓝底=赛程列表快鱼结构 | 🔷 深蓝底=球队详细信息快鱼结构 | 🟡 黄底=球员详细信息快鱼结构 | ⬜ 灰底=快鱼不可用")
c.font = Font(name="Microsoft YaHei", size=9)
c.alignment = left_align

lr += 1
ws.merge_cells(f'A{lr}:J{lr}')
c = ws.cell(row=lr, column=1, value="数据来源分布：快鱼可直接获取 ≈55项 | 需外部数据源 ≈22项 | AI综合生成 ≈6项")
c.font = Font(name="Microsoft YaHei", size=9, bold=True)
c.alignment = left_align

# Save
wb.save(path)
print(f"✅ 数据字段映射表已添加至: {path}")
print(f"共 {len(data_rows)} 行数据")

# Count by module
from collections import Counter
modules = Counter(d[7] for d in data_rows)
print("\n模块分布:")
for mod, count in modules.most_common():
    print(f"  {mod}: {count} 项")
