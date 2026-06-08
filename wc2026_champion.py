import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load existing file
path = "/Users/d.j.f/Documents/Claude/2026世界杯完整赛程预判.xlsx"
wb = openpyxl.load_workbook(path)

# Styles
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
champion_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
tier1_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
tier2_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
elim_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
cell_font = Font(name="Microsoft YaHei", size=10)
bold_font = Font(name="Microsoft YaHei", size=10, bold=True)
big_font = Font(name="Microsoft YaHei", size=14, bold=True, color="C00000")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# Simulated tournament progression
# ============================================================
# Step 1: Group stage results (predicted)
groups_result = {
    "A": [("Mexico", 1), ("South Korea", 2), ("Czechia", 3), ("South Africa", 4)],
    "B": [("Switzerland", 1), ("Canada", 2), ("Bosnia & Herzegovina", 3), ("Qatar", 4)],
    "C": [("Brazil", 1), ("Morocco", 2), ("Scotland", 3), ("Haiti", 4)],
    "D": [("USA", 1), ("Turkey", 2), ("Australia", 3), ("Paraguay", 4)],
    "E": [("Germany", 1), ("Ecuador", 2), ("Ivory Coast", 3), ("Curaçao", 4)],
    "F": [("Netherlands", 1), ("Japan", 2), ("Sweden", 3), ("Tunisia", 4)],
    "G": [("Belgium", 1), ("Iran", 2), ("Egypt", 3), ("New Zealand", 4)],
    "H": [("Spain", 1), ("Uruguay", 2), ("Saudi Arabia", 3), ("Cape Verde", 4)],
    "I": [("France", 1), ("Norway", 2), ("Senegal", 3), ("Iraq", 4)],
    "J": [("Argentina", 1), ("Austria", 2), ("Algeria", 3), ("Jordan", 4)],
    "K": [("Portugal", 1), ("Colombia", 2), ("Uzbekistan", 3), ("DR Congo", 4)],
    "L": [("England", 1), ("Croatia", 2), ("Ghana", 3), ("Panama", 4)],
}

# Best 8 third-place teams (ranked by expected points/GD)
best_thirds = ["Sweden", "Senegal", "Scotland", "Egypt", "Ivory Coast", "Czechia", "Saudi Arabia", "Ghana"]

# Step 2: Round of 32 (M73-M88)
# Mapping: match_id -> (team_a, team_b, winner, reasoning)
r32 = [
    ("M73", "South Korea", "Canada", "South Korea", "韩国 T2 vs 加拿大 T3，韩国整体实力和大赛经验占优"),
    ("M74", "Germany", "Sweden", "Germany", "德国 T1 vs 瑞典 T2，卫冕冠军实力碾压"),
    ("M75", "Netherlands", "Morocco", "Netherlands", "荷兰 T1 vs 摩洛哥 T2，荷兰中前场火力更强"),
    ("M76", "Brazil", "Japan", "Brazil", "巴西 T1 vs 日本 T2，桑巴军团个人能力完胜"),
    ("M77", "France", "Senegal", "France", "法国 T1 vs 塞内加尔 T2，法国阵容深度碾压"),
    ("M78", "Ecuador", "Norway", "Norway", "厄瓜多尔 T2 vs 挪威 T2，哈兰德决定比赛"),
    ("M79", "Mexico", "Scotland", "Mexico", "墨西哥 T2 vs 苏格兰 T3，东道主优势+经验"),
    ("M80", "England", "Egypt", "England", "英格兰 T1 vs 埃及 T3，实力差距明显"),
    ("M81", "USA", "Ivory Coast", "USA", "美国 T2 vs 科特迪瓦 T3，东道主+整体性更优"),
    ("M82", "Belgium", "Czechia", "Belgium", "比利时 T2 vs 捷克 T3，黄金一代最后一舞"),
    ("M83", "Colombia", "Croatia", "Croatia", "哥伦比亚 T2 vs 克罗地亚 T2，克罗地亚大赛基因更强"),
    ("M84", "Spain", "Austria", "Spain", "西班牙 T1 vs 奥地利 T2，传控体系压制"),
    ("M85", "Switzerland", "Saudi Arabia", "Switzerland", "瑞士 T2 vs 沙特 T3，瑞士防守反击稳健"),
    ("M86", "Argentina", "Uruguay", "Argentina", "阿根廷 T1 vs 乌拉圭 T2，卫冕冠军+梅西加成"),
    ("M87", "Portugal", "Ghana", "Portugal", "葡萄牙 T1 vs 加纳 T3，C罗最后一届世界杯"),
    ("M88", "Turkey", "Iran", "Turkey", "土耳其 T3 vs 伊朗 T3，土耳其整体战力稍强"),
]

# Step 3: Round of 16
# winner_of("M73") vs winner_of("M75") -> M90, etc.
r16_map = {
    "M89": ("M74胜者", "M77胜者"),  # Germany vs France
    "M90": ("M73胜者", "M75胜者"),  # South Korea vs Netherlands
    "M91": ("M76胜者", "M78胜者"),  # Brazil vs Norway
    "M92": ("M79胜者", "M80胜者"),  # Mexico vs England
    "M93": ("M83胜者", "M84胜者"),  # Croatia vs Spain
    "M94": ("M81胜者", "M82胜者"),  # USA vs Belgium
    "M95": ("M86胜者", "M88胜者"),  # Argentina vs Turkey
    "M96": ("M85胜者", "M87胜者"),  # Switzerland vs Portugal
}

r32_winners = {m[0]: m[3] for m in r32}

r16 = [
    ("M89", r32_winners["M74"], r32_winners["M77"], "France",
     "德国 vs 法国：巅峰对决！法国阵容年龄结构更优，姆巴佩正值巅峰，略胜一筹"),
    ("M90", r32_winners["M73"], r32_winners["M75"], "Netherlands",
     "韩国 vs 荷兰：荷兰整体实力碾压，韩国虽顽强但难敌橙色军团"),
    ("M91", r32_winners["M76"], r32_winners["M78"], "Brazil",
     "巴西 vs 挪威：哈兰德独木难支，巴西多点开花轻松晋级"),
    ("M92", r32_winners["M79"], r32_winners["M80"], "England",
     "墨西哥 vs 英格兰：英格兰人才济济，贝林厄姆+凯恩组合致命"),
    ("M93", r32_winners["M83"], r32_winners["M84"], "Spain",
     "克罗地亚 vs 西班牙：莫德里奇谢幕战，西班牙新生代已成熟"),
    ("M94", r32_winners["M81"], r32_winners["M82"], "Belgium",
     "美国 vs 比利时：比利时虽老化但经验丰富，德布劳内掌控中场"),
    ("M95", r32_winners["M86"], r32_winners["M88"], "Argentina",
     "阿根廷 vs 土耳其：阿根廷全队状态正佳，土耳其难以招架"),
    ("M96", r32_winners["M85"], r32_winners["M87"], "Portugal",
     "瑞士 vs 葡萄牙：葡萄牙攻击线豪华，C罗继续书写传奇"),
]

r16_winners = {m[0]: m[3] for m in r16}

# Step 4: Quarter-finals
qf = [
    ("M97", r16_winners["M89"], r16_winners["M90"], "France",
     "法国 vs 荷兰：法国攻防两端均占优，姆巴佩速度完爆荷兰防线"),
    ("M98", r16_winners["M93"], r16_winners["M94"], "Spain",
     "西班牙 vs 比利时：传控 vs 反击，西班牙控球率碾压，终场前绝杀"),
    ("M99", r16_winners["M91"], r16_winners["M92"], "Brazil",
     "巴西 vs 英格兰：世界杯经典对决！维尼修斯+罗德里戈双星闪耀，点球大战巴西胜出"),
    ("M100", r16_winners["M95"], r16_winners["M96"], "Argentina",
     "阿根廷 vs 葡萄牙：梅西 vs C罗终极一战！梅西策动全场，阿根廷 2-1 险胜"),
]

qf_winners = {m[0]: m[3] for m in qf}

# Step 5: Semi-finals
sf = [
    ("M101", qf_winners["M97"], qf_winners["M98"], "France",
     "法国 vs 西班牙：2024 欧洲杯半决赛重演！法国反击如刀，姆巴佩戴帽，3-1 晋级决赛"),
    ("M102", qf_winners["M99"], qf_winners["M100"], "Argentina",
     "巴西 vs 阿根廷：南美超级德比！梅西世界杯最后一舞，任意球绝杀，1-0 挺进决赛"),
]

sf_winners = {m[0]: m[3] for m in sf}

# Step 6: Final & Third place
third = ("M103", qf_winners["M99"], sf_winners["M102"], "Brazil",
         "巴西 vs 阿根廷（半决赛败者） → 巴西：南美双雄季军之争，巴西体能占优 2-0 获胜")

champion_match = ("M104", sf_winners["M101"], sf_winners["M102"], "France",
     "【冠军预测】法国 vs 阿根廷 —— 2022 世界杯决赛重演！\n"
     "姆巴佩 vs 梅西，新王 vs 旧王，速度 vs 智慧。\n"
     "法国队兵强马壮，三条线均衡：迈尼昂守门、萨利巴+科纳特中卫、楚阿梅尼中场、姆巴佩锋线，全员当打之年。\n"
     "阿根廷虽有梅西压阵，但核心阵容老化（梅西 39 岁、迪马利亚退役），高原透支后难敌法国青春风暴。\n"
     "最终：法国 3-2 阿根廷，姆巴佩梅开二度锁定胜局，法国时隔 8 年重夺大力神杯！")

# ============================================================
# Create "冠军预测" sheet
# ============================================================
ws = wb.create_sheet("冠军预测")

# ---- Section 1: Title ----
ws.merge_cells('A1:G1')
c = ws.cell(row=1, column=1, value="🏆 2026 世界杯冠军预测及完整推理")
c.font = Font(name="Microsoft YaHei", size=16, bold=True, color="1F4E79")
c.alignment = center

ws.merge_cells('A2:G2')
c = ws.cell(row=2, column=1, value="预测冠军：🇫🇷 法国 (France)")
c.font = Font(name="Microsoft YaHei", size=14, bold=True, color="C00000")
c.alignment = center

# ---- Section 2: Core reasoning summary ----
row = 4
ws.merge_cells(f'A{row}:G{row}')
c = ws.cell(row=row, column=1, value="一、核心推理逻辑")
c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="1F4E79")

reasons = [
    ("阵容深度", "法国拥有全球最深的足球人才池——姆巴佩、楚阿梅尼、卡马文加、萨利巴、科纳特、巴尔科拉、埃梅里，全队 26 人几乎人均五大联赛主力。没有明显短板位置。"),
    ("大赛经验", "法国近四届世界杯成绩：2014 八强、2018 冠军、2022 亚军、2026 ??。德尚执教已超 14 年，体系成熟稳定。连续两届进决赛绝非偶然。"),
    ("年龄结构", "法国核心阵容处于黄金年龄——姆巴佩 27 岁、楚阿梅尼 26 岁、萨利巴 25 岁，正值巅峰。对比阿根廷（梅西 39）、葡萄牙（C罗 41），年龄优势巨大。"),
    ("战术体系", "德尚的防守反击+快速转换打法非常适合杯赛。中场拦截能力世界顶级，前场姆巴佩+登贝莱的速度任何防线都难以招架。"),
    ("对手分析", "主要竞争对手各有短板：阿根廷阵容老化、巴西防守不稳、英格兰教练保守、西班牙锋无力、德国重建中。法国是短板最少的球队。"),
    ("历史规律", "世界杯历史上，只有意大利（1934/1938）和巴西（1958/1962）成功卫冕。但法国 2018 冠→2022 亚→2026 ? 呈现出强大的延续性，打破魔咒正当其时。"),
]

for i, (title, detail) in enumerate(reasons):
    row += 1
    c1 = ws.cell(row=row, column=1, value=title)
    c1.font = Font(name="Microsoft YaHei", size=10, bold=True)
    c1.alignment = center
    c1.border = thin_border
    ws.merge_cells(f'B{row}:G{row}')
    c2 = ws.cell(row=row, column=2, value=detail)
    c2.font = cell_font
    c2.alignment = left_align
    c2.border = thin_border

# ---- Section 3: Full bracket simulation ----
row += 2
ws.merge_cells(f'A{row}:G{row}')
c = ws.cell(row=row, column=1, value="二、淘汰赛完整模拟推演")
c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="1F4E79")

# Helper to write a stage
def write_stage(ws, start_row, title, data):
    row = start_row + 1
    ws.merge_cells(f'A{row}:G{row}')
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Microsoft YaHei", size=11, bold=True, color="2E75B6")

    headers = ["场次","主队","客队","胜方","推理依据"]
    widths = [8, 20, 20, 20, 70]
    row += 1
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for d in data:
        row += 1
        match_id, team_a, team_b, winner, reason = d
        vals = [match_id, team_a, team_b, winner, reason]
        is_winner = "法国" in winner or "France" in winner
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = bold_font if (is_winner and col == 4) else cell_font
            cell.alignment = center if col <= 4 else left_align
            cell.border = thin_border
            if is_winner and col == 4:
                cell.fill = champion_fill
    return row + 2

row = write_stage(ws, row, "▸ 32 强赛 (Round of 32)", r32)
row = write_stage(ws, row, "▸ 16 强赛 (Round of 16)", r16)
row = write_stage(ws, row, "▸ 1/4 决赛 (Quarter-finals)", qf)
row = write_stage(ws, row, "▸ 半决赛 (Semi-finals)", sf)

# Third place
row -= 1
ws.merge_cells(f'A{row}:G{row}')
c = ws.cell(row=row, column=1, value="▸ 季军赛 (Third Place)")
c.font = Font(name="Microsoft YaHei", size=11, bold=True, color="2E75B6")
row += 1
for col, (h, w) in enumerate(zip(["场次","主队","客队","胜方","推理依据"], [8,20,20,20,70]), 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
row += 1
for col, val in enumerate([third[0], third[1], third[2], third[3], third[4]], 1):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = cell_font
    cell.alignment = center if col <= 4 else left_align
    cell.border = thin_border

# Final
row += 1
ws.merge_cells(f'A{row}:G{row}')
c = ws.cell(row=row, column=1, value="▸ 🏆 决  赛 (FINAL)")
c.font = Font(name="Microsoft YaHei", size=11, bold=True, color="C00000")
row += 1
for col, (h, w) in enumerate(zip(["场次","主队","客队","🏆冠军","推理依据"], [8,20,20,20,70]), 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = center
    cell.border = thin_border
row += 1
cm = champion_match
for col, val in enumerate([cm[0], cm[1], cm[2], cm[3], cm[4]], 1):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = bold_font
    cell.fill = champion_fill
    cell.alignment = center if col <= 4 else left_align
    cell.border = thin_border

# ---- Section 4: Alternative scenarios ----
row += 2
ws.merge_cells(f'A{row}:G{row}')
c = ws.cell(row=row, column=1, value="三、夺冠概率 Top 5")
c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="1F4E79")

row += 1
prob_headers = ["排名","球队","夺冠概率","核心优势","最大隐患"]
for col, h in enumerate(prob_headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

probs = [
    ("🥇","法国","35%","阵容厚度冠绝全球 + 姆巴佩巅峰 + 德尚体系12年不变","心理压力：连续三届决赛期望值过高"),
    ("🥈","阿根廷","18%","梅西最后一舞的精神力量 + 世界杯卫冕冠军底蕴","核心老化：梅西39岁、后防重组未完成"),
    ("🥉","巴西","15%","攻击线天赋溢出（维尼修斯/罗德里戈/恩德里克）+ 20年无冠饥渴感","中场缺乏组织核心，后防不稳定"),
    ("4","英格兰","12%","黄金一代成熟期（贝林厄姆/福登/萨卡）+ 索斯盖特下课新帅红利","大赛心理脆弱，点球大战阴影"),
    ("5","西班牙","10%","传控体系 + 新生代崛起（亚马尔/佩德里/加维）","锋线终结能力不足，缺少顶级射手"),
]

for rank, team, prob, strength, weakness in probs:
    row += 1
    vals = [rank, team, prob, strength, weakness]
    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = cell_font
        cell.alignment = center if col <= 3 else left_align
        cell.border = thin_border

# ---- Section 5: Disclaimer ----
row += 2
ws.merge_cells(f'A{row}:G{row}')
c = ws.cell(row=row, column=1, value="⚠️ 免责声明：以上预测基于 2026 年 5 月已知信息和主观判断，纯属娱乐。足球比赛结果受伤病、红牌、天气、裁判、运气等无数因素影响，请勿当真或用于投注。")
c.font = Font(name="Microsoft YaHei", size=9, italic=True, color="808080")
c.alignment = left_align

# Column widths for this sheet
widths = [8, 20, 20, 20, 70, 0, 0]
for col, w in enumerate(widths, 1):
    if w > 0:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

ws.freeze_panes = "A4"

# Save
wb.save(path)
print(f"✅ 冠军预测已添加至: {path}")
print(f"🏆 预测冠军: 法国 (France)")
print(f"📊 夺冠概率: 35%")
print(f"🥈 亚军: 阿根廷")
print(f"🥉 季军: 巴西")
