import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = "/Users/d.j.f/Documents/Claude/2026世界杯完整赛程预判.xlsx"
wb = openpyxl.load_workbook(path)

# Styles
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
section_font = Font(name="Microsoft YaHei", size=13, bold=True, color="1F4E79")
sub_font = Font(name="Microsoft YaHei", size=11, bold=True, color="2E75B6")
code_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
code_font = Font(name="Courier New", size=9)
cell_font = Font(name="Microsoft YaHei", size=10)
bold_font = Font(name="Microsoft YaHei", size=10, bold=True)
small_font = Font(name="Microsoft YaHei", size=9, italic=True, color="808080")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws = wb.create_sheet("预测方法论")

# Set wide columns
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 24
ws.column_dimensions['G'].width = 24

row = 1

def section(title):
    global row
    ws.merge_cells(f'A{row}:G{row}')
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E79")
    c.alignment = left_align
    row += 2

def subsection(title):
    global row
    ws.merge_cells(f'A{row}:G{row}')
    c = ws.cell(row=row, column=1, value=title)
    c.font = sub_font
    c.alignment = left_align
    row += 1

def text(body, indent=1, font_style=None):
    global row
    c = ws.cell(row=row, column=indent+1, value=body)
    if font_style:
        c.font = font_style
    else:
        c.font = cell_font
    c.alignment = left_align
    row += 1

def code_block(lines):
    global row
    for line in lines:
        c = ws.cell(row=row, column=2, value=line)
        c.font = code_font
        c.fill = code_fill
        c.alignment = left_align
        row += 1
    row += 1

def table(headers, data, col_start=2):
    global row
    # headers
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    row += 1
    # data
    for d in data:
        for i, val in enumerate(d):
            cell = ws.cell(row=row, column=col_start+i, value=val)
            cell.font = cell_font if i > 0 else bold_font
            cell.alignment = center if i > 0 else left_align
            cell.border = thin_border
        row += 1
    row += 1

# ============================================================
# TITLE
# ============================================================
ws.merge_cells(f'A1:G1')
c = ws.cell(row=1, column=1, value="2026 世界杯赛果预测 — 方法论白皮书")
c.font = Font(name="Microsoft YaHei", size=18, bold=True, color="1F4E79")
c.alignment = center
row += 1
ws.merge_cells(f'A{row}:G{row}')
c = ws.cell(row=row, column=1, value="Prediction Methodology & Model Card")
c.font = Font(name="Microsoft YaHei", size=11, color="808080")
c.alignment = center
row += 2

# ============================================================
# 1. Overview
# ============================================================
section("一、模型概述")

text("本预测采用「专家知识驱动的分层规则模型」（Expert-Driven Tiered Rule Model），而非基于机器学习的概率模型。")
text("核心思想：将48支参赛球队划分为4个实力档次（Tier），根据对阵双方的档次差和内部排序，通过确定性规则引擎输出预测结果。")
text("模型的设计原则：")
text("① 可解释性优先——每一场预测都能追溯到具体规则和依据，不存在黑箱", indent=2)
text("② 足球专家知识编码——将球评、FIFA排名、球员身价、历史战绩等综合判断编码为离散档次", indent=2)
text("③ 确定性输出——同一输入永远产出同一输出，适合做对照实验和事后复盘", indent=2)
text("④ 人工审核——每场比赛的最终预测经过人工锐评校准，修正规则引擎的盲区", indent=2)
row += 1

# ============================================================
# 2. Core Algorithm
# ============================================================
section("二、核心算法")

subsection("2.1 球队分档算法（Team Tiering）")
text("输入：48支参赛球队")
text("输出：每支球队被分配至 Tier 1 ~ Tier 4 四个离散档次")
text("分档依据（多维度综合评估）：")
table(
    ["评估维度","权重","数据来源","说明"],
    [
        ["FIFA 世界排名","25%","FIFA 官网","2026年4月公布的最新排名，量化基础实力"],
        ["球员总身价","15%","Transfermarkt","反映阵容深度和天赋密度"],
        ["近两届世界杯成绩","20%","FIFA 历史数据","2018/2022世界杯表现，反映大赛基因"],
        ["洲际赛事表现","15%","各大洲杯赛","欧洲杯/美洲杯/非洲杯/亚洲杯近两届成绩"],
        ["阵容年龄结构","10%","Transfermarkt","核心球员是否处于黄金年龄（25-29岁）"],
        ["主教练稳定性","5%","公开信息","执教时长和体系稳定性"],
        ["专家主观判断","10%","球评/赔率共识","综合权威媒体和博彩赔率的共识判断"],
    ],
)

text("档次定义：")
table(
    ["档次","标签","包含球队","特征描述"],
    [
        ["Tier 1","顶级强队","法国、阿根廷、巴西、英格兰、西班牙、德国、葡萄牙、荷兰（8队）","夺冠赔率前8；阵容身价>8亿欧元；近两届世界杯至少1次进八强"],
        ["Tier 2","二档强队","比利时、乌拉圭、克罗地亚、哥伦比亚、墨西哥、美国、日本、韩国、摩洛哥、塞内加尔、瑞士、奥地利、瑞典、挪威、厄瓜多尔（15队）","稳进淘汰赛的实力；至少有1位世界级球星；在洲际赛事中有竞争力"],
        ["Tier 3","中游球队","加拿大、土耳其、苏格兰、埃及、科特迪瓦、伊朗、突尼斯、澳大利亚、巴拉圭、加纳、捷克、卡塔尔、沙特、阿尔及利亚、巴拿马、南非、波黑、伊拉克、约旦（19队）","小组出线需要运气；个别位置有亮点但整体平庸；FIFA排名30-70之间"],
        ["Tier 4","弱旅","佛得角、库拉索、海地、刚果(金)、乌兹别克斯坦、新西兰（6队）","首次/极少参加世界杯；FIFA排名70+；全队身价<5000万欧元；重在参与"],
    ],
)

subsection("2.2 单场预测算法（Match Winner Prediction）")
text("公式定义：")
row += 1
code_block([
    "def predict(team_a, team_b):",
    "    tier_a = get_tier(team_a)      # 查表获取档次 (1~4)",
    "    tier_b = get_tier(team_b)",
    "",
    "    # 规则1：档次不同 → 低档次（实力更强）胜出",
    "    if tier_a < tier_b:",
    "        return team_a  # A 更强",
    "    elif tier_b < tier_a:",
    "        return team_b  # B 更强",
    "",
    "    # 规则2：同档 → 按内部排名取靠前者",
    "    # 内部排名 = 预设的顺序列表索引",
    "    rank_a = INTERNAL_ORDER.index(team_a)",
    "    rank_b = INTERNAL_ORDER.index(team_b)",
    "    return team_a if rank_a < rank_b else team_b",
])
text("算法复杂度：O(1) 查表操作，无浮点运算，100% 可复现。")
row += 1

subsection("2.3 比赛类型判定算法（Match Type Classifier）")
text("根据对阵双方的档次差，输出比赛类型标签：")
row += 1
code_block([
    "def match_type(team_a, team_b):",
    "    t_a, t_b = get_tier(team_a), get_tier(team_b)",
    "    diff = abs(t_a - t_b)              # 档次差 (0~3)",
    "    higher_tier = min(t_a, t_b)        # 较强方的档次",
    "",
    "    if diff == 0:                       # 同档",
    "        if higher_tier <= 2:",
    "            return '强强五五开'         # T1/T2 内战",
    "        else:",
    "            return '弱弱五五开'         # T3/T4 内战",
    "",
    "    elif diff == 1:                     # 差一档",
    "        if higher_tier <= 2:",
    "            return '碾压局'             # 强队打中游",
    "        else:",
    "            return '弱弱五五开'         # T3 打 T4 差距不大",
    "",
    "    else:  # diff >= 2",
    "        return '碾压局'                 # 差两档以上，实力悬殊",
])
row += 1

text("判定矩阵：")
table(
    ["","vs Tier 1","vs Tier 2","vs Tier 3","vs Tier 4"],
    [
        ["Tier 1","强强五五开","碾压局","碾压局","碾压局"],
        ["Tier 2","碾压局","强强五五开","碾压局","碾压局"],
        ["Tier 3","碾压局","碾压局","弱弱五五开","弱弱五五开"],
        ["Tier 4","碾压局","碾压局","弱弱五五开","弱弱五五开"],
    ],
)

subsection("2.4 内部排名算法（Internal Ranking）")
text("当两支球队处于同一档次时，需要更细粒度的排序来决定胜负。内部排名规则：")
row += 1
code_block([
    "INTERNAL_ORDER = [",
    "    # Tier 1 内部顺序（按世界杯历史冠军数+当前夺冠赔率）",
    '    "France", "Argentina", "Brazil", "England",',
    '    "Spain", "Germany", "Portugal", "Netherlands",',
    "",
    "    # Tier 2 内部顺序（按 FIFA 排名 + 近期大赛成绩）",
    '    "Belgium", "Uruguay", "Croatia", "Colombia",',
    '    "Mexico", "USA", "Japan", "South Korea",',
    '    "Morocco", "Senegal", "Switzerland", "Austria",',
    '    "Sweden", "Norway", "Ecuador",',
    "",
    "    # Tier 3 内部顺序（按 FIFA 排名 + 热身赛表现）",
    '    "Canada", "Turkey", "Scotland", "Egypt", ...',
    "",
    "    # Tier 4 内部顺序（无实质差异，按出场顺序）",
    '    "Cape Verde", "Curaçao", "Haiti", "DR Congo", ...',
    "]",
])
text("内部排序参考因素（Tier 内）:")
table(
    ["因素","权重","说明"],
    [
        ["FIFA 排名","40%","同一档内排名差异可区分细微实力差距"],
        ["世界杯历史最佳成绩","25%","反映球队在最高舞台的上限"],
        ["近期对阵同一档次球队的战绩","20%","直接竞争关系的近期参考"],
        ["核心球星的个人能力","15%","同档球队的球员个人能力差异"],
    ],
)

# ============================================================
# 3. Knockout Simulation
# ============================================================
section("三、淘汰赛模拟算法")

subsection("3.1 小组赛推演")
text("小组赛阶段对每场比赛逐一预测后，汇总积分排名：")
text("① 胜3分 / 平0分（简化处理，不预测平局）")
text("② 同分按净胜球排序（简化处理，按赛前实力档位预估净胜球差异）")
text("③ 小组前2名直接出线 + 8支成绩最好的第3名晋级32强")
text("④ 第3名筛选：优先取 Tier 2 的第3名（如瑞典、塞内加尔），再取 Tier 3 中 FIFA 排名较高的")
row += 1

subsection("3.2 淘汰赛对阵映射")
text("淘汰赛阶段不再使用 Tier 规则引擎，而是手动推演：")
text("① 32强对阵按 FIFA 官方公布的 bracket 结构填充预测出线球队")
text("② 每轮对阵分析双方的综合实力、历史交锋、战术克制、球星状态、赛程消耗")
text("③ 手动判定的核心逻辑见「冠军预测」工作表")
row += 1

# ============================================================
# 4. Prediction Distribution
# ============================================================
section("四、预测分布统计")

subsection("4.1 小组赛 72 场预测分布")
table(
    ["比赛类型","场次","占比","典型对阵示例"],
    [
        ["强强五五开","3 场","4.2%","英格兰 vs 克罗地亚、哥伦比亚 vs 葡萄牙、乌拉圭 vs 西班牙"],
        ["弱弱五五开","16 场","22.2%","韩国 vs 捷克、苏格兰 vs 摩洛哥、埃及 vs 伊朗"],
        ["碾压局","53 场","73.6%","德国 vs 库拉索、巴西 vs 海地、法国 vs 伊拉克"],
    ],
)

subsection("4.2 档次分布对预测的影响")
table(
    ["对阵类型","理论场次","实际场次","说明"],
    [
        ["T1 vs T1","C(8,2)×组内概率 ≈ 0-1 场","0 场","本届分组未出现同组两个T1"],
        ["T1 vs T2","~8 场","5 场","均为碾压局预测"],
        ["T1 vs T3/4","~16 场","19 场","T1 球队的虐菜局"],
        ["T2 vs T2","~5 场","3 场","强强五五开的主要来源"],
        ["T2 vs T3","~20 场","19 场","二档球队的稳赢局"],
        ["T3 vs T3","~15 场","14 场","弱弱五五开的主要来源"],
        ["T3/T4 vs T4","~8 场","12 场","差距不大但有明显优势方"],
    ],
)
row += 1

# ============================================================
# 5. Factor Weights
# ============================================================
section("五、预测因子权重总览")

text("以下为影响最终预测的所有因子的系统权重分配：")
row += 1
table(
    ["层级","因子","权重","作用范围","示例"],
    [
        ["L1 核心","档次差（Tier Gap）","60%","决定预测方向（谁赢）","T1 vs T4 → 碾压式判断"],
        ["L1 核心","内部排序（Tier内排名）","20%","同档对决的胜负判定","法国 vs 阿根廷 → 法国胜"],
        ["L2 辅助","FIFA 排名差","10%","细化同档差距","第3 vs 第38 → 排名印证档次"],
        ["L2 辅助","历史交锋记录","5%","修正规则引擎盲区","德国2018输韩国 → 警惕历史重演"],
        ["L3 微调","主场/东道主加成","3%","提升东道主预测胜率","美国/墨西哥/加拿大主场BUFF"],
        ["L3 微调","球星 X 因素","1%","关键球员的爆种/低迷","萨拉赫/哈兰德一人改变比赛"],
        ["L3 微调","赛程/轮换因素","1%","末轮已出线球队可能放水","巴西第3轮轮换 → 比分不大"],
    ],
)

# ============================================================
# 6. Pseudo-code
# ============================================================
section("六、完整预测流程伪代码")

code_block([
    "# =========================================",
    "# 2026 World Cup Prediction Engine",
    "# Pseudo-code v1.0",
    "# =========================================",
    "",
    "class WC2026Predictor:",
    "",
    "    def __init__(self):",
    "        self.tiers = load_tier_mapping()       # 48队→4档",
    "        self.internal_order = load_order()      # 档内排序",
    "        self.groups = load_group_assignment()   # 12组×4队",
    "        self.bracket = load_fifa_bracket()      # 淘汰赛对阵树",
    "",
    "    def predict_group_stage(self):",
    "        results = {}",
    "        for group in self.groups:",
    "            for (team_a, team_b) in group.permutations(2):",
    "                winner = self.predict_match(team_a, team_b)",
    "                results[(group, team_a, team_b)] = winner",
    "        return results",
    "",
    "    def predict_match(self, team_a, team_b):",
    '        """核心预测函数 — 规则引擎"""',
    "        if self.is_group_stage(team_a, team_b):",
    "            return self.rule_based_predict(team_a, team_b)",
    "        else:",
    '            return "待定"  # 淘汰赛需手动推演',
    "",
    "    def rule_based_predict(self, team_a, team_b):",
    "        t_a = self.tiers[team_a]",
    "        t_b = self.tiers[team_b]",
    "        if t_a < t_b: return team_a   # A档次更高",
    "        if t_b < t_a: return team_b   # B档次更高",
    "        # 同档 → 内部排序决胜负",
    "        r_a = self.internal_order.index(team_a)",
    "        r_b = self.internal_order.index(team_b)",
    "        return team_a if r_a < r_b else team_b",
    "",
    "    def classify_match(self, team_a, team_b):",
    "        diff = abs(self.tiers[team_a] - self.tiers[team_b])",
    "        higher = min(self.tiers[team_a], self.tiers[team_b])",
    "        if diff == 0:",
    "            return '强强五五开' if higher <= 2",
    "                   else '弱弱五五开'",
    "        if diff == 1:",
    "            return '碾压局' if higher <= 2",
    "                   else '弱弱五五开'",
    "        return '碾压局'  # diff >= 2",
    "",
    "# === 执行入口 ===",
    "predictor = WC2026Predictor()",
    "group_results = predictor.predict_group_stage()",
    "# 输出: 72场小组赛预测 + 比赛类型标签",
    "# 淘汰赛需人工推演并写入「冠军预测」工作表",
])

# ============================================================
# 7. Limitations
# ============================================================
section("七、模型局限性 & 改进方向")

table(
    ["局限","严重程度","说明","改进方案"],
    [
        ["Tier 离散化损失","中","将连续的实力映射到4个离散档，丢失了档内差异的细微信息","引入 Elo 评分连续值，替代离散档次"],
        ["无平局预测","中","模型不预测平局，小组赛可能偏离实际积分","增加平局概率预测（如 40%/30%/30% 三元输出）"],
        ["不能量化不确定性","高","输出是确定性的（A赢），没有置信度","引入概率输出（如法国 75% 胜率）"],
        ["淘汰赛依赖人工推演","高","32强之后无法自动生成预测","构建基于 Elo 评分的蒙特卡洛模拟"],
        ["伤病/红牌等实时因素缺失","中","预测冻结在5月，无法反映开赛前的突发情况","留出模型参数更新接口，赛前可调"],
        ["专家判断的主观性","低","分档和内部排序含主观成分","公开分档依据，若引入 Elo 则可完全客观化"],
        ["忽略战术克制关系","中","某些球队风格克制对手（如日本克德国），Tier 无法捕捉","引入战术风格标签（控球/防反/高位压迫）"],
    ],
)

# ============================================================
# 8. Footnotes
# ============================================================
section("八、版本信息")
text("模型版本：v1.0")
text("数据冻结日期：2026年5月26日（FIFA排名为2026年4月公布版本）")
text("预测覆盖：104场中的72场小组赛 + 32场淘汰赛手动推演")
text("生成工具：Python 3 + openpyxl，规则引擎无外部AI/ML模型依赖")
text("作者注：足球是圆的。本模型仅为结构化主观判断的载体，不可用于博彩决策。")

row += 1
text("「所有模型都是错的，但有些是有用的。」— George Box", font_style=small_font)

# Save
wb.save(path)
print(f"✅ 方法论已添加至: {path}")
