#!/usr/bin/env python3
"""
评测分析脚本 v3：
- 支持历史数据沉淀(history.json)
- 生成带左侧日期导航的HTML报告
- 折线图展示趋势
- 完整不一致数据展示
"""
import json
import re
import os
import html as html_module
import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, 'raw_data.json')
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'outputs')
OUTPUT_HTML = os.path.join(OUTPUT_DIR, 'evaluation_report.html')
HISTORY_FILE = os.path.join(OUTPUT_DIR, 'history.json')

FIELD_MAP = [
    ('query', 'query'),
    ('grade', 'grade'),
    ('course', 'course'),
    ('app_qianwen_text', '千问'),
    ('app_doubao_speed_text', '豆包'),
    ('app_doubao_deepthink_text', '豆包深思'),
    ('app_yuanbao_text', '元宝'),
    ('ds', 'DeepSeek'),
    ('gemini-3.1-flash-image-preview', 'Gemini3.1'),
]

MODEL_FIELDS = [
    ('app_qianwen_text', '千问'),
    ('app_doubao_speed_text', '豆包'),
    ('app_doubao_deepthink_text', '豆包深思'),
    ('app_yuanbao_text', '元宝'),
    ('ds', 'DeepSeek'),
    ('gemini-3.1-flash-image-preview', 'Gemini3.1'),
]

COMPARISON_PAIRS = [
    ('app_doubao_speed_text', '千问-豆包'),
    ('app_doubao_deepthink_text', '千问-豆包深思'),
    ('app_yuanbao_text', '千问-元宝'),
    ('ds', '千问-DeepSeek'),
    ('gemini-3.1-flash-image-preview', '千问-Gemini3.1'),
]

# === 加载数据 ===
with open(INPUT_FILE, 'r', encoding='utf-8') as f:
    data = json.load(f)
print(f"共加载 {len(data)} 条记录")

# 确定评测日期（从数据中提取）
eval_date = data[0].get('ds', '2026-05-25') if data else '2026-05-25'

# === 生成Excel（按日期独立文件） ===
OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, f'原始数据-{eval_date}.xlsx')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

os.makedirs(OUTPUT_DIR, exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = "评测数据"
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
headers = [col_name for _, col_name in FIELD_MAP]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
for row_idx, record in enumerate(data, 2):
    for col_idx, (json_key, _) in enumerate(FIELD_MAP, 1):
        value = record.get(json_key, '')
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 15
for col in ['D', 'E', 'F', 'G', 'H', 'I']:
    ws.column_dimensions[col].width = 40
ws.freeze_panes = 'A2'
wb.save(OUTPUT_EXCEL)
print(f"Excel已保存: {OUTPUT_EXCEL}")

# === 答案提取函数 ===
def strip_markdown(text):
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'\*([^*]+)\*', r'\1', text)
    return text

def strip_latex_formatting(text):
    """去除仅做格式/样式的LaTeX命令，保留数学内容本身"""
    if not isinstance(text, str):
        return text
    # 辅助函数：匹配平衡花括号内容(支持嵌套)
    def match_braces(s, start):
        """从start位置的{开始，匹配到对应的}，返回内容和结束位置"""
        if start >= len(s) or s[start] != '{':
            return None, start
        depth = 0
        i = start
        while i < len(s):
            if s[i] == '{':
                depth += 1
            elif s[i] == '}':
                depth -= 1
                if depth == 0:
                    return s[start+1:i], i+1
            i += 1
        return s[start+1:], len(s)
    
    # 去除粗体/斜体/字体样式命令: \boldsymbol{x} -> x, \mathbf{x} -> x, etc.
    # 使用平衡括号匹配支持嵌套如 \boldsymbol{\frac{1}{8}}
    style_cmds = [
        'boldsymbol', 'mathbf', 'mathit', 'mathrm', 'mathsf', 'mathtt',
        'textbf', 'textit', 'textrm', 'textsf', 'texttt',
        'bf', 'it', 'rm', 'sf', 'tt',
        'bold', 'bm', 'pmb', 'boxed',
        'overline', 'underline', 'hat', 'bar', 'vec', 'tilde',
        'text', 'mbox', 'hbox',
    ]
    for cmd in style_cmds:
        pattern = '\\' + cmd + '{'
        while pattern in text:
            idx = text.find(pattern)
            brace_start = idx + len(pattern) - 1  # position of '{'
            content, end_pos = match_braces(text, brace_start)
            if content is not None:
                text = text[:idx] + content + text[end_pos:]
            else:
                break
    # \left, \right, \big, \Big etc. (delimiter sizing)
    text = re.sub(r'\\(?:left|right|big|Big|bigg|Bigg)\s*([|(){}[\].])?', r'\1', text)
    # \frac, \dfrac, \tfrac, \cfrac -> (a)/(b) 使用平衡括号匹配
    # v1.1: 添加 dfrac/tfrac/cfrac 支持（豆包/元宝常用\dfrac，千问用裸frac，旧版只处理\frac导致假分歧）
    for frac_pat in [r'\\dfrac\s*\{', r'\\tfrac\s*\{', r'\\cfrac\s*\{', r'\\frac\s*\{',
                      r'(?<![a-zA-Z])dfrac\s*\{', r'(?<![a-zA-Z])tfrac\s*\{',
                      r'(?<![a-zA-Z])cfrac\s*\{', r'(?<![a-zA-Z])frac\s*\{']:
        while True:
            m = re.search(frac_pat, text)
            if not m:
                break
            brace1_start = m.end() - 1  # position of first '{'
            num, pos1 = match_braces(text, brace1_start)
            if num is None:
                break
            # skip whitespace between } and next {
            pos2 = pos1
            while pos2 < len(text) and text[pos2] in ' \t':
                pos2 += 1
            if pos2 < len(text) and text[pos2] == '{':
                den, pos3 = match_braces(text, pos2)
                if den is not None:
                    text = text[:m.start()] + '(' + num + ')/(' + den + ')' + text[pos3:]
                    continue
            break
    # \sqrt{x} -> sqrt(x) 使用平衡括号匹配
    for sqrt_pat in [r'\\sqrt\s*\{', r'(?<![a-zA-Z])sqrt\s*\{']:
        while True:
            m = re.search(sqrt_pat, text)
            if not m:
                break
            brace_start = m.end() - 1
            content, end_pos = match_braces(text, brace_start)
            if content is not None:
                text = text[:m.start()] + 'sqrt(' + content + ')' + text[end_pos:]
            else:
                break
    # \times -> ×, \div -> ÷, \cdot -> ·, \cdots -> ...
    text = text.replace('\\times', '×').replace('\\div', '÷')
    text = text.replace('\\cdots', '...').replace('\\cdot', '·')
    text = text.replace('\\ldots', '...')
    # 裸命令(千问无反斜杠写法): times -> ×, quad -> 空格, ge -> ≥ 等
    text = re.sub(r'(?<![a-zA-Z])times(?![a-zA-Z])', '×', text)
    text = re.sub(r'(?<![a-zA-Z])div(?![a-zA-Z])', '÷', text)
    text = re.sub(r'(?<![a-zA-Z])cdot(?![a-zA-Z])', '·', text)
    text = re.sub(r'(?<![a-zA-Z])cdots(?![a-zA-Z])', '...', text)
    text = re.sub(r'(?<![a-zA-Z])quad(?![a-zA-Z])', ' ', text)
    text = re.sub(r'(?<![a-zA-Z])qquad(?![a-zA-Z])', ' ', text)
    text = re.sub(r'(?<![a-zA-Z])geq?(?![a-zA-Z])', '≥', text)
    text = re.sub(r'(?<![a-zA-Z])leq?(?![a-zA-Z])', '≤', text)
    text = re.sub(r'(?<![a-zA-Z])neq(?![a-zA-Z])', '≠', text)
    text = re.sub(r'(?<![a-zA-Z])infty(?![a-zA-Z])', '∞', text)
    text = re.sub(r'(?<![a-zA-Z])pm(?![a-zA-Z])', '±', text)
    # Remove remaining simple commands that are just formatting: \, \; \! \quad etc.
    text = re.sub(r'\\[,;!]', ' ', text)
    text = re.sub(r'\\(?:quad|qquad|hspace|vspace)\s*(?:\{[^}]*\})?', ' ', text)
    # \( \) \[ \] $ $$ delimiters -> remove them for plain comparison
    text = re.sub(r'\\\(|\\\)|\\\[|\\\]', '', text)
    text = re.sub(r'\$\$?', '', text)
    # Remove remaining backslash commands without braces (e.g. \alpha -> alpha)
    text = re.sub(r'\\([a-zA-Z]+)', r'\1', text)
    # Clean up braces that were part of LaTeX
    text = re.sub(r'[{}]', '', text)
    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extract_final_answer_line(text):
    """提取'最终答案'或'答案'标记后的完整答案内容（支持多行填空题）"""
    patterns = [
        r'最终答案[：:\s]*',
        r'(?:综合)?答案[：:\s]*\n',
        r'正确答案[是为]?[：:\s]*',
        r'结论[：:\s]*\n',
        r'总结[：:\s]*\n',
        r'解答[为是]?[：:\s]*\n',
    ]
    for pattern in patterns:
        # "解答"和"正确答案"标记特殊处理：使用最后一个出现位置
        is_last_match_marker = '解答' in pattern or '正确答案' in pattern
        if is_last_match_marker:
            matches = list(re.finditer(pattern, text))
            if not matches:
                continue
            match = matches[-1]  # 取最后一个
        else:
            match = re.search(pattern, text)
            if not match:
                continue
        after = text[match.end():]
        # 收集答案内容：可能是同一行的内容，也可能跨多行
        answer_lines = []
        empty_line_count = 0
        for line in after.split('\n'):
            line = line.strip()
            if not line:
                # 空行处理
                if not answer_lines:
                    # 前导空行跳过
                    continue
                empty_line_count += 1
                # 允许跳过最多1个空行继续收集（处理$$公式后空行的情况）
                if empty_line_count >= 2:
                    break
                continue
            # 有新内容，重置空行计数
            empty_line_count = 0
            # 如果遇到明显的非答案内容（新段落提示、追问等），停止
            if re.match(r'^(你要|要不要|需要|希望|如果|注意|提示|解[析释]|说明|方法)', line):
                break
            # 跳过纯markdown格式行(如 "---", "***", "* * *")
            if re.match(r'^[\-\*_\s]{3,}$', line):
                continue
            answer_lines.append(line)
            # 最多收集8行（从5提升到8，因为多小题+公式场景需要更多行）
            if len(answer_lines) >= 8:
                break
        
        if answer_lines:
            answer_text = '，'.join(answer_lines) if len(answer_lines) > 1 else answer_lines[0]
            # v1.9: 如果提取自"解答"标记且内容较长，尝试在文本中查找"答案"子标记
            # 豆包格式: "# 解答\n...(长解题过程)...\n答案: 鸡12只，免19只"
            # 此时应优先使用文末"答案"后的简洁内容
            if is_last_match_marker and len(answer_text) > 80 and '解答' in pattern:
                # 在原始text中查找"答案"子标记（文末简洁答案）
                sub_m = re.search(r'(?:最终)?答案[：:]\s*([^\n]{1,200})\s*$', text[-400:], re.MULTILINE)
                if sub_m:
                    sub_ans = sub_m.group(1).strip()
                    if sub_ans and len(sub_ans) >= 2 and len(sub_ans) < len(answer_text):
                        # 用文末简洁答案替代解题过程
                        return sub_ans
            # 过滤掉无效答案
            # 长度<=1且不含有效字符（汉字/字母/数字）的才过滤
            if len(answer_text) <= 1 and not re.search(r'[\u4e00-\u9fff\w]', answer_text):
                continue
            if re.match(r'^(填空|应为|汇总|如下|不唯一|解答|分析|参考|见下|见解析|答案见解析|详见解析|略)[：:）)\s]*$', answer_text):
                continue
            if answer_text.endswith('：') or answer_text.endswith(':'):
                continue
            if len(answer_text) > 300:
                continue
            # "解答"/"正确答案"标记特殊过滤：如果提取内容像是解题过程而非答案则跳过
            if is_last_match_marker:
                # 过滤解题过程描述（非简洁答案）
                if re.match(r'^(这道题|这个问题|本题|我们|首先|根据|为了|题目|由题|由已知|设|利用|从|按照)', answer_text):
                    continue
                # 过滤包含步骤编号的过程描述
                if re.match(r'^(第[一二三四五六七八九十\d]+[步个]|[\d]+[.、]\s*(?:分析|确定|计算|观察|根据|设|已知|实际))', answer_text):
                    continue
                # 过滤过长且含多步骤特征的内容(解题过程)
                if len(answer_text) > 100 and re.search(r'第[一二三]步|[123][.、]', answer_text):
                    continue
            return answer_text
    return None
    return None

def normalize_fill_in_answer(text):
    """标准化填空题答案：去除格式，提取核心数值和文本"""
    text = strip_latex_formatting(text)
    text = strip_markdown(text)
    # 去除emoji和特殊符号
    text = re.sub(r'[✅✓✔☑️✗✘❌⭕️]', '', text)
    # 去除括号包裹的数字，保留数字本身: (3) -> 3, （3） -> 3
    text = re.sub(r'[（(]\s*(\d+(?:\.\d+)?(?:/\d+)?)\s*[)）]', r'\1', text)
    # 标准化空格和标点
    text = re.sub(r'\s+', ' ', text).strip()
    text = re.sub(r'[，,]', ',', text)
    text = re.sub(r'[。.]$', '', text)
    return text

def extract_numbers_with_context(text):
    """从文本中提取所有数字及其紧邻的单位/上下文"""
    text = strip_latex_formatting(text)
    text = strip_markdown(text)
    # 去除emoji
    text = re.sub(r'[✅✓✔☑️✗✘❌⭕️]', '', text)
    # 提取: 数字 + 可选单位
    units = r'(?:立方米|平方米|米|千米|厘米|毫米|吨|千克|克|元|角|分|个|只|条|台|辆|本|支|块|片|张|把|棵|朵|座|间|艘|架|人|天|时|小时|分钟|秒|度|℃|%|倍|种|道|步|层|页|岁|周|月|年|亿|万|升|毫升|公斤|袋|颗|瓶|箱|包|盒|双|对|组|排|行|列|题|分)'
    # 匹配数字(可含小数/分数) + 可选单位
    pattern = r'(\d+(?:\.\d+)?(?:/\d+)?)\s*(' + units + r')?'
    results = re.findall(pattern, text)
    return [(num, unit) for num, unit in results if num]

def extract_answer(text):
    if not isinstance(text, str) or len(text.strip()) < 5 or text.strip() == 'none':
        return None, 'empty', None
    clean = strip_markdown(text)
    # 先做LaTeX格式清洗用于选择题/判断题
    clean_latex = strip_latex_formatting(clean)

    choice = extract_choice(clean_latex)
    if choice:
        return choice, 'choice', choice
    judgment = extract_judgment(clean_latex)
    if judgment:
        return judgment, 'judgment', judgment

    # 尝试提取"最终答案"整行作为填空题答案
    final_line = extract_final_answer_line(clean)
    if final_line:
        normalized = normalize_fill_in_answer(final_line)
        if normalized and len(normalized) > 1:
            return normalized, 'fill_in', final_line

    # 回退到单数值提取
    numerical = extract_numerical(clean_latex)
    if numerical:
        return numerical, 'numerical', numerical
    return None, 'complex', None

def extract_choice(text):
    patterns = [
        r'(?:最终)?答案[：:\s]*\n*\s*\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**',
        # 答案带括号: "最终答案：(D)" / "答案：（C）"
        r'(?:最终)?答案[：:\s]*\n*\s*[\(（]([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)[\)）]',
        r'参考答案[：:\s]*\n*\s*\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**',
        r'正确[答选]案?[是为：:\s]+[\(（]?\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**[\)）]?',
        r'正确选项[是为：:\s]+[\(（]?\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**[\)）]?',
        r'答案[是为]\s*[\(（]?\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**[\)）]?',
        r'对应选项[是为：:\s]+[\(（]?\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**[\)）]?',
        r'应选\s*\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**',
        r'[，,。]\s*选\s*\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**\s*[。✅\n]',
        r'(?:因此|所以|综上)[，,]?\s*(?:选项)?\s*([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)(?:是正确的|正确|为正确|均正确)',
        r'(?:因此|所以|综上)[，,]?\s*(?:正确|本题)?(?:答案|选项)?(?:选|是|为)[：:\s]*[\(（]?([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)[\)）]?',
        r'(?:建议|应该?)选[择取]?\s*\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**',
        r'选项\s*([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)(?:是正确|正确|都正确)',
        r'(?:是|为)\s*选项\s*[\(（]?\**([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\**[\)）]?',
        # "答案选X" 不要求前置逗号 (豆包新格式: "答案选A. √✅")
        r'答案选\s*([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)',
        # "这道题选X" / "本题选X"
        r'(?:这道题|本题)选\s*[\(（]?([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)[\)）]?',
        # 对应选项(X)✅ — 括号包裹
        r'对应选项\s*[\(（]([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)[\)）]',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            letters = sorted(set(re.findall(r'[A-E]', match.group(1))))
            if letters:
                return ','.join(letters)
    tail = text[-400:]
    # "选X✅" / "选X。" 尾部
    match = re.search(r'选([A-E])\s*[✅。]', tail)
    if match:
        return match.group(1)
    # "XX的是X✅" 豆包新格式: "说法不正确的是C✅" / "错误的说法是A✅" / "错误的描述是B✅"
    match = re.search(r'(?:正确|错误|不正确|不对)的?(?:选项|说法|答案|描述|表述|观点|叙述)?是\s*([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\s*[✅。\n]', tail)
    if match:
        letters = sorted(set(re.findall(r'[A-E]', match.group(1))))
        if letters:
            return ','.join(letters)
    # "所以XX的是X✅" 更宽松: "所以错误的是C✅" / "所以，不正确的说法是C✅"
    match = re.search(r'(?:因此|所以|综上)[，,]?\s*[^。\n]{0,15}的?(?:选项|说法|答案|描述|表述|观点|叙述)?是\s*([A-E][、,，\s]*(?:[A-E][、,，\s]*)*)\s*[✅。\n]', tail)
    if match:
        letters = sorted(set(re.findall(r'[A-E]', match.group(1))))
        if letters:
            return ','.join(letters)
    lines = tail.strip().split('\n')
    for line in reversed(lines[-5:]):
        line = line.strip()
        if re.match(r'^[A-E][、,，\s]*(?:[A-E][、,，\s]*)*$', line):
            letters = sorted(set(re.findall(r'[A-E]', line)))
            if letters and len(letters) <= 5:
                return ','.join(letters)
    match = re.search(r'最符合[^。]*的是\s*(?:选项\s*)?([A-E])', tail)
    if match:
        return match.group(1)
    # 通用："是选项 X" / "为 (X)" 尾部兜底
    match = re.search(r'(?:是|为)\s*选项?\s*[\(（]?([A-E])[\)）]?\s*[。\n]', tail)
    if match:
        return match.group(1)
    num_patterns = [
        r'(?:答案|正确答案)[是为：:\s]+([①②③④⑤][、,，\s]*(?:[①②③④⑤][、,，\s]*)*)',
        r'(?:正确|错误)的是[：:\s]*([①②③④⑤][、,，\s]*(?:[①②③④⑤][、,，\s]*)*)',
    ]
    for pattern in num_patterns:
        match = re.search(pattern, text)
        if match:
            nums = sorted(set(re.findall(r'[①②③④⑤]', match.group(1))))
            if nums:
                return ','.join(nums)
    return None

def extract_judgment(text):
    tail = text[-400:]
    patterns = [
        (r'答案[：:\s]*\n*\s*A[.\s（(]*正确', '正确'),
        (r'答案[：:\s]*\n*\s*B[.\s（(]*错误', '错误'),
        (r'答案[：:\s]*\n*\s*[×✗✘xX]', '错误'),
        (r'答案[：:\s]*\n*\s*[√✓✔]', '正确'),
        (r'判断[为是：:\s]+(?:对|正确|√)', '正确'),
        (r'判断[为是：:\s]+(?:错|错误|×)', '错误'),
        (r'说法[是]?正确的', '正确'),
        (r'说法[是]?错误的', '错误'),
        (r'本题[为是]?(?:正确|对)', '正确'),
        (r'本题[为是]?(?:错误|错)', '错误'),
        (r'选A[.\s]*[✅（(正确)]', '正确'),
        (r'选B[.\s]*[✅（(错误)]', '错误'),
        (r'答案[是为]?\s*A[.\s]*正确', '正确'),
        (r'答案[是为]?\s*B[.\s]*错误', '错误'),
        (r'正确的?答案是\s*A[.\s]*正确', '正确'),
        (r'正确的?答案是\s*B[.\s]*错误', '错误'),
        (r'(?:因此|所以|综上)[，,]?[^。]*(?:题[目干]?)?(?:说法|表述|判断)?(?:是)?正确的?[。\s]*$', '正确'),
        (r'(?:因此|所以|综上)[，,]?[^。]*(?:题[目干]?)?(?:说法|表述|判断)?(?:是)?错误的?[。\s]*$', '错误'),
    ]
    for pattern, answer in patterns:
        if re.search(pattern, tail):
            return answer
    return None

def extract_numerical(text):
    tail = text[-600:]
    units = r'(?:立方米|平方米|米|千米|厘米|毫米|吨|千克|克|元|角|分|个|只|条|台|辆|本|支|块|片|张|把|棵|朵|座|间|艘|架|人|天|时|小时|分钟|秒|度|℃|%|倍|种|道|步|层|页|岁|周|月|年|亿|万|升|毫升|公斤)'
    match = re.search(r'答[：:]\s*[^。\n]*?(\d+(?:\.\d+)?)\s*' + units, tail)
    if match:
        return match.group(1)
    matches = re.findall(r'=\s*(\d+(?:\.\d+)?)\s*(?:[（(]|' + units + r'|\s|$|[。，])', tail)
    if matches:
        return matches[-1]
    return None

def normalize_answer(answer):
    if answer is None:
        return None
    answer = str(answer).strip()
    answer = re.sub(r'[、，,\s]+', ',', answer)
    answer = answer.replace('①', '1').replace('②', '2').replace('③', '3').replace('④', '4').replace('⑤', '5')
    answer = answer.replace('对', '正确').replace('错', '错误')
    return answer.strip()

def normalize_for_comparison(text):
    """深度标准化：用于填空题/复杂答案的语义比对"""
    text = strip_latex_formatting(text)
    text = strip_markdown(text)
    # 去除emoji
    text = re.sub(r'[✅✓✔☑️✗✘❌⭕️💡]', '', text)
    # 去除"（也可XXX）"等括号内解释性内容
    text = re.sub(r'[（(]\s*(?:也可|或者|即|亦可|或按|如|例如|答案不唯一|意思相近|合理即可)[^)）]*[)）]', '', text)
    # 去除括号包裹(保留内容)
    text = re.sub(r'[（(]\s*([^)）]*?)\s*[)）]', r'\1', text)
    # 标准化标点
    text = re.sub(r'[，,]+', ',', text)
    text = re.sub(r'[。.]+$', '', text)
    text = re.sub(r'[：:]', ':', text)
    # 去除不影响答案语义的虚词/功能词
    text = re.sub(r'其中', '', text)
    text = re.sub(r'用(?=[全抽整总])', '', text)
    text = re.sub(r'进行', '', text)
    text = re.sub(r'(?:的话|来说)', '', text)
    # 去除多余空格
    text = re.sub(r'\s+', '', text)
    return text.strip()

def chinese_number_to_int(text):
    """将简单中文数字转为整数。仅处理纯中文数字字符串。"""
    text = text.strip()
    if not text:
        return None
    # 只处理纯中文数字字符串（全部字符都是数字相关）
    valid_chars = set('零一二三四五六七八九十百千万亿两半')
    if not all(c in valid_chars for c in text):
        return None
    simple = {'零':0,'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'两':2}
    if text in simple:
        return simple[text]
    if text == '十':
        return 10
    if text == '百':
        return 100
    if text == '千':
        return 1000
    if text == '万':
        return 10000
    # 处理十/百/千/万组合
    try:
        result = 0
        current = 0
        for char in text:
            if char in simple:
                current = simple[char]
            elif char == '十':
                result += (current if current else 1) * 10
                current = 0
            elif char == '百':
                result += (current if current else 1) * 100
                current = 0
            elif char == '千':
                result += (current if current else 1) * 1000
                current = 0
            elif char == '万':
                result = (result + current) * 10000
                current = 0
            elif char == '亿':
                result = (result + current) * 100000000
                current = 0
            elif char == '半':
                return None  # 半不好处理，返回None
        result += current
        return result if result > 0 else None
    except:
        return None

def answers_match(ans1, ans2, type1, type2):
    if ans1 is None or ans2 is None:
        return None
    norm1 = normalize_answer(ans1)
    norm2 = normalize_answer(ans2)
    if not norm1 or not norm2:
        return None
    # 直接相等
    if norm1 == norm2:
        return True

    # 选择题：集合比较
    if type1 == 'choice' and type2 == 'choice':
        return set(norm1.split(',')) == set(norm2.split(','))

    # 判断题
    if type1 == 'judgment' and type2 == 'judgment':
        return norm1 == norm2

    # 数值比较(两边都是纯数值)
    if type1 == 'numerical' and type2 == 'numerical':
        try:
            return abs(float(norm1) - float(norm2)) < 0.01
        except:
            return norm1 == norm2

    # 填空题 vs 填空题: 深度标准化后比对
    if type1 == 'fill_in' and type2 == 'fill_in':
        deep1 = normalize_for_comparison(ans1)
        deep2 = normalize_for_comparison(ans2)
        if deep1 and deep2 and deep1 == deep2:
            return True
        # 两方都在表达判断题答案（正确/对/错误/错）时，归一化比对
        if deep1 and deep2:
            judg_map = {'正确': '正确', '对': '正确', '错误': '错误', '错': '错误'}
            j1 = None
            j2 = None
            for kw, nv in judg_map.items():
                # 使用更宽松的边界：关键词在开头/末尾/被标点隔开即可
                if re.search(r'(?:^|[^一-龥a-zA-Z])' + kw + r'(?:$|[^一-龥a-zA-Z])', deep1) or deep1.startswith(kw) or deep1.endswith(kw):
                    j1 = nv
                    break
            for kw, nv in judg_map.items():
                if re.search(r'(?:^|[^一-龥a-zA-Z])' + kw + r'(?:$|[^一-龥a-zA-Z])', deep2) or deep2.startswith(kw) or deep2.endswith(kw):
                    j2 = nv
                    break
            if j1 and j2:
                # 双方都是短句判断 (<=20字)
                if len(deep1) <= 20 and len(deep2) <= 20:
                    return j1 == j2
        # 检查一方是否包含另一方（多小题情况：A回答了(1)(2)，B只回答了(2)）
        if deep1 and deep2:
            if deep1 in deep2 or deep2 in deep1:
                return True
        # 去除"（或：XXX）"备选说明后再做子串匹配
        if deep1 and deep2:
            clean1 = re.sub(r'或[:：][^；,，]*', '', deep1)
            clean2 = re.sub(r'或[:：][^；,，]*', '', deep2)
            clean1 = re.sub(r'[,，；;、\s]+', '', clean1)
            clean2 = re.sub(r'[,，；;、\s]+', '', clean2)
            if clean1 and clean2 and (clean1 in clean2 or clean2 in clean1):
                return True
        # 分段匹配：按分隔符拆分，检查各段关键答案是否在对方中出现
        if deep1 and deep2:
            seps = r'[,，；;、\n]+'
            # 去除序号前缀(1./2./(一)等)便于跨格式匹配
            numbering_re = re.compile(r'^(?:\d+[\.\、]|[(（][一二三四五六七八九十\d]+[)）])\s*')
            parts1_raw = [p.strip() for p in re.split(seps, deep1) if len(p.strip()) >= 2]
            parts2_raw = [p.strip() for p in re.split(seps, deep2) if len(p.strip()) >= 2]
            parts1 = [numbering_re.sub('', p) for p in parts1_raw]
            parts2 = [numbering_re.sub('', p) for p in parts2_raw]
            parts1 = [p for p in parts1 if len(p) >= 2]
            parts2 = [p for p in parts2 if len(p) >= 2]
            # 过滤掉过长的解释性段落(>20字)，只保留短答案段
            short_parts1 = [p for p in parts1 if len(p) <= 20]
            short_parts2 = [p for p in parts2 if len(p) <= 20]
            # 辅助：part-to-part匹配(双向子串)
            def part_matches_any(p, other_parts):
                for op in other_parts:
                    if p in op or op in p:
                        return True
                return False
            if short_parts1 and short_parts2:
                # 选平均段长更短的一方作为"答案方"，在对方段落中查找
                avg1 = sum(len(p) for p in short_parts1) / len(short_parts1)
                avg2 = sum(len(p) for p in short_parts2) / len(short_parts2)
                if avg1 <= avg2:
                    shorter, other = short_parts1, short_parts2
                    longer_text = deep2
                else:
                    shorter, other = short_parts2, short_parts1
                    longer_text = deep1
                # 检查每个段在对方全文中或对方段落中是否匹配
                match_count = sum(1 for p in shorter if p in longer_text or part_matches_any(p, other))
                if len(shorter) >= 3 and match_count / len(shorter) >= 0.6:
                    return True
                if len(shorter) >= 2 and match_count / len(shorter) >= 0.7:
                    return True
                # 也试反方向
                if avg1 <= avg2:
                    shorter, other = short_parts2, short_parts1
                    longer_text = deep1
                else:
                    shorter, other = short_parts1, short_parts2
                    longer_text = deep2
                match_count = sum(1 for p in shorter if p in longer_text or part_matches_any(p, other))
                if len(shorter) >= 3 and match_count / len(shorter) >= 0.6:
                    return True
                if len(shorter) >= 2 and match_count / len(shorter) >= 0.7:
                    return True
            # 也尝试用原始parts（不限长度），适用于多关键词答案
            if parts1 and parts2:
                shorter_p = parts1 if len(parts1) <= len(parts2) else parts2
                other_p = parts2 if len(parts1) <= len(parts2) else parts1
                longer_text = deep2 if len(parts1) <= len(parts2) else deep1
                match_count = sum(1 for p in shorter_p if p in longer_text or part_matches_any(p, other_p))
                if len(shorter_p) >= 3 and match_count / len(shorter_p) >= 0.6:
                    return True
            # 提取"X填：Y"/"X：Y"格式中冒号后的值，再做匹配
            def extract_values_after_colon(parts):
                vals = []
                for p in parts:
                    m = re.search(r'[:：]\s*(.+)$', p)
                    if m:
                        vals.append(m.group(1).strip())
                    else:
                        vals.append(p)
                return vals
            vals1 = extract_values_after_colon(parts1)
            vals2 = extract_values_after_colon(parts2)
            # 短值匹配
            short_vals1 = [v for v in vals1 if 2 <= len(v) <= 20]
            short_vals2 = [v for v in vals2 if 2 <= len(v) <= 20]
            if short_vals1 and short_vals2:
                # 双向检查
                for sv, lt in [(short_vals1, deep2), (short_vals2, deep1)]:
                    mc = sum(1 for v in sv if v in lt)
                    if len(sv) >= 2 and mc / len(sv) >= 0.7:
                        return True
        # 顺序无关的多部分答案比对：按连接词拆分后作为集合比较
        if deep1 and deep2:
            # 按"和/与/以及/，/；"等连接词拆分为独立部分
            conjunctions = r'(?:和|与|以及|[,，；;、])'
            parts_a = [p.strip() for p in re.split(conjunctions, deep1) if len(p.strip()) >= 2]
            parts_b = [p.strip() for p in re.split(conjunctions, deep2) if len(p.strip()) >= 2]
            # 双方都有多部分且部分数相同时，尝试集合匹配
            if len(parts_a) >= 2 and len(parts_b) >= 2 and len(parts_a) == len(parts_b):
                # normalize每个部分后作为集合比较（忽略顺序）
                set_a = set(re.sub(r'\s+', '', p) for p in parts_a)
                set_b = set(re.sub(r'\s+', '', p) for p in parts_b)
                if set_a == set_b:
                    return True
                # 更宽松：提取每个部分的核心数字+名词对，作为集合比较
                def extract_num_noun_key(part):
                    """从一个短答案片段中提取(数字,名词)key"""
                    m = re.search(r'(\d+(?:\.\d+)?(?:/\d+)?)\s*(?:杯|个|只|条|袋|颗|瓶|箱|盆|朵|块|张|棵|本|支|台|辆|把|元|克|千克|米|厘米|毫升|升|斤|吨)?\s*([\u4e00-\u9fff]{1,6})', part)
                    if m:
                        return (m.group(1), m.group(2))
                    # 反向：名词+数字
                    m2 = re.search(r'([\u4e00-\u9fff]{1,6})\s*[：:]*\s*(\d+(?:\.\d+)?(?:/\d+)?)', part)
                    if m2:
                        return (m2.group(2), m2.group(1))
                    return None
                keys_a = set(k for k in (extract_num_noun_key(p) for p in parts_a) if k)
                keys_b = set(k for k in (extract_num_noun_key(p) for p in parts_b) if k)
                if keys_a and keys_b and len(keys_a) >= 2 and keys_a == keys_b:
                    return True
        # 提取所有数字+单位按序比较（及集合比较）
        nums1 = extract_numbers_with_context(ans1)
        nums2 = extract_numbers_with_context(ans2)
        if nums1 and nums2 and len(nums1) == len(nums2):
            values1 = [n[0] for n in nums1]
            values2 = [n[0] for n in nums2]
            if values1 == values2:
                return True
            # 顺序无关：相同数字集合（处理多部分答案顺序颠倒的情况）
            if set(values1) == set(values2) and len(values1) >= 2:
                return True
            try:
                floats1 = sorted([float(v) for v in values1])
                floats2 = sorted([float(v) for v in values2])
                if all(abs(a-b) < 0.01 for a, b in zip(floats1, floats2)):
                    return True
            except:
                pass
        # 数字子集匹配（一方的数字全部出现在另一方中）
        if nums1 and nums2 and len(nums1) != len(nums2):
            values1 = set(n[0] for n in nums1)
            values2 = set(n[0] for n in nums2)
            if values1.issubset(values2) or values2.issubset(values1):
                return True
            # 放宽：答案数字交集>=70%（较少方）
            smaller = values1 if len(values1) <= len(values2) else values2
            overlap = values1 & values2
            if len(smaller) >= 3 and len(overlap) / len(smaller) >= 0.7:
                return True
        # 英文单词集合匹配（词库选词类题目）
        if deep1 and deep2:
            words1 = set(w.lower() for w in re.findall(r'[a-zA-Z]{3,}', deep1))
            words2 = set(w.lower() for w in re.findall(r'[a-zA-Z]{3,}', deep2))
            if len(words1) >= 2 and len(words2) >= 2:
                overlap_w = words1 & words2
                smaller_w = min(len(words1), len(words2))
                if smaller_w >= 2 and len(overlap_w) / smaller_w >= 0.75:
                    return True
        return False

    # 数值 vs 填空题 (或反过来): 检查fill_in答案中是否包含该数值作为最终答案
    num_ans = norm1 if type1 == 'numerical' else (norm2 if type2 == 'numerical' else None)
    fill_ans = ans1 if type1 == 'fill_in' else (ans2 if type2 == 'fill_in' else None)
    if num_ans and fill_ans:
        # 从fill_in答案中提取所有数字
        fill_nums = extract_numbers_with_context(fill_ans)
        fill_values = [n[0] for n in fill_nums]
        # 如果fill_in只有一个数字且等于num_ans，则一致
        if len(fill_values) == 1:
            try:
                if abs(float(fill_values[0]) - float(num_ans)) < 0.01:
                    return True
            except:
                if fill_values[0] == num_ans:
                    return True
        # 如果num_ans在fill_in的数字列表中（且是最后一个/主要答案）
        if num_ans in fill_values:
            return True
        try:
            num_float = float(num_ans)
            if any(abs(float(v) - num_float) < 0.01 for v in fill_values):
                return True
        except:
            pass
        return False

    # 选择题 vs 判断题：A=正确, B=错误 的跨类型映射
    if (type1 == 'choice' and type2 == 'judgment') or (type1 == 'judgment' and type2 == 'choice'):
        choice_val = norm1 if type1 == 'choice' else norm2
        judg_val = norm1 if type1 == 'judgment' else norm2
        # 仅单选A或B时适用（判断题场景）
        if choice_val in ['A', 'B']:
            choice_to_judg = {'A': '正确', 'B': '错误'}
            mapped = choice_to_judg.get(choice_val)
            if mapped == judg_val:
                return True
            else:
                return False

    # 选择题 vs 填空题：检查fill_in中是否包含对应选项
    choice_ans = norm1 if type1 == 'choice' else (norm2 if type2 == 'choice' else None)
    other_ans = ans2 if type1 == 'choice' else (ans1 if type2 == 'choice' else None)
    other_type = type2 if type1 == 'choice' else (type1 if type2 == 'choice' else None)
    if choice_ans and other_ans and other_type == 'fill_in':
        choice_set = set(choice_ans.split(','))
        other_clean = strip_latex_formatting(other_ans)

        # v1.6: 先尝试中文标签格式（豆包"高速马达：A、B、C"），比\b更精确
        labeled = re.findall(r'[一-鿿]{1,10}[：:]\s*\**([A-E](?:[、,，\s;；]*[A-E])*)\**', other_ans)
        if labeled:
            all_labels_letters = set()
            for group in labeled:
                all_labels_letters.update(re.findall(r'[A-E]', group))
            if all_labels_letters:
                # 严格相等
                if choice_set == all_labels_letters:
                    return True
                # 子集匹配：竞品答案包含多部分，千问只选了其中一部分（如千问{A,B,C} vs 竞品{A,B,C,D}）
                if choice_set.issubset(all_labels_letters) and len(all_labels_letters) > len(choice_set):
                    return True

        # 在fill_in答案中找所有孤立选项字母
        found_letters = set(re.findall(r'\b([A-E])\b', other_clean))
        if found_letters:
            if choice_set == found_letters:
                return True
            # 子集匹配（同上）
            if choice_set.issubset(found_letters) and len(found_letters) > len(choice_set):
                return True

        # 识别"答案选C"/"选C"模式
        select_match = re.search(r'(?:答案|选)[是为]?\s*([A-E](?:[、,，\s]*[A-E])*)', other_clean)
        if select_match:
            letters = set(re.findall(r'[A-E]', select_match.group(1)))
            if choice_set == letters:
                return True

        return False

    # 填空题 vs 判断题：检查fill_in中是否包含判断关键词
    judg_ans = norm1 if type1 == 'judgment' else (norm2 if type2 == 'judgment' else None)
    fill_ans_j = ans1 if type1 == 'fill_in' else (ans2 if type2 == 'fill_in' else None)
    if judg_ans and fill_ans_j:
        fill_clean = fill_ans_j.strip()
        # 从fill_in中提取判断结论
        if judg_ans == '正确' or judg_ans == '对,正确':
            if re.search(r'(正确|对$|对[,，。]|√|是正确的|判断正确)', fill_clean):
                # 确保没有"错误"在后面
                if not re.search(r'(错误|不正确|×|是错误的)', fill_clean.split('正确')[-1] if '正确' in fill_clean else ''):
                    return True
        elif judg_ans == '错误' or judg_ans == '错,错误':
            if re.search(r'(错误|错$|错[,，。]|×|不正确|是错误的)', fill_clean):
                return True
        return False

    # 最后兜底
    deep1 = normalize_for_comparison(ans1)
    deep2 = normalize_for_comparison(ans2)
    if deep1 and deep2 and deep1 == deep2:
        return True
    # 中文数字 vs 阿拉伯数字兜底
    if deep1 and deep2:
        cn1 = chinese_number_to_int(deep1)
        cn2 = chinese_number_to_int(deep2)
        if cn1 is not None and cn2 is not None:
            return cn1 == cn2
        # 一方是阿拉伯数字字符串
        try:
            if cn1 is not None and float(deep2) == cn1:
                return True
            if cn2 is not None and float(deep1) == cn2:
                return True
        except:
            pass

    return False

# === 执行答案提取 ===
print("\n提取答案中...")
all_answers = []
for idx, record in enumerate(data):
    ra = {}
    for json_key, _ in MODEL_FIELDS:
        text = record.get(json_key, '')
        answer, ans_type, display_answer = extract_answer(text)
        ra[json_key] = (answer, ans_type, display_answer)
    all_answers.append(ra)

for json_key, col_name in MODEL_FIELDS:
    extracted = sum(1 for ra in all_answers if ra[json_key][0] is not None)
    print(f"  {col_name}: {extracted}/{len(data)} ({extracted/len(data)*100:.1f}%)")

# === "无法作答"检测函数 ===
def is_cannot_answer(text):
    """检测模型响应是否表示无法作答（不应计入可比对分母）"""
    if not isinstance(text, str) or len(text.strip()) < 10:
        return False
    # 明确的"无法作答"信号（在文本前半部分出现更有意义）
    cant_patterns = [
        r'关键信息缺失',
        r'无法(?:进行)?(?:解答|判断|确定|计算|分析)',
        r'(?:图片|题目).*(?:不清|不完整|缺失|看不到|无法识别)',
        r'(?:需要|请).*(?:补充|提供|拍摄).*(?:完整|清晰)',
        r'信息不足.*无法',
        r'缺少.*(?:关键|必要|具体).*(?:信息|数据|条件)',
        r'无法.*(?:看清|辨认|识别).*(?:图片|文字|内容)',
    ]
    # 仅检查文本前400字（后面可能有模型的建议/追问）
    check_text = text[:400]
    for pat in cant_patterns:
        if re.search(pat, check_text):
            return True
    return False

def is_different_question(text1, text2):
    """检测两个模型的解题文本是否在回答不同的题目。
    同一query图片URL可能被API分发给不同模型时返回了不同题目，
    此时跨模型比对无意义，应标记为'可能为不同题目'排除。

    策略：双方文本的前200字中都有足够中文内容（≥5个不同关键词），
    但关键词完全无重叠（0交集），同时双方文本长度都不短（≥200字），
    表明两个模型确实在回答完全不同的题目。
    """
    if not text1 or not text2:
        return False
    if len(text1) < 200 or len(text2) < 200:
        return False

    # 提取前300字符
    t1_head = strip_markdown(text1)[:300]
    t2_head = strip_markdown(text2)[:300]

    def extract_keywords(text):
        cleaned = re.sub(r'\\\\[a-zA-Z]+(?:\{[^}]*\})*', '', text)
        cleaned = re.sub(r'\\$[^\\$]*\\$', '', cleaned)
        cleaned = re.sub(r'\\d+', '', cleaned)
        words = set()
        segments = re.split(r'[，。；：、！？\\n,.;:!?\\s]+', cleaned)
        for seg in segments:
            chinese_chars = re.findall(r'[一-鿿]{2,4}', seg)
            words.update(chinese_chars)
        return words

    w1 = extract_keywords(t1_head)
    w2 = extract_keywords(t2_head)

    # 双方都需要≥5个不同中文关键词才具备判定条件
    if len(w1) < 5 or len(w2) < 5:
        return False

    # 零关键词交集 → 不同题目
    overlap = w1 & w2
    return len(overlap) == 0

# === 一致性比对 ===
print("\n一致性比对...")

comparison_results = {}
all_disagreements = {}  # 完整不一致数据

for comp_key, pair_name in COMPARISON_PAIRS:
    matched = 0
    total_comparable = 0
    disagreements = []
    skipped_cant_answer = 0
    uncomparables = []  # 无法判断一致性的记录

    for idx, record in enumerate(data):
        qw_answer, qw_type, qw_display = all_answers[idx]['app_qianwen_text']
        comp_answer, comp_type, comp_display = all_answers[idx][comp_key]

        if qw_answer is not None and comp_answer is not None:
            # 检查是否有一方明确表示"无法作答"——不应计入可比对分母
            qw_text = record.get('app_qianwen_text', '')
            comp_text = record.get(comp_key, '')
            if is_cannot_answer(qw_text) or is_cannot_answer(comp_text):
                skipped_cant_answer += 1
                reason = '千问无法作答' if is_cannot_answer(qw_text) else '对方模型无法作答'
                uncomparables.append({
                    'idx': idx,
                    'query': record.get('query', ''),
                    'grade': record.get('grade', ''),
                    'course': record.get('course', ''),
                    'reason': reason,
                    'qw_answer': qw_display if qw_display else str(qw_answer)[:80],
                    'comp_answer': comp_display if comp_display else str(comp_answer)[:80],
                })
                continue

            total_comparable += 1
            is_match = answers_match(qw_answer, comp_answer, qw_type, comp_type)
            if is_match:
                matched += 1
            else:
                # 截图URL映射
                image_key_map = {
                    'app_doubao_speed_text': 'app_doubao_speed_image_url',
                    'app_doubao_deepthink_text': 'app_doubao_deepthink_image_url',
                    'app_yuanbao_text': 'app_yuanbao_image_url',
                    'ds': None,
                    'gemini-3.1-flash-image-preview': None,  # Gemini无独立截图字段
                }
                comp_image_key = image_key_map.get(comp_key)
                disagreements.append({
                    'idx': idx,
                    'query': record.get('query', ''),
                    'grade': record.get('grade', ''),
                    'course': record.get('course', ''),
                    'qw_doc': record.get('app_qianwen_text', ''),
                    'comp_doc': record.get(comp_key, ''),
                    'qw_answer': qw_display if qw_display else qw_answer,
                    'comp_answer': comp_display if comp_display else comp_answer,
                    'qw_image_url': record.get('app_qianwen_image_url', ''),
                    'comp_image_url': record.get(comp_image_key, '') if comp_image_key else '',
                })

        else:
            # 至少一方答案无法提取（complex类型）或文本为空
            qw_text_raw = record.get('app_qianwen_text', '')
            comp_text_raw = record.get(comp_key, '')
            qw_empty = not qw_text_raw or str(qw_text_raw).strip().lower() in ('none', 'null', '')
            comp_empty = not comp_text_raw or str(comp_text_raw).strip().lower() in ('none', 'null', '')

            if qw_empty and comp_empty:
                reason = '双方模型均未返回解题结果'
            elif qw_empty:
                reason = '千问未返回解题结果'
            elif comp_empty:
                reason = '对方模型未返回解题结果'
            elif qw_answer is None and comp_answer is None:
                reason = '双方均无法提取答案'
            elif qw_answer is None:
                reason = '千问答案无法提取'
            else:
                reason = '对方模型答案无法提取'

            uncomparables.append({
                'idx': idx,
                'query': record.get('query', ''),
                'grade': record.get('grade', ''),
                'course': record.get('course', ''),
                'reason': reason,
                'qw_answer': (qw_display if qw_display else str(qw_answer)[:80]) if qw_answer else '(无内容)' if qw_empty else '(无法提取)',
                'comp_answer': (comp_display if comp_display else str(comp_answer)[:80]) if comp_answer else '(无内容)' if comp_empty else '(无法提取)',
            })

    rate = matched / total_comparable * 100 if total_comparable > 0 else 0
    comparison_results[pair_name] = {
        'total_comparable': total_comparable,
        'matched': matched,
        'rate': round(rate, 1),
        'uncomparable_count': len(uncomparables),
    }
    all_disagreements[pair_name] = disagreements
    all_disagreements[pair_name + '_uncomparable'] = uncomparables
    skip_info = f" (跳过{skipped_cant_answer}条无法作答, {len(uncomparables)}条无法判断)" if skipped_cant_answer or uncomparables else ""
    print(f"  {pair_name}: {matched}/{total_comparable} 一致率 {rate:.1f}%{skip_info}")

# === 按年级/学科统计 ===
all_courses = sorted(set(r.get('course', '未知') for r in data))
all_grades = sorted(set(r.get('grade', '未知') for r in data))

grade_stats = {}
course_stats = {}
for comp_key, pair_name in COMPARISON_PAIRS:
    grade_stats[pair_name] = {}
    course_stats[pair_name] = {}
    for grade in all_grades:
        m, c = 0, 0
        for idx, record in enumerate(data):
            if record.get('grade', '') != grade: continue
            qw_a, qw_t, _ = all_answers[idx]['app_qianwen_text']
            comp_a, comp_t, _ = all_answers[idx][comp_key]
            if qw_a is not None and comp_a is not None:
                c += 1
                if answers_match(qw_a, comp_a, qw_t, comp_t): m += 1
        grade_stats[pair_name][grade] = {'comparable': c, 'matched': m, 'rate': round(m/c*100, 1) if c > 0 else 0}
    for course in all_courses:
        m, c = 0, 0
        for idx, record in enumerate(data):
            if record.get('course', '') != course: continue
            qw_a, qw_t, _ = all_answers[idx]['app_qianwen_text']
            comp_a, comp_t, _ = all_answers[idx][comp_key]
            if qw_a is not None and comp_a is not None:
                c += 1
                if answers_match(qw_a, comp_a, qw_t, comp_t): m += 1
        course_stats[pair_name][course] = {'comparable': c, 'matched': m, 'rate': round(m/c*100, 1) if c > 0 else 0}

# === 保存历史数据 ===
print("\n保存历史数据...")

# === 生成评测明细Excel ===
print("\n生成评测明细Excel...")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

detail_wb = Workbook()
detail_ws = detail_wb.active
detail_ws.title = "评测明细"

# 表头
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ['序号', 'query', '学段', '学科', '千问解题doc', '千问答案抽取',
           '对比模型', '对比模型解题doc', '对比模型答案抽取', '比对结果']
detail_ws.append(headers)
for cell in detail_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# 填充数据：每条query × 每个对比模型一行
row_num = 1
for idx, record in enumerate(data):
    qw_answer, qw_type, qw_display = all_answers[idx]['app_qianwen_text']
    qw_doc = record.get('app_qianwen_text', '') or ''
    query = record.get('query', '')
    grade = record.get('grade', '')
    course = record.get('course', '')

    for comp_key, pair_name in COMPARISON_PAIRS:
        comp_answer, comp_type, comp_display = all_answers[idx][comp_key]
        comp_doc = record.get(comp_key, '') or ''
        comp_name = pair_name.replace('千问-', '')

        # 确定比对结果
        if qw_answer is None or comp_answer is None:
            result = '无法比对'
        else:
            is_match = answers_match(qw_answer, comp_answer, qw_type, comp_type)
            result = '一致' if is_match else '不一致'

        row_num += 1
        detail_ws.append([
            idx + 1,
            query,
            grade,
            course,
            qw_doc[:2000] if qw_doc else '',
            str(qw_answer) if qw_answer else '未提取',
            comp_name,
            comp_doc[:2000] if comp_doc else '',
            str(comp_answer) if comp_answer else '未提取',
            result,
        ])
        # 设置比对结果颜色
        result_cell = detail_ws.cell(row=row_num, column=10)
        if result == '一致':
            result_cell.font = Font(color="22C55E")
        elif result == '不一致':
            result_cell.font = Font(color="EF4444")
        else:
            result_cell.font = Font(color="94A3B8")

# 设置列宽
col_widths = [6, 40, 8, 8, 50, 25, 12, 50, 25, 10]
for i, w in enumerate(col_widths, 1):
    detail_ws.column_dimensions[chr(64+i) if i <= 26 else 'A'].width = w
# 用正确方式设置列宽
from openpyxl.utils import get_column_letter
for i, w in enumerate(col_widths, 1):
    detail_ws.column_dimensions[get_column_letter(i)].width = w

# 冻结首行
detail_ws.freeze_panes = 'A2'

DETAIL_FILE = os.path.join(OUTPUT_DIR, f'评测明细-{eval_date}.xlsx')
detail_wb.save(DETAIL_FILE)
print(f"评测明细Excel已保存: {DETAIL_FILE} ({row_num} 行)")

history = []
if os.path.exists(HISTORY_FILE):
    with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
        history = json.load(f)

# 构建不一致数据的JSON（用于history存储和HTML内嵌）
disagreements_json = {}
uncomparables_json = {}
for pair_name, items in all_disagreements.items():
    if pair_name.endswith('_uncomparable'):
        # 无法判断一致性的数据
        real_pair = pair_name.replace('_uncomparable', '')
        uncomparables_json[real_pair] = []
        for d in items:
            uncomparables_json[real_pair].append({
                'idx': d['idx'] + 1,
                'query': d['query'],
                'grade': d['grade'],
                'course': d['course'],
                'reason': d['reason'],
                'qw_answer': d['qw_answer'],
                'comp_answer': d['comp_answer'],
            })
    else:
        disagreements_json[pair_name] = []
        for d in items:
            disagreements_json[pair_name].append({
                'idx': d['idx'] + 1,
                'query': d['query'],
                'grade': d['grade'],
                'course': d['course'],
                'qw_doc': d['qw_doc'],
                'comp_doc': d['comp_doc'],
                'qw_answer': d['qw_answer'],
                'comp_answer': d['comp_answer'],
                'qw_image_url': d.get('qw_image_url', ''),
                'comp_image_url': d.get('comp_image_url', ''),
            })

# 模型级别统计：提取率 + 空文本率
model_stats = {}
for json_key, col_name in MODEL_FIELDS:
    total = len(data)
    empty = sum(1 for record in data if not record.get(json_key) or str(record.get(json_key, '')).strip().lower() in ('none', 'null', ''))
    extracted = sum(1 for ra in all_answers if ra[json_key][0] is not None)
    model_stats[col_name] = {
        'total': total,
        'empty_text': empty,
        'extracted': extracted,
        'extraction_rate': round(extracted / total * 100, 1) if total > 0 else 0,
        'empty_rate': round(empty / total * 100, 1) if total > 0 else 0,
    }

# 如果同日期已有记录则替换
history = [h for h in history if h['date'] != eval_date]
history.append({
    'date': eval_date,
    'total_questions': len(data),
    'results': comparison_results,
    'grade_stats': grade_stats,
    'course_stats': course_stats,
    'model_stats': model_stats,
    'disagreements': disagreements_json,
    'uncomparables': uncomparables_json,
})
history.sort(key=lambda x: x['date'])

with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
    json.dump(history, f, ensure_ascii=False, indent=2)
print(f"历史数据已保存: {HISTORY_FILE} (共 {len(history)} 次评测)")

# === 生成HTML报告 ===
print("\n生成HTML报告...")

def esc(s):
    """HTML转义"""
    return html_module.escape(str(s)) if s else ''

def truncate_text(text, max_len=200):
    if not text or len(text) <= max_len:
        return text
    return text[:max_len] + '...'

# HTML模板
html_content = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=5.0, user-scalable=yes">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<title>多模型解题答案一致性评测报告</title>
<script src="https://cdn.bootcdn.net/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script>if(typeof Chart==='undefined'){document.write('<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"><\\/script>')}</script>
<link rel="stylesheet" href="https://cdn.bootcdn.net/ajax/libs/KaTeX/0.16.9/katex.min.css" onerror="this.href='https://cdnjs.cloudflare.com/ajax/libs/KaTeX/0.16.9/katex.min.css'">
<script>KATEX_INLINE_PLACEHOLDER</script>
<script>KATEX_AUTORENDER_INLINE_PLACEHOLDER</script>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
html { overflow-x:hidden; }
body { font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif; background:#ffffff; color:#1e293b; min-height:100vh; margin:0; max-width:100vw; overflow-x:hidden; }
.page-wrap { width:100%; max-width:100vw; overflow-x:hidden; position:relative; }

/* === PC端：左侧栏布局 === */
@media (min-width: 769px) {
    .page-wrap { display:flex; overflow:hidden; height:100vh; }
    .sidebar { width:120px; background:linear-gradient(180deg, #ede9fe 0%, #e0d9fc 100%); border-right:1px solid #d4c8f8; padding:12px 0; position:fixed; top:0; left:0; bottom:0; overflow-y:auto; z-index:100; }
    .mobile-header { display:none; }
    .main { margin-left:120px; flex:1; padding:32px; min-width:0; height:100vh; overflow-y:auto; background:#f8f9fc; }
}

/* === 移动端：顶部日期栏 === */
@media (max-width: 768px) {
    body { display:block; overflow-y:auto; overflow-x:hidden; background:#f8f9fc; width:100%; }
    .sidebar { display:none; }
    .mobile-header { display:block; position:sticky; top:0; z-index:200; background:linear-gradient(180deg, #ede9fe 0%, #e0d9fc 100%); border-bottom:1px solid #d4c8f8; padding:12px 16px; }
    .mobile-header .mh-title { font-size:14px; color:#4c1d95; font-weight:600; margin-bottom:8px; }
    .mobile-header .date-scroll { display:flex; gap:8px; overflow-x:auto; -webkit-overflow-scrolling:touch; padding-bottom:4px; scrollbar-width:none; }
    .mobile-header .date-scroll::-webkit-scrollbar { display:none; }
    .mobile-header .date-chip { flex-shrink:0; padding:6px 14px; border-radius:20px; font-size:13px; background:#ffffff; color:#6d28d9; border:1px solid #d4c8f8; cursor:pointer; white-space:nowrap; transition:all 0.2s; }
    .mobile-header .date-chip.active { background:#7c3aed; color:#ffffff; border-color:#7c3aed; }
    .main { margin-left:0; padding:16px; min-width:0; width:100%; overflow-x:hidden; background:#f8f9fc; }
    .card { max-width:100%; overflow:hidden; }
}

/* === 侧栏样式（PC端） === */
.sidebar h3 { padding:0 10px; font-size:11px; color:#7c3aed; text-transform:uppercase; letter-spacing:1px; margin-bottom:8px; }
.sidebar .logo { padding:10px 10px 14px; border-bottom:1px solid #d4c8f8; margin-bottom:10px; }
.sidebar .logo h2 { font-size:13px; color:#4c1d95; }
.sidebar .logo p { font-size:9px; color:#7c3aed; margin-top:2px; }
.sidebar .nav-item { display:block; padding:7px 10px; font-size:12px; color:#6d28d9; cursor:pointer; transition:all 0.15s; border-left:3px solid transparent; }
.sidebar .nav-item:hover { background:rgba(124,58,237,0.08); color:#4c1d95; }
.sidebar .nav-item.active { background:rgba(124,58,237,0.12); color:#4c1d95; border-left-color:#7c3aed; }
.sidebar .nav-item .date-label { font-weight:600; }
.sidebar .nav-item .date-count { font-size:11px; color:#8b5cf6; margin-top:2px; }

/* === 主内容区 === */
.header { margin-bottom:24px; position:relative; }
.header h1 { font-size:22px; color:#1e293b; line-height:1.4; }
.header .subtitle { color:#6b7280; font-size:13px; margin-top:6px; }
.header .download-group { display:flex; gap:8px; flex-wrap:wrap; margin-top:12px; }
.download-btn { display:inline-block; padding:7px 14px; background:#6366f1; color:#ffffff; border:none; border-radius:6px; font-size:12px; text-decoration:none; transition:all 0.2s; white-space:nowrap; }
.download-btn:hover { background:#4f46e5; color:#ffffff; }
@media (min-width:769px) { .header h1 { font-size:28px; } .header .download-group { position:absolute; top:0; right:0; margin-top:0; } }

.card { background:#ffffff; border:1px solid #e5e7eb; border-radius:14px; padding:20px; margin-bottom:20px; box-shadow:0 1px 3px rgba(99,102,241,0.04); }
@media (min-width:769px) { .card { padding:28px; margin-bottom:24px; } }
.card h2 { font-size:16px; color:#1e293b; margin-bottom:14px; padding-bottom:10px; border-bottom:1px solid #e5e7eb; line-height:1.5; }
@media (min-width:769px) { .card h2 { font-size:18px; margin-bottom:16px; padding-bottom:12px; } }
.card h3 { font-size:15px; color:#374151; margin:20px 0 12px; }

/* === 概览网格 === */
.overview-grid { display:grid; grid-template-columns:repeat(auto-fit, minmax(140px, 1fr)); gap:12px; }
@media (min-width:900px) { .overview-grid { grid-template-columns:repeat(5,1fr); gap:14px; } }
.stat-box { background:#f5f3ff; border:1px solid #e0d9fc; border-radius:12px; padding:16px; text-align:center; }
@media (min-width:769px) { .stat-box { padding:20px; } }
.stat-box .label { font-size:12px; color:#6b7280; margin-bottom:6px; }
.stat-box .value { font-size:24px; font-weight:700; }
@media (min-width:769px) { .stat-box .value { font-size:28px; } }
.stat-box .sub { font-size:11px; color:#9ca3af; margin-top:4px; }
.alert-light { margin-top:8px; padding:6px 8px; background:rgba(239,68,68,0.06); border:1px solid rgba(239,68,68,0.2); border-radius:6px; display:flex; align-items:center; gap:6px; }
.alert-dot { width:8px; height:8px; border-radius:50%; background:#ef4444; box-shadow:0 0 6px #ef4444; flex-shrink:0; animation:pulse-dot 1.5s infinite; }
.alert-text { font-size:11px; color:#dc2626; line-height:1.3; text-align:left; }
@keyframes pulse-dot { 0%,100%{opacity:1} 50%{opacity:0.4} }

.rate-high { color:#16a34a !important; }
.rate-mid { color:#d97706 !important; }
.rate-low { color:#dc2626 !important; }
.note-box { background:rgba(99,102,241,0.04); border-left:3px solid #6366f1; padding:12px 16px; border-radius:0 8px 8px 0; margin:12px 0; font-size:13px; color:#4338ca; }

/* === 表格 === */
.table-scroll-wrap { overflow-x:auto; -webkit-overflow-scrolling:touch; margin-top:12px; border-radius:8px; border:1px solid #e5e7eb; }
table { width:100%; border-collapse:collapse; font-size:13px; min-width:500px; }
th,td { padding:10px 12px; text-align:left; border-bottom:1px solid #f3f4f6; white-space:nowrap; }
th { background:#f5f3ff; color:#6366f1; font-weight:600; font-size:11px; text-transform:uppercase; letter-spacing:0.5px; position:sticky; top:0; z-index:1; }
td { color:#374151; }
tr:hover { background:rgba(99,102,241,0.03); }
@media (max-width:768px) { table { font-size:12px; } th,td { padding:8px 10px; } }

.tag { display:inline-block; padding:2px 7px; border-radius:4px; font-size:11px; font-weight:500; }
.tag-green { background:rgba(22,163,74,0.08); color:#16a34a; }
.tag-yellow { background:rgba(217,119,6,0.08); color:#d97706; }
.tag-red { background:rgba(220,38,38,0.08); color:#dc2626; }

/* === 图表 === */
.chart-container { position:relative; height:220px; margin:16px 0; }
@media (min-width:769px) { .chart-container { height:280px; } }

/* === Tab切换 === */
.tabs { display:flex; gap:0; margin-bottom:0; border-bottom:1px solid #e5e7eb; overflow-x:auto; -webkit-overflow-scrolling:touch; scrollbar-width:none; }
.tabs::-webkit-scrollbar { display:none; }
.tab-btn { padding:10px 16px; font-size:13px; color:#6b7280; cursor:pointer; border-bottom:2px solid transparent; transition:all 0.15s; background:none; border-top:none; border-left:none; border-right:none; white-space:nowrap; flex-shrink:0; }
.tab-btn:hover { color:#1e293b; }
.tab-btn.active { color:#6366f1; border-bottom-color:#6366f1; }
.tab-content { display:none; }
.tab-content.active { display:block; }

/* === 不一致详情表格 === */
.detail-table-wrap { overflow:auto; -webkit-overflow-scrolling:touch; border:1px solid #e5e7eb; border-radius:8px; margin-top:12px; }
@media (min-width:769px) { .detail-table-wrap { max-height:600px; } }
@media (max-width:768px) { .detail-table-wrap { max-height:none; } }
.detail-table-wrap table { margin-top:0; min-width:480px; }
.detail-table-wrap th { background:#f5f3ff; }
.detail-table-wrap th:nth-child(1) { width:40px; }
.detail-table-wrap th:nth-child(2) { width:25%; min-width:100px; }
.detail-table-wrap th:nth-child(3) { width:25%; min-width:100px; }
.detail-table-wrap th:nth-child(4) { width:25%; min-width:100px; }
.detail-table-wrap th:nth-child(5) { width:60px; }
.detail-table-wrap td { overflow:hidden; text-overflow:ellipsis; white-space:nowrap; font-size:12px; color:#4b5563; }
.detail-table-wrap td:nth-child(3),
.detail-table-wrap td:nth-child(4) { white-space:normal; word-break:break-all; }
.btn-detail { background:#e5e7eb; border:none; color:#374151; padding:4px 8px; border-radius:4px; cursor:pointer; font-size:11px; white-space:nowrap; }
.btn-detail:hover { background:#d1d5db; }
.doc-cell { max-width:300px; max-height:60px; overflow:hidden; text-overflow:ellipsis; font-size:11px; line-height:1.4; cursor:pointer; position:relative; color:#6b7280; }
.doc-cell:hover { max-height:none; overflow:visible; position:relative; z-index:10; }
.doc-cell::after { content:'点击展开'; position:absolute; bottom:0; right:0; font-size:10px; color:#9ca3af; background:#ffffff; padding:0 4px; }
.doc-cell:hover::after { display:none; }
.answer-tag { display:inline-block; background:rgba(0,0,0,0.03); color:#4b5563; padding:3px 8px; border-radius:4px; font-size:11px; font-weight:400; margin:3px 3px 3px 0; line-height:1.4; }

/* === Modal === */
.modal-overlay { display:none; position:fixed; top:0; left:0; right:0; bottom:0; background:rgba(0,0,0,0.4); z-index:1000; justify-content:center; align-items:center; }
.modal-overlay.show { display:flex; }
.modal { background:#ffffff; border:1px solid #e5e7eb; border-radius:16px; width:90%; max-width:900px; max-height:85vh; overflow-y:auto; padding:24px; box-shadow:0 20px 60px rgba(99,102,241,0.15); }
@media (max-width:768px) { .modal { width:100%; max-width:100%; max-height:100vh; height:100vh; border-radius:0; padding:16px; padding-top:60px; } }
.modal h3 { margin-bottom:16px; color:#1e293b; font-size:15px; }
.modal .doc-content { background:#f8f9fc; border:1px solid #e5e7eb; border-radius:8px; padding:12px; margin:12px 0; font-size:13px; line-height:1.7; white-space:pre-wrap; word-break:break-word; max-height:300px; overflow-y:auto; color:#374151; }
@media (min-width:769px) { .modal .doc-content { padding:16px; max-height:800px; } }
.modal .doc-content .katex { font-size:1.1em; color:#1e293b; }
.modal .doc-content .katex-display { margin:8px 0; overflow-x:auto; }
.modal .close-btn { position:fixed; top:16px; right:16px; background:#6366f1; border:none; color:#ffffff; padding:8px 16px; border-radius:6px; cursor:pointer; font-size:13px; z-index:1001; }
@media (min-width:769px) { .modal .close-btn { position:sticky; top:0; float:right; } }
.modal .close-btn:hover { background:#4f46e5; }

/* === 趋势曲线标题区 === */
.trend-header { display:flex; align-items:center; justify-content:space-between; margin-bottom:16px; padding-bottom:12px; border-bottom:1px solid #e5e7eb; flex-wrap:wrap; gap:8px; }
.trend-header h2 { margin:0 !important; padding:0 !important; border:none !important; }
#chart-legend { display:flex; gap:12px; flex-wrap:wrap; }
@media (max-width:768px) { #chart-legend { gap:8px; } #chart-legend span { font-size:11px !important; } }

/* === 方法说明弹窗适配 === */
#method-modal-overlay .modal table { min-width:auto; }
#method-modal-overlay .modal th, #method-modal-overlay .modal td { white-space:normal; word-break:break-word; }
@media (max-width:768px) { #method-modal-overlay .modal th { width:80px !important; font-size:12px; } #method-modal-overlay .modal td { font-size:12px; } }
</style>
</head>
<body>
<div class="page-wrap">

<!-- PC侧栏 -->
<div class="sidebar">
<div class="logo">
<h2>评测报告</h2>
<p>多模型答案一致性</p>
</div>
<h3>评测日期</h3>
<div id="nav-dates"></div>
</div>

<!-- 移动端顶部导航 -->
<div class="mobile-header">
<div class="mh-title">千问与竞对模型解题答案一致性评测</div>
<div class="date-scroll" id="mobile-dates"></div>
</div>

<div class="main">
<div class="header">
<h1>千问与竞对模型解题【答案一致性】评测 <span id="header-date" style="font-size:14px;font-weight:normal;color:#6b7280;"></span></h1>
<p class="subtitle">以千问为基准，与其他模型进行答案一致性对比分析 · 构建时间: BUILDTIME_PLACEHOLDER · 数据日期: <span id="header-date"></span></p>
<div class="download-group">
<a href="javascript:void(0)" class="download-btn" style="color:#ffffff;background:#7c3aed;border-color:#7c3aed;" onclick="document.getElementById('method-modal-overlay').style.display='flex'">评测方法说明</a>
</div>
</div>

<!-- 模块一：总体概览 -->
<div class="card" id="section-overview">
<h2 style="display:flex;align-items:baseline;gap:12px;flex-wrap:wrap;">总体一致率概览 <span id="overview-date" style="font-size:13px;color:#6b7280;"></span><span style="font-size:12px;color:#6b7280;font-weight:normal;margin-left:auto;">说明：一致率 = 一致题数 / 可比对题数 × 100%（DS抓取待打通暂缺）</span></h2>
<div class="overview-grid" id="overview-grid"></div>
<!-- 一致率异动分析面板（有预警时显示） -->
<div class="anomaly-panel" id="anomaly-panel" style="display:none;margin-top:12px;padding:12px 16px;background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;">
<div style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
<span style="font-size:15px;">⚠️</span>
<span style="font-weight:700;color:#92400e;font-size:13px;">一致率异动分析</span>
</div>
<div id="anomaly-content" style="font-size:12px;color:#78350f;line-height:1.5;"></div>
</div>
</div>

<!-- 模块二：趋势折线图 -->
<div class="card">
<div class="trend-header">
<h2>答案一致率趋势曲线</h2>
<div id="chart-legend"></div>
</div>
<div class="chart-container"><canvas id="trendChart"></canvas></div>
</div>

<!-- 模块三：按年级统计 -->
<div class="card" id="section-grade">
<h2>按学段统计一致率</h2>
<div class="table-scroll-wrap" id="grade-table-wrap"></div>
</div>

<!-- 模块四：按学科统计 -->
<div class="card" id="section-course">
<h2>按学科统计一致率</h2>
<div class="table-scroll-wrap" id="course-table-wrap" style="max-height:320px;overflow-y:auto;"></div>
</div>

<!-- 模块五：不一致详情 -->
<div class="card" id="section-details">
<h2>答案不一致的数据明细</h2>
<div class="tabs" id="detail-tabs"></div>
<div id="detail-contents"></div>
</div>

<!-- 模块六：无法判断一致性的数据 -->
<div class="card" id="section-uncomparables">
<h2>无法判断一致性的数据</h2>
<p style="font-size:13px;color:#6b7280;margin-bottom:12px;">以下记录因模型无法作答或答案无法结构化提取，未纳入一致性统计。</p>
<div class="tabs" id="uncomp-tabs"></div>
<div id="uncomp-contents"></div>
</div>

<!-- 方法说明弹窗 -->
<div class="modal-overlay" id="method-modal-overlay" onclick="if(event.target===this)this.style.display='none'" style="display:none;">
<div class="modal" style="max-width:750px;">
<button class="close-btn" onclick="document.getElementById('method-modal-overlay').style.display='none'">关闭</button>
<h2 style="margin-bottom:16px;">评测方法说明</h2>
<table>
<tr><th style="width:130px;">评测基准</th><td>以千问(app_qianwen_text)为参照，分别与豆包、豆包深思、元宝、DeepSeek、Gemini3.1进行答案比对</td></tr>
<tr><th>答案提取</th><td>多层正则匹配：选择题(A-E) → 判断题(正确/错误) → 填空/结论题 → 数值题。复杂解答/论述/多步证明题暂无法结构化提取</td></tr>
<tr><th>一致性判定</th><td>选择题按选项集合比对；判断题语义归一化；数值题精度比对（误差&lt;0.01）；填空题深度归一化+子串匹配+数字集合比较；LaTeX格式统一去除后比对</td></tr>
<tr><th>DeepSeek说明</th><td>ds字段在源数据中仅为日期值，不含模型解题内容，DS抓取待打通</td></tr>
</table>
<h3 style="margin-top:20px;margin-bottom:10px;font-size:14px;color:#1e293b;">关于"无法判断"</h3>
<p style="font-size:13px;color:#374151;line-height:1.8;margin-bottom:8px;">以下情况记录不参与一致率计算（不计入分母），标记为"无法判断"：</p>
<table style="font-size:13px;">
<tr><th style="width:160px;">模型明确无法作答</th><td>模型回复中包含"图片不清/关键信息缺失/无法解答"等信号，表示未尝试给出答案。这不等于"答错"，不应计为不一致。<br><span style="color:#9ca3af;">示例：豆包回复"这道题缺少关键配图，无法确定答案"</span></td></tr>
<tr><th>答案无法结构化提取</th><td>模型给出了完整解题过程，但算法无法从中抽取可比对的结构化答案（如长篇论述、多步证明、开放式分析）。此时无法判定是否一致。<br><span style="color:#9ca3af;">示例：数学证明题，两模型各写500字推导，可能结论相同但表述不同</span></td></tr>
<tr><th>一方文本为空</th><td>API未返回该模型的解题结果（接口超时或抓取失败），无内容可比对</td></tr>
</table>
<h3 style="margin-top:20px;margin-bottom:10px;font-size:14px;color:#1e293b;">关于"不可用query"</h3>
<p style="font-size:13px;color:#374151;line-height:1.8;">当前数据源API不提供query有效性标注字段。所有query均为图片URL，各模型硬性输出解题结果，未预先判断图片是否为有效考题。少数情况下某模型会自行判断"这不是一道题"（如元宝），但验证发现该判断不可靠（曾将有效数学题误判为非题目），故暂不依据单一模型判断做全局query排除。</p>
<div style="text-align:right;margin-top:20px;"><a href="javascript:void(0)" onclick="document.getElementById('method-modal-overlay').style.display='none'" style="color:#6366f1;text-decoration:none;font-size:14px;font-weight:500;">关闭</a></div>
</div>
</div>
</div>

<!-- 详情弹窗 -->
<div class="modal-overlay" id="modal-overlay" onclick="if(event.target===this)closeModal()">
<div class="modal" id="modal-content"></div>
</div>

<script>
// === 数据注入 ===
const historyData = HISTORY_DATA_PLACEHOLDER;
const pairNames = PAIR_NAMES_PLACEHOLDER;

// === 版本检测（每次会话最多刷新一次） ===
const BUILD_TIME = 'BUILDTIME_PLACEHOLDER';
(function(){
    if (location.hostname === 'localhost' || location.hostname === '127.0.0.1') return;
    if (sessionStorage.getItem('_ver_checked')) return;
    sessionStorage.setItem('_ver_checked', '1');
    var vurl = 'https://paultunggm-pixel.github.io/consistency-eval/version.json?t=' + Date.now();
    var xhr = new XMLHttpRequest();
    xhr.open('GET', vurl, true);
    xhr.timeout = 5000;
    xhr.onload = function(){
        if (xhr.status === 200) {
            try {
                var remote = JSON.parse(xhr.responseText);
                if (remote.build_time && remote.build_time > BUILD_TIME) {
                    location.reload(true);
                }
            } catch(e) {}
        }
    };
    xhr.send();
})();

// === 初始化 ===
let currentDateIdx = historyData.length - 1;
const pairColors = ['#6366f1','#0891b2','#16a34a','#d97706','#dc2626'];

function rateClass(rate) {
    if (rate >= 75) return 'rate-high';
    if (rate >= 55) return 'rate-mid';
    return 'rate-low';
}
function tagClass(rate) {
    if (rate >= 75) return 'tag-green';
    if (rate >= 55) return 'tag-yellow';
    return 'tag-red';
}

// 渲染左侧导航
function renderNav() {
    const container = document.getElementById('nav-dates');
    container.innerHTML = historyData.map((h, i) => `
        <div class="nav-item ${i === currentDateIdx ? 'active' : ''}" onclick="switchDate(${i})">
            <div class="date-label">${h.date}</div>
            <div class="date-count">${h.total_questions} 题</div>
        </div>
    `).reverse().join('');
    // 移动端顶部日期chips
    var mobileEl = document.getElementById('mobile-dates');
    if (mobileEl) {
        mobileEl.innerHTML = historyData.map((h, i) => `<div class="date-chip ${i === currentDateIdx ? 'active' : ''}" onclick="switchDate(${i})">${h.date} (${h.total_questions}题)</div>`).reverse().join('');
    }
}

// 切换日期
function switchDate(idx) {
    currentDateIdx = idx;
    renderNav();
    renderOverview();
    renderGradeTable();
    renderCourseTable();
    renderDetails();
    renderUncomparables();
}

// 渲染总体概览
function renderOverview() {
    const h = historyData[currentDateIdx];
    document.getElementById('overview-date').textContent = `(${h.date})`;
    document.getElementById('header-date').textContent = h.date;
    const grid = document.getElementById('overview-grid');
    var curUncomp = h.uncomparables || {};
    grid.innerHTML = pairNames.map((name, i) => {
        const r = h.results[name] || {rate:0, matched:0, total_comparable:0, uncomparable_count:0};
        const color = pairColors[i];
        const uncompCount = r.uncomparable_count || (curUncomp[name] ? curUncomp[name].length : 0);
        // 警示红灯：与前三次均值对比，偏差>=10%时触发
        let alertHtml = '';
        const prevRuns = historyData.filter((_, i) => i < currentDateIdx).slice(-3);
        if (prevRuns.length >= 3) {
            const avgRate = prevRuns.reduce((sum, run) => {
                const pr = run.results[name] || {rate:0};
                return sum + pr.rate;
            }, 0) / prevRuns.length;
            const diff = r.rate - avgRate;
            if (Math.abs(diff) >= 10) {
                const direction = diff > 0 ? '高' : '低';
                alertHtml = `<div class="alert-light">
                    <span class="alert-dot"></span>
                    <span class="alert-text">该指标本次数值比前三次平均值${direction}${Math.abs(diff).toFixed(1)}%</span>
                </div>`;
            }
        }
        return `<div class="stat-box">
            <div class="label">${name}</div>
            <div class="value" style="color:${color}">${r.rate}%</div>
            <div class="sub">${r.matched} / ${r.total_comparable} 题一致</div>
            <div class="sub" style="color:#9ca3af;font-size:11px;margin-top:2px;">无法判断 ${uncompCount} 题</div>
            ${alertHtml}
        </div>`;
    }).join('');
    // v1.7: 渲染异动分析面板
    renderAnomalyAnalysis();
}

// v1.7: 一致率异动分析
function renderAnomalyAnalysis() {
    const panel = document.getElementById('anomaly-panel');
    const content = document.getElementById('anomaly-content');
    if (!panel || !content) return;

    const h = historyData[currentDateIdx];
    if (currentDateIdx < 3) { panel.style.display = 'none'; return; }

    const prevRuns = historyData.filter((_, i) => i < currentDateIdx).slice(-3);
    if (prevRuns.length < 3) { panel.style.display = 'none'; return; }

    const anomalies = [];
    pairNames.forEach(name => {
        const r = (h.results[name] || {rate:0});
        const avgRate = prevRuns.reduce((sum, run) => sum + (run.results[name]||{rate:0}).rate, 0) / prevRuns.length;
        const diff = r.rate - avgRate;
        if (Math.abs(diff) >= 10) {
            anomalies.push({name, rate: r.rate, avgRate: avgRate.toFixed(1), diff: diff.toFixed(1),
                direction: diff > 0 ? '↑上升' : '↓下降', matched: r.matched, comparable: r.total_comparable});
        }
    });

    if (anomalies.length === 0) { panel.style.display = 'none'; return; }
    panel.style.display = 'block';

    const ms = h.model_stats || {};
    let leftHtml = '';
    let rightHtml = '';

    // 左列：异动指标 + 数据源诊断
    leftHtml += '<div style="font-weight:700;margin-bottom:4px;">📊 异动指标</div>';
    anomalies.forEach(a => {
        leftHtml += '<div style="margin-bottom:2px;padding:3px 8px;background:rgba(251,191,36,0.15);border-radius:4px;display:inline-block;margin-right:4px;font-size:11px;">'
            + '<b>' + a.name + '</b> ' + a.rate + '% '
            + '<span style="color:' + (parseFloat(a.diff)>0?'#16a34a':'#dc2626') + '">' + a.direction + Math.abs(parseFloat(a.diff)) + 'pp</span>'
            + '</div>';
    });

    // 空文本率诊断
    if (Object.keys(ms).length > 0) {
        const highEmpty = [];
        Object.entries(ms).forEach(([name, stats]) => {
            if (stats.empty_rate > 20) {
                highEmpty.push(name + ' <b>' + stats.empty_rate + '%</b>');
            }
        });
        if (highEmpty.length > 0) {
            leftHtml += '<div style="margin-top:6px;font-weight:700;margin-bottom:2px;">🔍 空文本率异常</div>';
            leftHtml += '<div style="font-size:11px;">' + highEmpty.join(' &nbsp;|&nbsp; ') + '</div>';
            leftHtml += '<div style="font-size:11px;color:#b45309;margin-top:2px;">→ 千问-元宝样本过少不可靠</div>';
        }
    }

    // 右列：提取诊断 + 建议
    if (Object.keys(ms).length > 0) {
        const lowExtract = [];
        Object.entries(ms).forEach(([name, stats]) => {
            const extractFailRate = 100 - stats.extraction_rate;
            if (extractFailRate > 15 && stats.empty_rate < 50) {
                lowExtract.push(name + ' <b>' + extractFailRate.toFixed(1) + '%</b>');
            }
        });
        if (lowExtract.length > 0) {
            rightHtml += '<div style="font-weight:700;margin-bottom:2px;">📝 提取异常</div>';
            rightHtml += '<div style="font-size:11px;margin-bottom:4px;">' + lowExtract.join(' &nbsp;|&nbsp; ') + '</div>';
        }
    }

    rightHtml += '<div style="font-weight:700;margin-bottom:2px;">💡 排查方向</div>';
    rightHtml += '<div style="font-size:11px;">· 确认API数据源是否正常（竞品大面积空返回）</div>';
    rightHtml += '<div style="font-size:11px;">· 检查学科/学段分布是否与历史显著偏移</div>';
    rightHtml += '<div style="font-size:11px;">· 排查"同URL不同题"问题</div>';

    content.innerHTML = '<div style="display:flex;gap:16px;flex-wrap:wrap;">'
        + '<div style="flex:1;min-width:260px;">' + leftHtml + '</div>'
        + '<div style="flex:1;min-width:200px;">' + rightHtml + '</div>'
        + '</div>';
}

// 渲染折线图
function renderTrendChart() {
    if (typeof Chart === 'undefined') {
        document.getElementById('trendChart').parentElement.innerHTML = '<p style="color:#f87171;padding:20px;">Chart.js CDN加载失败，趋势图无法显示。请检查网络或刷新页面。</p>';
        return;
    }
    const ctx = document.getElementById('trendChart').getContext('2d');
    const labels = historyData.map(h => h.date);
    const datasets = pairNames.map((name, i) => ({
        label: name,
        data: historyData.map(h => (h.results[name] || {rate:0}).rate),
        borderColor: pairColors[i],
        backgroundColor: pairColors[i] + '20',
        tension: 0.3,
        pointRadius: 4,
        pointHoverRadius: 6,
        borderWidth: 2,
    }));
    // 自定义图例渲染到标题右侧
    var legendEl = document.getElementById('chart-legend');
    if (legendEl) {
        legendEl.innerHTML = pairNames.map(function(name, i) {
            return '<span style="display:flex;align-items:center;gap:5px;font-size:12px;color:#6b7280;"><span style="width:12px;height:3px;background:' + pairColors[i] + ';border-radius:2px;display:inline-block;"></span>' + name + '</span>';
        }).join('');
    }
    new Chart(ctx, {
        type: 'line',
        data: { labels, datasets },
        options: {
            responsive: true, maintainAspectRatio: false,
            plugins: {
                legend: { display: false },
                datalabels: { display: false }
            },
            scales: {
                x: { ticks:{color:'#6b7280'}, grid:{color:'rgba(99,102,241,0.08)', borderDash:[4,4]} },
                y: {
                    ticks:{color:'#6b7280',callback:v=>v+'%'},
                    grid:{color:'rgba(99,102,241,0.08)', borderDash:[4,4]},
                    ...(() => {
                        const allVals = datasets.flatMap(ds => ds.data.filter(v => v > 0));
                        if (allVals.length === 0) return {min:0, max:100};
                        const lo = Math.min(...allVals), hi = Math.max(...allVals);
                        const yMin = Math.max(0, Math.floor(lo - 2));
                        const yMax = Math.min(100, Math.ceil(hi + 2));
                        const range = yMax - yMin;
                        const step = range <= 10 ? 2 : range <= 20 ? 4 : 5;
                        return { min: yMin, max: yMax, ticks: { color:'#64748b', callback: function(v){return v+'%';}, stepSize: step } };
                    })()
                }
            }
        },
        plugins: [{
            id: 'dataLabels',
            afterDatasetsDraw: function(chart) {
                var ctx2 = chart.ctx;
                ctx2.font = '11px sans-serif';
                ctx2.textAlign = 'center';
                chart.data.datasets.forEach(function(dataset, i) {
                    var meta = chart.getDatasetMeta(i);
                    meta.data.forEach(function(point, index) {
                        var value = dataset.data[index];
                        if (value > 0) {
                            ctx2.fillStyle = dataset.borderColor;
                            ctx2.fillText(value + '%', point.x, point.y - 10);
                        }
                    });
                });
            }
        }]
    });
}

// 渲染年级表格
function renderGradeTable() {
    const h = historyData[currentDateIdx];
    const grades = Object.keys(h.grade_stats[pairNames[0]] || {}).sort();
    let html = '<table><thead><tr><th>年级</th>';
    pairNames.forEach(n => { html += `<th>${n}</th>`; });
    html += '</tr></thead><tbody>';
    grades.forEach(g => {
        html += `<tr><td><strong>${g}</strong></td>`;
        pairNames.forEach(n => {
            const s = (h.grade_stats[n] || {})[g] || {rate:0,matched:0,comparable:0};
            if (s.comparable > 0) {
                html += `<td><span class="tag ${tagClass(s.rate)}">${s.rate}%</span> <small style="color:#475569">(${s.matched}/${s.comparable})</small></td>`;
            } else {
                html += '<td><small style="color:#475569">N/A</small></td>';
            }
        });
        html += '</tr>';
    });
    html += '</tbody></table>';
    document.getElementById('grade-table-wrap').innerHTML = html;
}

// 渲染学科表格
function renderCourseTable() {
    const h = historyData[currentDateIdx];
    const courses = Object.keys(h.course_stats[pairNames[0]] || {}).sort();
    let html = '<table><thead><tr><th>学科</th>';
    pairNames.forEach(n => { html += `<th>${n}</th>`; });
    html += '</tr></thead><tbody>';
    courses.forEach(c => {
        html += `<tr><td><strong>${c}</strong></td>`;
        pairNames.forEach(n => {
            const s = (h.course_stats[n] || {})[c] || {rate:0,matched:0,comparable:0};
            if (s.comparable > 0) {
                html += `<td><span class="tag ${tagClass(s.rate)}">${s.rate}%</span> <small style="color:#475569">(${s.matched}/${s.comparable})</small></td>`;
            } else {
                html += '<td><small style="color:#475569">N/A</small></td>';
            }
        });
        html += '</tr>';
    });
    html += '</tbody></table>';
    document.getElementById('course-table-wrap').innerHTML = html;
}

// 渲染不一致详情
function renderDetails() {
    const tabsEl = document.getElementById('detail-tabs');
    const contentsEl = document.getElementById('detail-contents');
    const curDisagreements = historyData[currentDateIdx].disagreements || {};
    let tabsHtml = '';
    let contentsHtml = '';

    pairNames.forEach((name, i) => {
        const items = curDisagreements[name] || [];
        const isActive = i === 0 ? ' active' : '';
        tabsHtml += `<button class="tab-btn${isActive}" onclick="switchTab(${i})">${name} (${items.length})</button>`;

        // 确定对比模型名称
        const compName = name.replace('千问-', '');
        contentsHtml += `<div class="tab-content${isActive}" id="tab-${i}">`;
        if (items.length === 0) {
            contentsHtml += '<p style="color:#6b7280;padding:20px;">无不一致数据或无法比对</p>';
        } else {
            contentsHtml += `<div class="detail-table-wrap"><table>
            <thead><tr><th>#</th><th>query</th><th>千问答案抽取</th><th>${compName}答案抽取</th><th>操作</th></tr></thead><tbody>`;
            items.forEach((item, j) => {
                const queryDisplay = item.query.length > 60 ? item.query.substring(0,60)+'...' : item.query;
                contentsHtml += `<tr>
                    <td>${item.idx}</td>
                    <td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${item.query.replace(/"/g,'&quot;')}">${queryDisplay}</td>
                    <td><span class="answer-tag">${renderAnswerContent(item.qw_answer)}</span></td>
                    <td><span class="answer-tag">${renderAnswerContent(item.comp_answer)}</span></td>
                    <td><button class="btn-detail" onclick="showDetail('${name}',${j})">详情</button></td>
                </tr>`;
            });
            contentsHtml += '</tbody></table></div>';
        }
        contentsHtml += '</div>';
    });

    tabsEl.innerHTML = tabsHtml;
    contentsEl.innerHTML = contentsHtml;
    // 渲染答案标签中的LaTeX
    doKaTeX(contentsEl);
}

// 统一的KaTeX渲染函数，带安全检查
function doKaTeX(el) {
    if (typeof renderMathInElement !== 'function') return;
    try {
        renderMathInElement(el, {
            delimiters: [
                {left: '$$', right: '$$', display: true},
                {left: '$', right: '$', display: false},
                {left: '\\\\(', right: '\\\\)', display: false},
                {left: '\\\\[', right: '\\\\]', display: true},
            ],
            throwOnError: false,
            trust: true,
        });
    } catch(e) { console.warn('KaTeX error:', e); }
}

function switchTab(idx) {
    document.querySelectorAll('#detail-tabs .tab-btn').forEach((b,i) => b.classList.toggle('active', i===idx));
    document.querySelectorAll('#detail-contents .tab-content').forEach((c,i) => c.classList.toggle('active', i===idx));
}

function switchUncompTab(idx) {
    document.querySelectorAll('#uncomp-tabs .tab-btn').forEach((b,i) => b.classList.toggle('active', i===idx));
    document.querySelectorAll('#uncomp-contents .tab-content').forEach((c,i) => c.classList.toggle('active', i===idx));
}

function renderUncomparables() {
    var tabsEl = document.getElementById('uncomp-tabs');
    var contentsEl = document.getElementById('uncomp-contents');
    var curUncomparables = historyData[currentDateIdx].uncomparables || {};
    var tabsHtml = '';
    var contentsHtml = '';

    pairNames.forEach(function(name, i) {
        var items = curUncomparables[name] || [];
        var isActive = i === 0 ? ' active' : '';
        tabsHtml += '<button class="tab-btn' + isActive + '" onclick="switchUncompTab(' + i + ')">' + name + ' (' + items.length + ')</button>';

        var compName = name.replace('千问-', '');
        contentsHtml += '<div class="tab-content' + isActive + '">';
        if (items.length === 0) {
            contentsHtml += '<p style="color:#6b7280;padding:20px;">所有数据均可比对</p>';
        } else {
            contentsHtml += '<div class="detail-table-wrap"><table><thead><tr><th style="width:40px;">#</th><th style="width:25%;">query</th><th style="width:20%;">原因</th><th style="width:22%;">千问答案</th><th style="width:22%;">' + compName + '答案</th></tr></thead><tbody>';
            items.forEach(function(item) {
                var queryDisplay = item.query.length > 50 ? item.query.substring(0,50)+'...' : item.query;
                var reasonColor = item.reason.indexOf('无法作答') >= 0 ? '#fbbf24' : '#94a3b8';
                contentsHtml += '<tr><td>' + item.idx + '</td><td style="max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="' + item.query.replace(/"/g,'&quot;') + '">' + queryDisplay + '</td><td style="color:' + reasonColor + ';font-size:12px;">' + item.reason + '</td><td style="font-size:12px;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">' + item.qw_answer + '</td><td style="font-size:12px;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">' + item.comp_answer + '</td></tr>';
            });
            contentsHtml += '</tbody></table></div>';
        }
        contentsHtml += '</div>';
    });

    tabsEl.innerHTML = tabsHtml;
    contentsEl.innerHTML = contentsHtml;
}

let imgRotationState = {};
function rotateImg(imgId, deg) {
    imgRotationState[imgId] = ((imgRotationState[imgId] || 0) + deg) % 360;
    const el = document.getElementById(imgId);
    if (el) el.style.transform = 'rotate(' + imgRotationState[imgId] + 'deg)';
}
function renderQuery(query) {
    // 如果query是图片URL，则直接渲染为img标签，并提供旋转按钮
    if (/^https?:\\/\\/.+\\.(jpg|jpeg|png|gif|webp|bmp)(\\?.*)?$/i.test(query)) {
        const imgId = 'qimg-' + Math.random().toString(36).slice(2,8);
        return `<div style="margin:12px 0;">
            <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
                <button onclick="rotateImg('${imgId}',-90)" style="background:#e5e7eb;border:none;color:#374151;padding:4px 10px;border-radius:4px;cursor:pointer;font-size:12px;" title="逆时针旋转90°">↺ 左转</button>
                <button onclick="rotateImg('${imgId}',90)" style="background:#e5e7eb;border:none;color:#374151;padding:4px 10px;border-radius:4px;cursor:pointer;font-size:12px;" title="顺时针旋转90°">↻ 右转</button>
                <span style="font-size:11px;color:#9ca3af;">点击按钮可摆正图片方向</span>
            </div>
            <div style="overflow:auto;text-align:center;">
                <img id="${imgId}" src="${query}" style="max-width:100%;max-height:800px;border-radius:8px;border:1px solid #334155;image-orientation:from-image;transition:transform 0.3s ease;" onerror="this.onerror=null;this.outerHTML='<p style=\\'color:#f87171\\'>图片加载失败: '+this.src+'</p>'" />
            </div>
            <p style="font-size:11px;color:#475569;margin-top:4px;word-break:break-all;">${query}</p>
        </div>`;
    }
    return `<p style="color:#374151;font-size:13px;margin:8px 0;word-break:break-all;">${query}</p>`;
}

function escapeHtmlButKeepLatex(text) {
    // 先保护LaTeX公式片段（$$...$$, $...$, \\(...\\), \\[...\\]）
    const latexBlocks = [];
    let idx = 0;
    // 保护 $$ ... $$
    text = text.replace(/\\$\\$([\\s\\S]*?)\\$\\$/g, (m) => { latexBlocks.push(m); return '%%LATEX'+(idx++)+'%%'; });
    // 保护 $ ... $（非贪婪，单行内）
    text = text.replace(/\\$([^$\\n]+?)\\$/g, (m) => { latexBlocks.push(m); return '%%LATEX'+(idx++)+'%%'; });
    // 保护 \\[ ... \\]
    text = text.replace(/\\\\\\[([\\s\\S]*?)\\\\\\]/g, (m) => { latexBlocks.push(m); return '%%LATEX'+(idx++)+'%%'; });
    // 保护 \\( ... \\)
    text = text.replace(/\\\\\\(([\\s\\S]*?)\\\\\\)/g, (m) => { latexBlocks.push(m); return '%%LATEX'+(idx++)+'%%'; });
    // HTML转义非LaTeX部分
    text = text.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
    // 还原LaTeX
    text = text.replace(/%%LATEX(\\d+)%%/g, (_, i) => latexBlocks[parseInt(i)]);
    return text;
}

function wrapNakedLatex(text) {
    // 为裸露的LaTeX添加定界符，使KaTeX能识别和渲染
    
    // 0a. 修复千问粘连写法: sinfrac{a}{b} -> sin frac{a}{b}, cosfrac -> cos frac
    text = text.replace(/([^a-zA-Z])(sin|cos|tan|log|ln|lim|max|min)(frac\\{)/g, '$1$2 $3');
    text = text.replace(/^(sin|cos|tan|log|ln|lim|max|min)(frac\\{)/gm, '$1 $2');
    
    // 0b. 修复千问风格无反斜杠LaTeX: frac{a}{b} -> \\frac{a}{b}
    // 带花括号参数的命令
    var braceCmds = 'dfrac|tfrac|cfrac|frac|sqrt|text|mathrm|mathbf|boldsymbol|overline|underline|hat|bar|vec|tilde|binom|begin|end|int|sum|prod|lim|mathbb|mathcal';
    // 匹配: 非反斜杠非字母前导 + 命令名 + {
    text = text.replace(new RegExp('([^\\\\\\\\a-zA-Z])(' + braceCmds + ')(\\\\{)', 'g'), '$1\\\\$2$3');
    // 行首的情况
    text = text.replace(new RegExp('^(' + braceCmds + ')(\\\\{)', 'gm'), '\\\\$1$2');
    
    // 0c. 不带花括号的独立命令 (times, div, cdot, alpha, sin, cos等)
    var standaloneCmds = 'times|div|cdot|cdots|ldots|alpha|beta|gamma|delta|epsilon|zeta|eta|theta|iota|kappa|lambda|mu|nu|xi|pi|rho|sigma|tau|upsilon|phi|chi|psi|omega|Gamma|Delta|Theta|Lambda|Xi|Pi|Sigma|Upsilon|Phi|Psi|Omega|infty|pm|mp|le|ge|leq|geq|ne|neq|approx|equiv|perp|parallel|circ|angle|triangle|therefore|because|quad|qquad|left|right|sin|cos|tan|log|ln|lim|max|min|sum|prod|int|partial|nabla|forall|exists|in|subset|supset|cup|cap|wedge|vee|neg|to|mapsto|rightarrow|leftarrow|Rightarrow|Leftarrow';
    text = text.replace(new RegExp('([^\\\\\\\\a-zA-Z])(' + standaloneCmds + ')([^a-zA-Z{])', 'g'), '$1\\\\$2$3');
    text = text.replace(new RegExp('^(' + standaloneCmds + ')([^a-zA-Z{])', 'gm'), '\\\\$1$2');
    
    // 去除多余重复反斜杠 (\\\\\\\\frac -> \\\\frac)
    text = text.replace(/\\\\{2,}([a-zA-Z])/g, '\\\\$1');
    
    // 1. 保护已有定界符内容
    var prot = [];
    var pi = 0;
    text = text.replace(/\\$\\$[\\s\\S]*?\\$\\$/g, function(m) { prot.push(m); return '%%P'+(pi++)+'%%'; });
    text = text.replace(/\\$[^$\\n]+?\\$/g, function(m) { prot.push(m); return '%%P'+(pi++)+'%%'; });
    text = text.replace(/\\\\\\([\\s\\S]*?\\\\\\)/g, function(m) { prot.push(m); return '%%P'+(pi++)+'%%'; });
    text = text.replace(/\\\\\\[[\\s\\S]*?\\\\\\]/g, function(m) { prot.push(m); return '%%P'+(pi++)+'%%'; });
    
    // 2. \\\\begin{...}...\\\\end{...} -> $$...$$
    text = text.replace(/(\\\\begin\\{[^}]+\\}[\\s\\S]*?\\\\end\\{[^}]+\\})/g, function(m) {
        var wrapped = '$$' + m + '$$';
        prot.push(wrapped); return '%%P'+(pi++)+'%%';
    });
    
    // 3. 识别独占一行或多行的完整数学表达式（元宝风格：整行都是LaTeX）
    // 如: f(x) = -\\frac{1}{2} + \\sum_{n=1}^{\\infty} ...
    // 行尾允许: } 数字 ) ] 字母(单位如V/A/m)
    text = text.replace(/^([^\\n]*\\\\(?:frac|sum|prod|int|left|sin|cos|quad|tau|alpha|beta|gamma|delta|theta|omega|Omega|times|ge|le|infty)[^\\n]*[}\\d)\\]a-zA-Z])\\s*$/gm, function(m) {
        if (m.indexOf('$') >= 0 || m.indexOf('%%P') >= 0) return m;
        // 排除明显的自然语言行（中文字符占比过高）
        var chineseCount = (m.match(/[\\u4e00-\\u9fff]/g) || []).length;
        if (chineseCount > m.length * 0.3) return m;
        var wrapped = '$$' + m.trim() + '$$';
        prot.push(wrapped); return '%%P'+(pi++)+'%%';
    });
    
    // 4. 包裹含\\frac/\\sqrt/\\binom的行内表达式（支持嵌套花括号如 \\frac{L}{R_{eq}}）
    // 使用平衡括号匹配helper
    function matchBrace(s, pos) {
        // 从s[pos]='{'开始匹配到对应'}', 返回结束位置(}后一位)或-1
        if (pos >= s.length || s[pos] !== '{') return -1;
        var depth = 0;
        for (var j = pos; j < s.length; j++) {
            if (s[j] === '{') depth++;
            else if (s[j] === '}') { depth--; if (depth === 0) return j + 1; }
        }
        return -1;
    }
    // 查找并包裹 \\frac{...}{...}, \\sqrt{...}, \\binom{...}{...} 及后续连续数学内容
    var fracRe = /\\\\(?:frac|sqrt|binom)\\{/g;
    var fracMatch;
    var result4 = '';
    var lastIdx4 = 0;
    while ((fracMatch = fracRe.exec(text)) !== null) {
        var cmdStart = fracMatch.index;
        var bracePos = fracMatch.index + fracMatch[0].length - 1; // position of first {
        var end1 = matchBrace(text, bracePos);
        if (end1 === -1) continue;
        var end = end1;
        // 如果是frac/binom，尝试匹配第二个{...}
        if (fracMatch[0].indexOf('frac') >= 0 || fracMatch[0].indexOf('binom') >= 0) {
            var p = end;
            while (p < text.length && text[p] === ' ') p++;
            if (p < text.length && text[p] === '{') {
                var end2 = matchBrace(text, p);
                if (end2 !== -1) end = end2;
            }
        }
        // 尝试继续扩展：后续连续数学（如 = 2 \\times 3 等）
        var contRe = /^(?:[^\\n，。；]*?(?:\\\\[a-zA-Z]+(?:\\{[^}]*\\})*|[0-9]+))/;
        var tail = text.substring(end);
        var contMatch = tail.match(contRe);
        if (contMatch) end += contMatch[0].length;
        var expr = text.substring(cmdStart, end);
        if (expr.indexOf('$') >= 0 || expr.indexOf('%%P') >= 0) {
            fracRe.lastIndex = end;
            continue;
        }
        result4 += text.substring(lastIdx4, cmdStart) + ' $' + expr + '$ ';
        lastIdx4 = end;
        fracRe.lastIndex = end;
    }
    result4 += text.substring(lastIdx4);
    text = result4;
    
    // 5. 包裹其他 \\cmd 或 \\cmd{...} 模式 (行内, 支持嵌套花括号)
    var cmdRe5 = /\\\\[a-zA-Z]+/g;
    var cmdMatch5;
    var result5 = '';
    var lastIdx5 = 0;
    while ((cmdMatch5 = cmdRe5.exec(text)) !== null) {
        var cs = cmdMatch5.index;
        var ce = cmdMatch5.index + cmdMatch5[0].length;
        // 跳过已保护/已包裹的
        if (text.substring(Math.max(0, cs-2), cs).indexOf('$') >= 0) {
            continue;
        }
        // 收集后续{...}参数
        var pos5 = ce;
        while (pos5 < text.length && text[pos5] === '{') {
            var be = matchBrace(text, pos5);
            if (be === -1) break;
            pos5 = be;
        }
        // 收集 _^ 下标上标
        while (pos5 < text.length && (text[pos5] === '_' || text[pos5] === '^')) {
            pos5++;
            if (pos5 < text.length && text[pos5] === '{') {
                var be2 = matchBrace(text, pos5);
                if (be2 !== -1) pos5 = be2;
            } else if (pos5 < text.length && /[0-9a-zA-Z]/.test(text[pos5])) {
                pos5++;
            }
        }
        var expr5 = text.substring(cs, pos5);
        if (expr5.trim().length < 4) continue;
        if (expr5.indexOf('$') >= 0 || expr5.indexOf('%%P') >= 0) continue;
        // 已经在Step 4中被处理的不再重复
        if (cs < lastIdx5) continue;
        result5 += text.substring(lastIdx5, cs) + ' $' + expr5.trim() + '$ ';
        lastIdx5 = pos5;
        cmdRe5.lastIndex = pos5;
    }
    result5 += text.substring(lastIdx5);
    text = result5;
    
    // 5b. 包裹含下标/上标的独立数学表达式(如 R_{eq}, e^{-8t}, i(0_+))
    // 匹配: 字母数字开头 + (_或^) + {内容}或单字符 的模式
    // 注意：跳过已在 $...$ 内部的（Step 4/5产出的）
    text = text.replace(/(?<![\\\\$a-zA-Z])([a-zA-Z][a-zA-Z0-9()]*(?:[_^](?:\\{[^}]*\\}|[0-9a-zA-Z]))(?:[_^](?:\\{[^}]*\\}|[0-9a-zA-Z]))*)/g, function(m, expr, offset) {
        // 只处理确实含有_或^的
        if (expr.indexOf('_') < 0 && expr.indexOf('^') < 0) return m;
        // 不处理已在$内或%%P内的
        if (m.indexOf('$') >= 0 || m.indexOf('%%P') >= 0) return m;
        // 检查是否已在$...$内部（向左查找最近的$是否未闭合）
        var leftContext = text.substring(Math.max(0, offset - 100), offset);
        var dollarCount = (leftContext.match(/\\$/g) || []).length;
        if (dollarCount % 2 === 1) return m; // 在奇数$后面 = 在$...$内部
        // 跳过太短的（如 a_1 单独一个可能是编号）
        if (expr.length < 4) return m;
        // 跳过明显不是数学的（纯下划线变量名如 file_name）
        if (/^[a-z]+_[a-z]+$/i.test(expr) && expr.indexOf('{') < 0) return m;
        return ' $' + expr + '$ ';
    });
    
    // 6. 还原保护内容
    text = text.replace(/%%P(\\d+)%%/g, function(_, i) { return prot[parseInt(i)]; });
    
    // 7. 清理空的行内$对 ($ 空内容 $)
    text = text.replace(/\\$\\s+\\$/g, '');
    
    return text;
}

function renderAnswerContent(text) {
    // 答案标签的轻量LaTeX渲染：按中文逗号分段处理，每段分别判断是否为纯数学
    // 不做整行包裹(Step3)，因为答案可能是中文+公式混合
    if (!text) return '';
    var segments = text.split('\\uff0c');  // 只按中文全角逗号(extract_final_answer_line连接符)
    var rendered = segments.map(function(seg) {
        seg = seg.trim();
        if (!seg) return '';
        // 已有$$定界符的直接保留
        if (seg.indexOf('$$') >= 0) return seg;
        // 已有$...$定界符的直接保留
        if (/\\$[^$]+\\$/.test(seg)) return seg;
        // 判断是否为纯数学段(中文<20%且含数学特征)
        var cnCount = (seg.match(/[\\u4e00-\\u9fff]/g) || []).length;
        var hasMath = /[=^_{}]|\\\\[a-zA-Z]|(?<![a-zA-Z])(?:frac|quad|ge|le|sin|cos|tau|alpha|beta|Omega)(?![a-zA-Z])/.test(seg);
        if (hasMath && cnCount < seg.length * 0.2) {
            // 纯数学段 -> 补裸命令反斜杠后整段包裹$...$
            seg = seg.replace(/(?<![\\\\a-zA-Z])(quad|geq?|leq?|times|infty|tau|alpha|beta|gamma|Omega|pm|text)(?![a-zA-Z])/g, '\\\\$1');
            return '$' + seg + '$';
        } else if (hasMath) {
            // 混合段(中文+公式) -> 用wrapNakedLatex inline处理
            return wrapNakedLatex(seg);
        }
        return seg;
    });
    var result = rendered.join('\\uff0c');
    return escapeHtmlButKeepLatex(result);
}

function renderDocContent(text) {
    // 先去除 markdown格式(粗体/斜体标记): **text** -> text, *text* -> text
    // 注意不要破坏已有的 $$ 定界符
    text = text.replace(/\\*\\*\\s*(\\$\\$[\\s\\S]*?\\$\\$)\\s*\\*\\*/g, '$1'); // **$$...$$** -> $$...$$
    text = text.replace(/\\*\\*([^*\\n]+?)\\*\\*/g, '$1'); // **text** -> text
    text = text.replace(/(?<!\\*)\\*([^*\\n]+?)\\*(?!\\*)/g, '$1'); // *text* -> text
    // 去除 markdown 列表前缀 (*, -, #) 以避免干扰
    text = text.replace(/^(\\s*)[*\\-]\\s+/gm, '$1');
    text = text.replace(/^#{1,6}\\s+/gm, '');
    // 为裸露LaTeX添加定界符
    text = wrapNakedLatex(text);
    // 然后HTML转义（保护LaTeX定界符内的内容）
    return escapeHtmlButKeepLatex(text);
}

function showDetail(pairName, itemIdx) {
    const curDisagreements = historyData[currentDateIdx].disagreements || {};
    const item = curDisagreements[pairName][itemIdx];
    const compName = pairName.replace('千问-', '');
    const modal = document.getElementById('modal-content');
    modal.innerHTML = `
        <button class="close-btn" onclick="closeModal()">关闭</button>
        <h3>第 ${item.idx} 题 · ${item.grade} · ${item.course}</h3>
        <h4 style="color:#94a3b8;margin-top:12px;font-size:13px;">题目 (Query)：</h4>
        ${renderQuery(item.query)}
        <div style="display:flex;gap:8px;margin:12px 0;flex-wrap:wrap;">
            <span class="answer-tag">千问答案: ${renderAnswerContent(item.qw_answer)}</span>
            <span class="answer-tag">${compName}答案: ${renderAnswerContent(item.comp_answer)}</span>
        </div>
        <h4 style="color:#94a3b8;margin-top:16px;font-size:13px;">千问解题：</h4>
        <div style="display:flex;gap:12px;align-items:flex-start;flex-wrap:wrap;">
            ${item.qw_image_url ? `<div style="flex:0 0 280px;min-width:200px;max-width:45%;"><span class="answer-tag" style="font-size:10px;">千问解题截图</span><br><img src="${item.qw_image_url}" style="width:100%;max-height:800px;border:1px solid #e5e7eb;border-radius:6px;margin-top:4px;cursor:pointer;object-fit:contain;" onclick="window.open('${item.qw_image_url}')" onerror="this.style.display='none'" loading="lazy"></div>` : ''}
            <div style="flex:1;min-width:250px;" class="doc-content">${renderDocContent(item.qw_doc)}</div>
        </div>
        <h4 style="color:#94a3b8;margin-top:16px;font-size:13px;">${compName}解题：</h4>
        <div style="display:flex;gap:12px;align-items:flex-start;flex-wrap:wrap;">
            ${item.comp_image_url ? `<div style="flex:0 0 280px;min-width:200px;max-width:45%;"><span class="answer-tag" style="font-size:10px;">${compName}解题截图</span><br><img src="${item.comp_image_url}" style="width:100%;max-height:800px;border:1px solid #e5e7eb;border-radius:6px;margin-top:4px;cursor:pointer;object-fit:contain;" onclick="window.open('${item.comp_image_url}')" onerror="this.style.display='none'" loading="lazy"></div>` : ''}
            <div style="flex:1;min-width:250px;" class="doc-content">${renderDocContent(item.comp_doc)}</div>
        </div>
    `;
    // 渲染LaTeX公式
    doKaTeX(modal);
    document.getElementById('modal-overlay').classList.add('show');
    document.body.style.overflow = 'hidden';
}

function closeModal() {
    document.getElementById('modal-overlay').classList.remove('show');
    document.body.style.overflow = '';
}

// 键盘ESC关闭弹窗
document.addEventListener('keydown', e => { if(e.key==='Escape') closeModal(); });

// 初始化 (每步独立try-catch，避免CDN加载失败导致全页空白)
var initErrors = [];
try { renderNav(); } catch(e) { initErrors.push('renderNav: '+e.message); console.error('renderNav failed:', e); }
try { renderOverview(); } catch(e) { initErrors.push('renderOverview: '+e.message); console.error('renderOverview failed:', e); }
try { renderTrendChart(); } catch(e) { initErrors.push('renderTrendChart: '+e.message); console.error('renderTrendChart failed:', e); }
try { renderGradeTable(); } catch(e) { initErrors.push('renderGradeTable: '+e.message); console.error('renderGradeTable failed:', e); }
try { renderCourseTable(); } catch(e) { initErrors.push('renderCourseTable: '+e.message); console.error('renderCourseTable failed:', e); }
try { renderDetails(); } catch(e) { initErrors.push('renderDetails: '+e.message); console.error('renderDetails failed:', e); }
try { renderUncomparables(); } catch(e) { initErrors.push('renderUncomparables: '+e.message); console.error('renderUncomparables failed:', e); }
if (initErrors.length > 0) {
    var errDiv = document.createElement('div');
    errDiv.style.cssText = 'position:fixed;top:10px;right:10px;background:#dc2626;color:white;padding:12px 16px;border-radius:8px;z-index:99999;font-size:13px;max-width:400px;';
    errDiv.innerHTML = '<strong>初始化错误:</strong><br>' + initErrors.join('<br>');
    document.body.appendChild(errDiv);
}
</script>
</div><!-- .page-wrap -->
</body>
</html>'''

# 注入数据到HTML (escape </script> to prevent tag truncation)
history_json = json.dumps(history, ensure_ascii=False).replace('</script>', '<\\/script>')
html_content = html_content.replace('HISTORY_DATA_PLACEHOLDER', history_json)
html_content = html_content.replace('PAIR_NAMES_PLACEHOLDER', json.dumps([p[1] for p in COMPARISON_PAIRS], ensure_ascii=False))
html_content = html_content.replace('BUILDTIME_PLACEHOLDER', datetime.datetime.now().strftime('%Y-%m-%d %H:%M'))

# 内联KaTeX JS（避免CDN加载失败）
katex_js_path = '/tmp/katex.min.js'
autorender_js_path = '/tmp/auto-render.min.js'
if os.path.exists(katex_js_path):
    with open(katex_js_path, 'r') as f:
        katex_js = f.read()
    html_content = html_content.replace('KATEX_INLINE_PLACEHOLDER', katex_js)
else:
    # fallback: use CDN
    html_content = html_content.replace('<script>KATEX_INLINE_PLACEHOLDER</script>',
        '<script src="https://cdn.bootcdn.net/ajax/libs/KaTeX/0.16.9/katex.min.js"></script>')

if os.path.exists(autorender_js_path):
    with open(autorender_js_path, 'r') as f:
        autorender_js = f.read()
    html_content = html_content.replace('KATEX_AUTORENDER_INLINE_PLACEHOLDER', autorender_js)
else:
    html_content = html_content.replace('<script>KATEX_AUTORENDER_INLINE_PLACEHOLDER</script>',
        '<script src="https://cdn.bootcdn.net/ajax/libs/KaTeX/0.16.9/contrib/auto-render.min.js"></script>')

with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"HTML报告已保存: {OUTPUT_HTML}")
print(f"\n=== 完成 ===")
