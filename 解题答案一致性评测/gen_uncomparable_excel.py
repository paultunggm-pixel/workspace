#!/usr/bin/env python3
"""从 history.json 的 uncomparables 字段生成无法判断明细 Excel"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
HISTORY_FILE = os.path.join(SCRIPT_DIR, 'outputs', 'history.json')
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'outputs')

with open(HISTORY_FILE, 'r') as f:
    history = json.load(f)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

for entry in history:
    date = entry['date']
    uncomps = entry.get('uncomparables', {})

    wb = Workbook()
    ws = wb.active
    ws.title = '无法判断明细'

    headers = ['序号', '对比模型对', '原因', 'query', '学段', '学科', '千问答案', '对方答案']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    for pair_name, items in uncomps.items():
        for item in items:
            vals = [
                item.get('idx', ''),
                pair_name,
                item.get('reason', ''),
                item.get('query', ''),
                item.get('grade', ''),
                item.get('course', ''),
                (item.get('qw_answer', '') or '(无内容)')[:200],
                (item.get('comp_answer', '') or '(无内容)')[:200],
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
            row += 1

    # 列宽: 序号, 对比模型对, 原因, query, 学段, 学科, 千问答案, 对方答案
    widths = [6, 14, 22, 50, 8, 10, 40, 40]
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    for letter, w in zip(col_letters, widths):
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = 'A2'

    out_path = os.path.join(OUTPUT_DIR, f'无法判断明细-{date}.xlsx')
    wb.save(out_path)
    print(f'已生成: {out_path}')

print('全部完成!')
